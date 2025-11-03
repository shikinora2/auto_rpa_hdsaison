# Tên tệp: logic_convert_pdf.py
import os
import re
import json
import base64
import logging
from datetime import datetime
from dateutil.relativedelta import relativedelta
import fitz  # PyMuPDF

# Cấu hình logging cơ bản
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- HÀM HELPER ---

def find_value(regex_pattern: str, text: str) -> str | None:
    """
    Tìm kiếm một giá trị bằng regex, tương tự hàm findValue trong JS.
    Sử dụng re.DOTALL để '.' có thể khớp với ký tự xuống dòng.
    Sử dụng re.IGNORECASE để không phân biệt hoa thường.
    """
    try:
        match = re.search(regex_pattern, text, re.IGNORECASE | re.DOTALL)
        if match and match.group(1):
            return match.group(1).replace('VNĐ', '').strip()
    except Exception as e:
        logging.warning(f"Lỗi regex với pattern '{regex_pattern}': {e}")
    return None

def calculate_end_date(disbursement_date_str: str | None, loan_term_str: str | None) -> str | None:
    """
    Tính toán ngày kết thúc hợp đồng nếu không tìm thấy.
    Tương tự logic trong JS.
    """
    if not disbursement_date_str or not loan_term_str:
        return None
    
    try:
        # Tìm số tháng (ví dụ: "6 tháng" -> 6)
        months_match = re.search(r'(\d+)', loan_term_str)
        if not months_match:
            return None
            
        months = int(months_match.group(1))
        
        # Chuyển đổi ngày giải ngân
        start_date = datetime.strptime(disbursement_date_str, '%d/%m/%Y')
        
        # Cộng số tháng
        end_date = start_date + relativedelta(months=months)
        
        return end_date.strftime('%d/%m/%Y')
    except Exception as e:
        logging.warning(f"Không thể tính ngày kết thúc từ '{disbursement_date_str}' và '{loan_term_str}': {e}")
        return None

# --- HÀM TRÍCH XUẤT CHÍNH ---

def extract_data_from_text(full_text: str, payment_page_text: str) -> dict | None:
    """
    Trích xuất dữ liệu từ văn bản đã được bóc tách.
    Đây là logic được chuyển đổi trực tiếp từ 'extractDataForExportTab' trong JS.
    """
    if not full_text or not payment_page_text:
        logging.error("Thiếu văn bản đầy đủ hoặc văn bản trang thanh toán.")
        return None

    # 1. Số hợp đồng (ưu tiên từ trang thanh toán)
    so_hop_dong_match = re.search(r'Số Hợp Đồng:\s*([A-Z0-9]+)', payment_page_text, re.IGNORECASE)
    so_hop_dong = so_hop_dong_match.group(1).strip() if so_hop_dong_match else find_value(r'Số:\s*([A-Z0-9]+)', full_text)

    if not so_hop_dong:
        logging.error("Không tìm thấy Số Hợp Đồng. Dừng trích xuất.")
        return None

    # 2. Họ tên
    ho_ten_raw = find_value(r'1\.1\.\s*Họ tên:\s*(.*?)\s*1\.2\.', full_text)
    ho_ten = ho_ten_raw.title() if ho_ten_raw else None # .title() tương đương toTitleCase

    # 3. Ngày sinh
    ngay_sinh = find_value(r'1\.2\.\s*Ngày sinh:\s*([0-9\/]+)', full_text) or \
                find_value(r'Ngày sinh:\s*([0-9\/]+)', full_text)

    # 4. Số điện thoại
    sdt = find_value(r'1\.7\.\s*Điện thoại di động:\s*(\d+)', full_text) or \
          find_value(r'Điện thoại di động:\s*(\d+)', full_text)

    # 5. Số CCCD (Chuỗi regex phức tạp từ JS)
    so_cccd = (
        find_value(r'1\.4\.\s*Số CCCD\/Thẻ căn cước\/Giấy tờ khác:\s*([0-9]+)', full_text) or
        find_value(r'1\.4\.\s*Số CCCD:\s*([0-9]+)', full_text) or
        find_value(r'1\.3\.\s*CMND\/CCCD:\s*([0-9]+)', full_text) or
        find_value(r'CMND\/CCCD:\s*([0-S0-9]+)', full_text) or
        find_value(r'Số CMND\/CCCD:\s*([0-9]+)', full_text) or
        # Mẫu cũ
        find_value(r'1\.4\.\s*Số CMND\/Thẻ CCCD\/Hộ chiếu\/Giấy tờ khác:\s*([0-9]+)', full_text) or
        find_value(r'Số CMND\/Thẻ CCCD\/Hộ chiếu\/Giấy tờ khác:\s*([0-9]+)', full_text) or
        find_value(r'Số CMND\/Thẻ CCCD:\s*([0-9]+)', full_text) or
        find_value(r'CMND\/Thẻ CCCD:\s*([0-9]+)', full_text)
    )

    # 6. Ngày kết thúc (và các trường phụ trợ)
    ngay_ket_thuc_raw = find_value(r'5\.8\.\s*Ngày Thanh Toán Cuối Cùng:\s*([0-9\/]+)', full_text) or \
                        find_value(r'Ngày Thanh Toán Cuối Cùng:\s*([0-9\/]+)', full_text)
    
    ngay_giai_ngan = find_value(r'2\.6\.\s*Ngày giải ngân dự kiến:\s*([0-9\/]+)', full_text) or \
                     find_value(r'Ngày giải ngân:\s*([0-9\/]+)', full_text)
                     
    thoi_han_vay = find_value(r'2\.5\. Thời Hạn Vay:\s*(.*?)\s*3\.', full_text)

    ngay_ket_thuc = ngay_ket_thuc_raw or calculate_end_date(ngay_giai_ngan, thoi_han_vay)

    return {
        "soHopDong": so_hop_dong,
        "hoTen": ho_ten,
        "ngaySinh": ngay_sinh,
        "sdt": sdt,
        "soCCCD": so_cccd,
        "ngayKetThuc": ngay_ket_thuc
    }

# --- HÀM XỬ LÝ PDF ---

def process_pdf_bytes(pdf_bytes: bytes) -> dict | None:
    """
    Nhận vào nội dung (bytes) của một file PDF, bóc tách văn bản
    và gọi hàm extract_data_from_text.
    """
    full_text = ""
    payment_page_text = ""
    
    try:
        # Mở PDF từ bytes
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for page in doc:
                page_text = page.get_text()
                full_text += page_text + "\n"
                
                # Tìm trang thanh toán (theo logic của 'processSinglePdfForExtraction')
                if ("Số Hợp Đồng:" in page_text and "Khoản Thanh Toán Hàng Tháng:" in page_text) or \
                   ("HƯỚNG DẪN THANH TOÁN" in page_text): # Thêm fallback từ logic 'processSinglePdf'
                    if not payment_page_text: # Chỉ lấy trang đầu tiên tìm thấy
                        payment_page_text = page_text
        
        if not payment_page_text:
            logging.warning("Không tìm thấy trang thanh toán chứa 'Số Hợp Đồng'.")
            # Nếu không tìm thấy, vẫn thử trích xuất với full_text
            # vì 'soHopDong' có thể được tìm thấy ở chỗ khác
            payment_page_text = full_text 
            
        return extract_data_from_text(full_text, payment_page_text)
        
    except Exception as e:
        logging.error(f"Lỗi khi xử lý file PDF bytes: {e}", exc_info=True)
        return None

# --- HÀM CHÍNH (PUBLIC) ---

def process_file(file_path: str) -> dict | None:
    """
    Hàm chính để xử lý một file.
    Tự động phát hiện file .pdf hoặc .json (chứa base64).
    """
    file_name = os.path.basename(file_path)
    logging.info(f"Đang xử lý file: {file_name}")
    pdf_bytes = None
    
    try:
        if file_path.lower().endswith(".json"):
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if 'pdf_base64' not in data:
                logging.error(f"File JSON {file_name} không có key 'pdf_base64'.")
                return None
                
            pdf_bytes = base64.b64decode(data['pdf_base64'])
            logging.info("Đã giải mã Base64 từ file JSON.")
            
        elif file_path.lower().endswith(".pdf"):
            with open(file_path, 'rb') as f:
                pdf_bytes = f.read()
            logging.info("Đã đọc file PDF.")
            
        else:
            logging.warning(f"Bỏ qua file không hỗ trợ: {file_name}")
            return None
            
        # Sau khi có pdf_bytes, tiến hành xử lý
        if pdf_bytes:
            extracted_data = process_pdf_bytes(pdf_bytes)
            if extracted_data:
                logging.info(f"Trích xuất thành công: {extracted_data.get('soHopDong')}")
                return extracted_data
            else:
                logging.error(f"Không thể trích xuất dữ liệu từ {file_name}")
                return None
                
    except Exception as e:
        logging.error(f"Lỗi nghiêm trọng khi xử lý file {file_name}: {e}", exc_info=True)
        return None

def process_directory(directory_path: str) -> list:
    """
    Xử lý tất cả các file .pdf và .json trong một thư mục.
    """
    all_data = []
    logging.info(f"Quét thư mục: {directory_path}")
    
    for filename in os.listdir(directory_path):
        file_path = os.path.join(directory_path, filename)
        
        if os.path.isfile(file_path) and (filename.lower().endswith('.pdf') or filename.lower().endswith('.json')):
            extracted_data = process_file(file_path)
            if extracted_data:
                all_data.append(extracted_data)
        else:
            logging.debug(f"Bỏ qua file hoặc thư mục con: {filename}")
            
    return all_data

# === PHẦN BỔ SUNG ĐỂ XUẤT EXCEL ===
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    logging.error("Lỗi: Thư viện 'openpyxl' chưa được cài đặt.")
    logging.error("Vui lòng chạy: pip install openpyxl")


def export_data_to_excel(data_list: list, output_path: str):
    """
    Nhận vào một danh sách (list) các dictionary
    và xuất ra một file Excel (.xlsx).
    """
    if not data_list:
        logging.warning("Không có dữ liệu để xuất Excel.")
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách hợp đồng"

        # 1. Ghi Header
        headers = [
            'STT',
            'Số hợp đồng',
            'Họ tên',
            'Ngày sinh',
            'SĐT',
            'Số CCCD',
            'Ngày kết thúc'
        ]
        ws.append(headers)

        # 2. Ghi Dữ liệu
        for i, data in enumerate(data_list):
            row = [
                i + 1,
                data.get('soHopDong', ''),
                data.get('hoTen', ''),
                data.get('ngaySinh', ''),
                data.get('sdt', ''),
                data.get('soCCCD', ''),
                data.get('ngayKetThuc', '')
            ]
            ws.append(row)

        # 3. Tự động điều chỉnh độ rộng cột (tương tự logic JS)
        col_widths = [5, 15, 25, 12, 12, 15, 15]
        for i, width in enumerate(col_widths):
            try:
                # Cột bắt đầu từ 1 (A=1, B=2,...)
                ws.column_dimensions[get_column_letter(i + 1)].width = width
            except Exception as e:
                logging.warning(f"Lỗi khi set độ rộng cột {i+1}: {e}")

        # 4. Lưu file
        wb.save(output_path)
        logging.info(f"Đã lưu file Excel thành công vào: {output_path}")
    
    except ImportError:
        logging.error("Vui lòng cài đặt 'openpyxl' để sử dụng tính năng này (pip install openpyxl)")
        raise # Ném lỗi ra ngoài để UI có thể bắt
    except Exception as e:
        logging.error(f"Lỗi khi xuất file Excel: {e}", exc_info=True)
        raise # Ném lỗi ra ngoài để UI có thể bắt

# --- KHỐI ĐỂ CHẠY THỬ NGHIỆM ---

if __name__ == "__main__":
    """
    Khối này sẽ chạy khi bạn thực thi file này trực tiếp.
    Nó sẽ quét thư mục hiện tại, xử lý tất cả các file .pdf và .json,
    sau đó in kết quả ra và lưu vào file 'extracted_contracts.xlsx'.
    """
    logging.info("--- CHẠY TẬP LỆNH TRÍCH XUẤT HỢP ĐỒNG ---")
    
    # Thư mục để quét (mặc định là thư mục chứa file .py này)
    target_directory = "." 
    
    all_extracted_data = process_directory(target_directory)
    
    if all_extracted_data:
        print("\n" + "="*50)
        print(f"✅ TRÍCH XUẤT THÀNH CÔNG {len(all_extracted_data)} HỢP ĐỒNG")
        print("="*50)
        
        # In ra màn hình (dưới dạng JSON)
        print(json.dumps(all_extracted_data, indent=2, ensure_ascii=False))
        
        # Lưu vào file Excel
        output_filename = "extracted_contracts.xlsx"
        try:
            export_data_to_excel(all_extracted_data, output_filename)
        except Exception as e:
            logging.error(f"Không thể lưu file output Excel: {e}")
            
    else:
        logging.warning("Không tìm thấy hoặc không xử lý được file .pdf/.json nào trong thư mục.")
