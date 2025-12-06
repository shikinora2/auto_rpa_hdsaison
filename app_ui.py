import customtkinter
from customtkinter import filedialog
from tkinter import messagebox
import threading
import rpa_logic  # Logic RPA (Tải file & Cào chi tiết)
import logic_convert_pdf # Logic (Trích xuất file local)
from datetime import date
import os
import json
import base64
import zalo_logic

customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

class InputDialog(customtkinter.CTkToplevel):
    """Custom input dialog với customtkinter"""
    def __init__(self, parent, title, prompt):
        super().__init__(parent)
        
        self.result = None
        
        # Cấu hình window
        self.title(title)
        self.geometry("400x180")
        self.resizable(False, False)
        
        # Center window
        self.transient(parent)
        self.grab_set()
        
        # Icon và prompt
        icon_frame = customtkinter.CTkFrame(self, fg_color="transparent")
        icon_frame.pack(pady=(20, 10), padx=20, fill="x")
        
        icon_label = customtkinter.CTkLabel(
            icon_frame,
            text="✏️",
            font=customtkinter.CTkFont(size=24)
        )
        icon_label.pack(side="left", padx=(0, 10))
        
        prompt_label = customtkinter.CTkLabel(
            icon_frame,
            text=prompt,
            font=customtkinter.CTkFont(size=13)
        )
        prompt_label.pack(side="left", fill="x", expand=True)
        
        # Entry
        self.entry = customtkinter.CTkEntry(
            self,
            height=35,
            font=customtkinter.CTkFont(size=13),
            placeholder_text="Nhập tên kịch bản..."
        )
        self.entry.pack(pady=10, padx=20, fill="x")
        self.entry.focus()
        
        # Buttons
        button_frame = customtkinter.CTkFrame(self, fg_color="transparent")
        button_frame.pack(pady=10, padx=20, fill="x")
        
        cancel_btn = customtkinter.CTkButton(
            button_frame,
            text="Cancel",
            command=self.on_cancel,
            fg_color="gray",
            hover_color="#5A6268",
            width=100,
            height=32
        )
        cancel_btn.pack(side="right", padx=(5, 0))
        
        ok_btn = customtkinter.CTkButton(
            button_frame,
            text="OK",
            command=self.on_ok,
            width=100,
            height=32
        )
        ok_btn.pack(side="right")
        
        # Bind Enter key
        self.entry.bind("<Return>", lambda e: self.on_ok())
        self.entry.bind("<Escape>", lambda e: self.on_cancel())
        
        # Wait for window
        self.wait_window()
    
    def on_ok(self):
        self.result = self.entry.get()
        self.destroy()
    
    def on_cancel(self):
        self.result = None
        self.destroy()
    
    def get_input(self):
        return self.result

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Tool Automation - v1.3.0")
        self.geometry("850x700")
        
        # Print version để debug
        print("="*60)
        print("HD SAISON RPA Tool v1.3.0")
        print("Headless mode variable: headless_mode_var")
        print("="*60)

        # === CẤU HÌNH GRID CHO CỬA SỔ CHÍNH ===
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # === CẤU HÌNH GRID CHO CỬA SỔ CHÍNH ===
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Biến kiểm soát luồng
        self.rpa_thread = None
        self.pause_event = threading.Event()
        self.stop_event = threading.Event()

        # Thư mục lưu trữ dữ liệu ứng dụng
        self.app_data_dir = "app_data"
        if not os.path.exists(self.app_data_dir):
            os.makedirs(self.app_data_dir)

        self.config_file = os.path.join(self.app_data_dir, "config.json")

        # === TẠO TABVIEW ===
        self.tabview = customtkinter.CTkTabview(self, width=600, height=650)
        self.tabview.grid(row=0, column=0, padx=15, pady=15, sticky="nsew")

        # Tạo 5 tabs
        self.tabview.add("Trang Chủ")
        self.tabview.add("Tác Vụ")
        self.tabview.add("Auto Zalo")
        self.tabview.add("Kiểm Tra Hợp Đồng")
        self.tabview.add("Gemini & Sheet")

        # Cấu hình grid cho từng tab (Trang Chủ sẽ được config trong create_home_tab)
        self.tabview.tab("Tác Vụ").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Tác Vụ").grid_rowconfigure(3, weight=1)

        self.tabview.tab("Auto Zalo").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Auto Zalo").grid_rowconfigure(1, weight=1)

        self.tabview.tab("Kiểm Tra Hợp Đồng").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Kiểm Tra Hợp Đồng").grid_rowconfigure(0, weight=1)

        self.tabview.tab("Gemini & Sheet").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Gemini & Sheet").grid_rowconfigure(1, weight=1)

        # Biến lưu dữ liệu khách hàng cho Zalo
        self.zalo_customer_data = []  # List of dict: [{name, phone, address, ...}, ...]
        self.zalo_excel_path = None

        # Biến quản lý tài khoản Zalo
        self.current_account_id = None  # ID tài khoản đang chọn
        self.account_manager = zalo_logic.ZaloAccountManager()  # Khởi tạo ngay

        # Biến quản lý trạng thái tạm dừng
        self.is_paused = False

        # === TAB 1: TRANG CHỦ ===
        self.create_home_tab()

        # === TAB 2: TÁC VỤ ===
        self.create_tasks_tab()

        # === TAB 3: AUTO ZALO ===
        self.create_zalo_tab()

        # === TAB 4: KIỂM TRA HỢP ĐỒNG ===
        self.create_contract_check_tab()

        # === TAB 5: GEMINI & SHEET ===
        self.create_gemini_sheet_tab()

        self.load_config()
        self.load_zalo_session_info()  # Load thông tin session Zalo

    def create_home_tab(self):
        """Tạo nội dung cho tab Trang Chủ"""
        home_tab = self.tabview.tab("Trang Chủ")
        
        # Cấu hình grid cho home_tab: 1 cột, 2 hàng
        home_tab.grid_columnconfigure(0, weight=1)
        home_tab.grid_rowconfigure(0, weight=0)  # Hàng trên (đăng nhập) - auto height
        home_tab.grid_rowconfigure(1, weight=1)  # Hàng dưới (log) - chiếm hết không gian còn lại
        
        # === TRÊN: ĐĂNG NHẬP + TÍNH NĂNG ===
        top_frame = customtkinter.CTkFrame(home_tab)
        top_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        # Đăng nhập
        self.login_title = customtkinter.CTkLabel(
            top_frame,
            text="ĐĂNG NHẬP HỆ THỐNG",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        self.login_title.pack(pady=(10, 8), padx=10)

        # Frame chứa các input
        login_inputs = customtkinter.CTkFrame(top_frame, fg_color="transparent")
        login_inputs.pack(pady=(0, 5), padx=15, fill="x")

        self.username_label = customtkinter.CTkLabel(
            login_inputs,
            text="Tên đăng nhập:",
            font=customtkinter.CTkFont(size=12)
        )
        self.username_label.pack(pady=(3, 2), anchor="w")
        self.username_entry = customtkinter.CTkEntry(
            login_inputs,
            placeholder_text="Nhập tên đăng nhập",
            height=30
        )
        self.username_entry.pack(pady=2, fill="x")

        self.password_label = customtkinter.CTkLabel(
            login_inputs,
            text="Mật khẩu:",
            font=customtkinter.CTkFont(size=12)
        )
        self.password_label.pack(pady=(6, 2), anchor="w")
        self.password_entry = customtkinter.CTkEntry(
            login_inputs,
            placeholder_text="Nhập mật khẩu",
            show="*",
            height=30
        )
        self.password_entry.pack(pady=2, fill="x")

        # Checkboxes
        self.show_password_check = customtkinter.CTkCheckBox(
            login_inputs,
            text="Hiện mật khẩu",
            command=self.toggle_password_visibility,
            font=customtkinter.CTkFont(size=11)
        )
        self.show_password_check.pack(pady=(6, 2), anchor="w")

        self.save_creds_check = customtkinter.CTkCheckBox(
            login_inputs,
            text="Lưu thông tin đăng nhập",
            font=customtkinter.CTkFont(size=11)
        )
        self.save_creds_check.pack(pady=2, anchor="w")

        # Divider
        separator1 = customtkinter.CTkFrame(top_frame, height=2, fg_color="gray30")
        separator1.pack(fill="x", padx=15, pady=8)

        # Tính năng
        features_label = customtkinter.CTkLabel(
            top_frame,
            text="TRẠNG THÁI HỆ THỐNG",
            font=customtkinter.CTkFont(weight="bold", size=13)
        )
        features_label.pack(pady=(5, 8), padx=10)

        features_frame = customtkinter.CTkFrame(top_frame, fg_color="transparent")
        features_frame.pack(pady=(0, 10), padx=15, fill="x")

        # Checkbox chế độ headless
        self.headless_mode_var = customtkinter.BooleanVar(value=False)
        self.headless_checkbox = customtkinter.CTkCheckBox(
            features_frame,
            text="⚙️ Chạy ngầm",
            variable=self.headless_mode_var,
            font=customtkinter.CTkFont(size=11)
        )
        self.headless_checkbox.pack(pady=5, anchor="w")
        
        # Spacer để đảm bảo nội dung không bị che
        spacer = customtkinter.CTkFrame(top_frame, height=20, fg_color="transparent")
        spacer.pack(fill="x")
        
        # === DƯỚI: LOG TRẠNG THÁI ===
        log_container = customtkinter.CTkFrame(home_tab)
        log_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        
        log_title = customtkinter.CTkLabel(
            log_container,
            text="📋 NHẬT KÝ HỆ THỐNG",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        log_title.pack(pady=(8, 5), padx=10, anchor="w")

        self.log_textbox = customtkinter.CTkTextbox(
            log_container,
            state="disabled",
            wrap="word",
            font=customtkinter.CTkFont(size=11)
        )
        self.log_textbox.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def create_tasks_tab(self):
        """Tạo nội dung cho tab Tác Vụ (Các nút điều khiển)"""
        tasks_tab = self.tabview.tab("Tác Vụ")
        
        # === BỘ LỌC NGÀY ===
        self.date_frame = customtkinter.CTkFrame(tasks_tab)
        self.date_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        self.date_title = customtkinter.CTkLabel(
            self.date_frame,
            text="BỘ LỌC NGÀY",
            font=customtkinter.CTkFont(weight="bold", size=15)
        )
        self.date_title.pack(pady=(10, 5))

        # Grid container cho date picker
        date_grid = customtkinter.CTkFrame(self.date_frame, fg_color="transparent")
        date_grid.pack(pady=(5, 10), padx=15, fill="x")
        date_grid.grid_columnconfigure((1, 2, 3), weight=1)

        today = date.today()
        current_year = today.year
        self.days = [f"{i:02d}" for i in range(1, 32)]
        self.months = [f"{i:02d}" for i in range(1, 13)]
        self.years = [str(y) for y in range(current_year - 5, current_year + 2)]

        # Từ ngày
        self.start_date_label = customtkinter.CTkLabel(
            date_grid,
            text="Từ:",
            font=customtkinter.CTkFont(size=13, weight="bold")
        )
        self.start_date_label.grid(row=0, column=0, padx=6, pady=6, sticky="w")

        self.start_day_combo = customtkinter.CTkComboBox(
            date_grid,
            values=self.days,
            width=60,
            height=30,
            font=customtkinter.CTkFont(size=13)
        )
        self.start_day_combo.grid(row=0, column=1, padx=3, pady=6, sticky="ew")
        self.start_day_combo.set(today.strftime("%d"))

        self.start_month_combo = customtkinter.CTkComboBox(
            date_grid,
            values=self.months,
            width=60,
            height=30,
            font=customtkinter.CTkFont(size=13)
        )
        self.start_month_combo.grid(row=0, column=2, padx=3, pady=6, sticky="ew")
        self.start_month_combo.set(today.strftime("%m"))

        self.start_year_combo = customtkinter.CTkComboBox(
            date_grid,
            values=self.years,
            width=80,
            height=30,
            font=customtkinter.CTkFont(size=13)
        )
        self.start_year_combo.grid(row=0, column=3, padx=3, pady=6, sticky="ew")
        self.start_year_combo.set(str(current_year))

        # Đến ngày
        self.end_date_label = customtkinter.CTkLabel(
            date_grid,
            text="Đến:",
            font=customtkinter.CTkFont(size=13, weight="bold")
        )
        self.end_date_label.grid(row=1, column=0, padx=6, pady=6, sticky="w")

        self.end_day_combo = customtkinter.CTkComboBox(
            date_grid,
            values=self.days,
            width=60,
            height=30,
            font=customtkinter.CTkFont(size=13)
        )
        self.end_day_combo.grid(row=1, column=1, padx=3, pady=6, sticky="ew")
        self.end_day_combo.set(today.strftime("%d"))

        self.end_month_combo = customtkinter.CTkComboBox(
            date_grid,
            values=self.months,
            width=60,
            height=30,
            font=customtkinter.CTkFont(size=13)
        )
        self.end_month_combo.grid(row=1, column=2, padx=3, pady=6, sticky="ew")
        self.end_month_combo.set(today.strftime("%m"))

        self.end_year_combo = customtkinter.CTkComboBox(
            date_grid,
            values=self.years,
            width=80,
            height=30,
            font=customtkinter.CTkFont(size=13)
        )
        self.end_year_combo.grid(row=1, column=3, padx=3, pady=6, sticky="ew")
        self.end_year_combo.set(str(current_year))

        # === CHỌN THƯ MỤC VÀ HÌNH THỨC ===
        settings_frame = customtkinter.CTkFrame(tasks_tab)
        settings_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        
        # Chọn thư mục
        folder_container = customtkinter.CTkFrame(settings_frame, fg_color="transparent")
        folder_container.pack(pady=8, padx=12, fill="x")
        
        self.folder_label = customtkinter.CTkLabel(
            folder_container,
            text="Thư mục lưu:",
            font=customtkinter.CTkFont(size=13)
        )
        self.folder_label.pack(side="left", padx=(0, 5))

        self.folder_entry = customtkinter.CTkEntry(
            folder_container,
            state="disabled",
            height=30,
            font=customtkinter.CTkFont(size=12)
        )
        self.folder_entry.pack(side="left", fill="x", expand=True, padx=3)

        self.folder_button = customtkinter.CTkButton(
            folder_container,
            text="Chọn",
            command=self.select_folder,
            width=70,
            height=30,
            font=customtkinter.CTkFont(size=12)
        )
        self.folder_button.pack(side="left", padx=(3, 0))

        default_save_path = os.path.abspath("downloads_contracts")
        self.folder_entry.configure(state="normal")
        self.folder_entry.insert(0, default_save_path)
        self.folder_entry.configure(state="disabled")

        # Hình thức lưu
        format_container = customtkinter.CTkFrame(settings_frame, fg_color="transparent")
        format_container.pack(pady=(0, 8), padx=12, fill="x")

        self.save_format_label = customtkinter.CTkLabel(
            format_container,
            text="Định dạng lưu:",
            font=customtkinter.CTkFont(size=13)
        )
        self.save_format_label.pack(side="left", padx=(0, 8))

        self.save_format_button = customtkinter.CTkSegmentedButton(
            format_container,
            values=["PDF", "JSON"],
            command=self.on_save_format_change,
            font=customtkinter.CTkFont(size=12),
            height=30
        )
        self.save_format_button.pack(side="left", fill="x", expand=True)
        self.save_format_button.set("PDF")

        # === TÁC VỤ ONLINE (RPA) ===
        self.rpa_frame = customtkinter.CTkFrame(tasks_tab)
        self.rpa_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 10))
        
        self.rpa_label = customtkinter.CTkLabel(
            self.rpa_frame,
            text="TÁC VỤ TỰ ĐỘNG HÓA (ONLINE)",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        self.rpa_label.pack(pady=(10, 8))

        # Grid buttons
        rpa_buttons = customtkinter.CTkFrame(self.rpa_frame, fg_color="transparent")
        rpa_buttons.pack(fill="x", padx=10, pady=(0, 10))
        rpa_buttons.grid_columnconfigure((0, 1), weight=1)

        # Hàng 1
        self.check_button = customtkinter.CTkButton(
            rpa_buttons,
            text="Kiểm tra số lượng",
            command=self.start_check_thread,
            fg_color="#00695C",
            hover_color="#004D40",
            height=36,
            font=customtkinter.CTkFont(size=13)
        )
        self.check_button.grid(row=0, column=0, padx=4, pady=4, sticky="ew")

        self.start_button = customtkinter.CTkButton(
            rpa_buttons,
            text="Tải File (PDF)",
            command=self.start_rpa_thread,
            height=36,
            font=customtkinter.CTkFont(size=13)
        )
        self.start_button.grid(row=0, column=1, padx=4, pady=4, sticky="ew")

        # Hàng 2
        self.scrape_details_button = customtkinter.CTkButton(
            rpa_buttons,
            text="Lấy Chi Tiết (Excel)",
            command=self.start_detail_scrape_thread,
            fg_color="#004D40",
            hover_color="#00695C",
            height=36,
            font=customtkinter.CTkFont(size=13)
        )
        self.scrape_details_button.grid(row=1, column=0, padx=4, pady=4, sticky="ew")

        self.open_excel_button = customtkinter.CTkButton(
            rpa_buttons,
            text="Mở Thư Mục Excel",
            command=self.open_excel_folder,
            fg_color="#1565C0",
            hover_color="#0D47A1",
            height=36,
            font=customtkinter.CTkFont(size=13)
        )
        self.open_excel_button.grid(row=1, column=1, padx=4, pady=4, sticky="ew")

        # Hàng 3 - Điều khiển
        self.pause_button = customtkinter.CTkButton(
            rpa_buttons,
            text="Tạm Dừng",
            command=self.toggle_pause,
            state="disabled",
            fg_color="gray",
            text_color_disabled="white",
            height=32,
            font=customtkinter.CTkFont(size=12)
        )
        self.pause_button.grid(row=2, column=0, padx=4, pady=4, sticky="ew")

        self.stop_button = customtkinter.CTkButton(
            rpa_buttons,
            text="Kết Thúc (RPA)",
            command=self.stop_rpa,
            state="disabled",
            fg_color="#D32F2F",
            hover_color="#B71C1C",
            height=32,
            font=customtkinter.CTkFont(size=12)
        )
        self.stop_button.grid(row=2, column=1, padx=4, pady=4, sticky="ew")

        # === TÁC VỤ OFFLINE (LOCAL) ===
        self.local_frame = customtkinter.CTkFrame(tasks_tab)
        self.local_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))

        self.local_label = customtkinter.CTkLabel(
            self.local_frame,
            text="TÁC VỤ XỬ LÝ FILE (OFFLINE)",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        self.local_label.pack(pady=(10, 8))

        self.extract_button = customtkinter.CTkButton(
            self.local_frame,
            text="Trích xuất File (PDF/JSON) sang Excel",
            command=self.start_extraction_thread,
            fg_color="#4E342E",
            hover_color="#6D4C41",
            height=36,
            font=customtkinter.CTkFont(size=13)
        )
        self.extract_button.pack(fill="x", padx=10, pady=(0, 10))

    def create_zalo_tab(self):
        """Tạo nội dung cho tab Auto Zalo với ScrollableFrame"""
        zalo_tab = self.tabview.tab("Auto Zalo")

        # === TIÊU ĐỀ ===
        zalo_title = customtkinter.CTkLabel(
            zalo_tab,
            text="TỰ ĐỘNG HÓA ZALO",
            font=customtkinter.CTkFont(weight="bold", size=16)
        )
        zalo_title.grid(row=0, column=0, pady=(10, 5), sticky="ew", padx=10)

        # === SCROLLABLE FRAME ===
        scrollable_frame = customtkinter.CTkScrollableFrame(zalo_tab, fg_color="transparent")
        scrollable_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        scrollable_frame.grid_columnconfigure(0, weight=1)

        # === 1. QUẢN LÝ TÀI KHOẢN ZALO ===
        account_frame = customtkinter.CTkFrame(scrollable_frame)
        account_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        account_frame.grid_columnconfigure(0, weight=0)  # Cột nút (trái)
        account_frame.grid_columnconfigure(1, weight=1)  # Cột nhập liệu (phải)

        account_title = customtkinter.CTkLabel(
            account_frame,
            text="Quản Lý Tài Khoản Zalo",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        account_title.grid(row=0, column=0, columnspan=2, pady=(10, 8), padx=10, sticky="w")

        # === CỘT TRÁI: CÁC NÚT ===
        buttons_left_frame = customtkinter.CTkFrame(account_frame, fg_color="transparent")
        buttons_left_frame.grid(row=1, column=0, padx=(15, 10), pady=5, sticky="n")

        # Nút kiểm tra và cập nhật
        check_button = customtkinter.CTkButton(
            buttons_left_frame, text="Kiểm tra & Cập nhật",
            command=self.check_zalo_status, height=32, width=150,
            font=customtkinter.CTkFont(size=12)
        )
        check_button.pack(pady=3)

        # Nút mở Zalo
        open_zalo_btn = customtkinter.CTkButton(
            buttons_left_frame, text="Mở Zalo",
            command=self.open_zalo_window, height=32, width=150,
            fg_color="#0068FF", hover_color="#0052CC",
            font=customtkinter.CTkFont(size=12)
        )
        open_zalo_btn.pack(pady=3)

        # Frame chứa nút thêm/xóa tài khoản
        account_action_frame = customtkinter.CTkFrame(buttons_left_frame, fg_color="transparent")
        account_action_frame.pack(pady=3)

        # Nút thêm tài khoản
        add_account_btn = customtkinter.CTkButton(
            account_action_frame, text="+", width=70, height=28,
            command=self.add_zalo_account, font=customtkinter.CTkFont(size=16)
        )
        add_account_btn.pack(side="left", padx=2)

        # Nút xóa tài khoản
        delete_account_btn = customtkinter.CTkButton(
            account_action_frame, text="X", width=70, height=28,
            command=self.delete_zalo_account, fg_color="#DC3545", hover_color="#C82333",
            font=customtkinter.CTkFont(size=16)
        )
        delete_account_btn.pack(side="left", padx=2)

        # === CỘT PHẢI: CÁC Ô NHẬP LIỆU VÀ THÔNG TIN ===
        info_right_frame = customtkinter.CTkFrame(account_frame, fg_color="transparent")
        info_right_frame.grid(row=1, column=1, padx=(10, 15), pady=5, sticky="ew")
        info_right_frame.grid_columnconfigure(1, weight=1)

        # Tài khoản
        account_label = customtkinter.CTkLabel(
            info_right_frame, text="Tài khoản:", font=customtkinter.CTkFont(size=12)
        )
        account_label.grid(row=0, column=0, padx=(0, 10), pady=5, sticky="w")

        self.account_combobox = customtkinter.CTkComboBox(
            info_right_frame, values=["Chưa có tài khoản"],
            command=self.on_account_selected, font=customtkinter.CTkFont(size=12)
        )
        self.account_combobox.grid(row=0, column=1, pady=5, sticky="ew")
        self.account_combobox.set("Chưa có tài khoản")

        # Họ tên Zalo
        zalo_name_title = customtkinter.CTkLabel(
            info_right_frame, text="Họ tên:", font=customtkinter.CTkFont(size=12)
        )
        zalo_name_title.grid(row=1, column=0, padx=(0, 10), pady=5, sticky="w")

        self.zalo_name_label = customtkinter.CTkLabel(
            info_right_frame, text="Chưa cập nhật", font=customtkinter.CTkFont(size=12),
            text_color="gray"
        )
        self.zalo_name_label.grid(row=1, column=1, pady=5, sticky="w")

        # Phiên đăng nhập
        session_title = customtkinter.CTkLabel(
            info_right_frame, text="Phiên đăng nhập:", font=customtkinter.CTkFont(size=12)
        )
        session_title.grid(row=2, column=0, padx=(0, 10), pady=5, sticky="w")

        self.zalo_session_label = customtkinter.CTkLabel(
            info_right_frame, text="Chưa đăng nhập", font=customtkinter.CTkFont(size=12),
            text_color="gray"
        )
        self.zalo_session_label.grid(row=2, column=1, pady=5, sticky="w")

        # Trạng thái
        status_title = customtkinter.CTkLabel(
            info_right_frame, text="Trạng thái:", font=customtkinter.CTkFont(size=12)
        )
        status_title.grid(row=3, column=0, padx=(0, 10), pady=5, sticky="w")

        self.zalo_status_label = customtkinter.CTkLabel(
            info_right_frame, text="Inactive", font=customtkinter.CTkFont(size=12),
            text_color="red"
        )
        self.zalo_status_label.grid(row=3, column=1, pady=5, sticky="w")

        # === 2. NHẬP DỮ LIỆU KHÁCH HÀNG ===
        data_frame = customtkinter.CTkFrame(scrollable_frame)
        data_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        data_frame.grid_columnconfigure(0, weight=0)  # Cột nút (trái)
        data_frame.grid_columnconfigure(1, weight=1)  # Cột thông tin (phải)

        data_title = customtkinter.CTkLabel(
            data_frame,
            text="📁 Nhập Dữ Liệu Khách Hàng",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        data_title.grid(row=0, column=0, columnspan=2, pady=(10, 8), padx=10, sticky="w")

        # === CỘT TRÁI: CÁC NÚT ===
        data_buttons_frame = customtkinter.CTkFrame(data_frame, fg_color="transparent")
        data_buttons_frame.grid(row=1, column=0, padx=(15, 10), pady=(0, 15), sticky="n")

        # Nút chọn file Excel
        self.select_excel_button = customtkinter.CTkButton(
            data_buttons_frame,
            text="Chọn File Excel",
            command=self.select_zalo_excel,
            height=32, width=150,
            font=customtkinter.CTkFont(size=13)
        )
        self.select_excel_button.pack(pady=3)

        # Nút nhập từ Sheet
        self.import_from_sheet_button = customtkinter.CTkButton(
            data_buttons_frame,
            text="Nhập từ Sheet",
            command=self.import_zalo_from_sheet,
            height=32, width=150,
            fg_color="#0F9D58",
            hover_color="#0B8043",
            font=customtkinter.CTkFont(size=13)
        )
        self.import_from_sheet_button.pack(pady=3)

        # === CỘT PHẢI: THÔNG TIN FILE ===
        info_frame = customtkinter.CTkFrame(data_frame, fg_color="transparent")
        info_frame.grid(row=1, column=1, padx=(10, 15), pady=(0, 15), sticky="ew")

        # Hiển thị file đã chọn
        self.zalo_file_label = customtkinter.CTkLabel(
            info_frame,
            text="Chưa chọn nguồn dữ liệu",
            font=customtkinter.CTkFont(size=12),
            text_color="gray"
        )
        self.zalo_file_label.pack(anchor="w", pady=(0, 5))

        # Hiển thị số lượng khách hàng
        self.zalo_customer_count_label = customtkinter.CTkLabel(
            info_frame,
            text="Số khách hàng: 0",
            font=customtkinter.CTkFont(size=12, weight="bold"),
            text_color="#0068FF"
        )
        self.zalo_customer_count_label.pack(anchor="w")

        # === 3. KẾT BẠN HÀNG LOẠT ===
        friend_frame = customtkinter.CTkFrame(scrollable_frame)
        friend_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        friend_frame.grid_columnconfigure(0, weight=0)  # Cột nút (trái)
        friend_frame.grid_columnconfigure(1, weight=1)  # Cột nhập liệu (phải)

        friend_title = customtkinter.CTkLabel(
            friend_frame,
            text="👥 Kết Bạn Hàng Loạt",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        friend_title.grid(row=0, column=0, columnspan=2, pady=(10, 8), padx=10, sticky="w")

        # === CỘT TRÁI: CÁC NÚT ===
        friend_buttons_frame = customtkinter.CTkFrame(friend_frame, fg_color="transparent")
        friend_buttons_frame.grid(row=1, column=0, padx=(15, 10), pady=(0, 15), sticky="n")

        # Nút kết bạn
        self.add_friend_button = customtkinter.CTkButton(
            friend_buttons_frame,
            text="Kết Bạn Hàng Loạt",
            command=self.add_friends_bulk,
            fg_color="#FFC107",
            hover_color="#E0A800",
            text_color="black",
            height=36, width=150,
            font=customtkinter.CTkFont(size=13, weight="bold")
        )
        self.add_friend_button.pack(pady=3)

        # Nút tạm dừng/tiếp tục (Zalo)
        self.zalo_pause_button = customtkinter.CTkButton(
            friend_buttons_frame,
            text="Tạm dừng",
            command=self.toggle_pause,
            fg_color="#6C757D",
            hover_color="#5A6268",
            height=36, width=150,
            font=customtkinter.CTkFont(size=13, weight="bold"),
            state="disabled"  # Mặc định disabled
        )
        self.zalo_pause_button.pack(pady=3)

        # === CỘT PHẢI: Ô NHẬP LIỆU ===
        friend_input_frame = customtkinter.CTkFrame(friend_frame, fg_color="transparent")
        friend_input_frame.grid(row=1, column=1, padx=(10, 15), pady=(0, 15), sticky="nsew")

        # Hướng dẫn
        friend_help = customtkinter.CTkLabel(
            friend_input_frame,
            text="Gửi lời mời kết bạn đến tất cả số điện thoại trong danh sách",
            font=customtkinter.CTkFont(size=11),
            text_color="gray"
        )
        friend_help.pack(anchor="w", pady=(0, 5))

        # Hướng dẫn biến có sẵn cho lời chào
        greeting_help = customtkinter.CTkLabel(
            friend_input_frame,
            text="Biến có sẵn: {my_name}, {contract_id}, {name}, {phone}, {gender} (Nam→anh, Nữ→chị), {address}, {cccd}, {dob}",
            font=customtkinter.CTkFont(size=11),
            text_color="gray"
        )
        greeting_help.pack(anchor="w", pady=(0, 3))

        # Label cho ô nhập lời chào
        greeting_label = customtkinter.CTkLabel(
            friend_input_frame,
            text="Lời chào khi kết bạn:",
            font=customtkinter.CTkFont(size=12)
        )
        greeting_label.pack(anchor="w", pady=(0, 3))

        # TextBox nhập lời chào
        self.friend_greeting_textbox = customtkinter.CTkTextbox(
            friend_input_frame,
            height=80,
            font=customtkinter.CTkFont(size=12)
        )
        self.friend_greeting_textbox.pack(fill="both", expand=True, pady=(0, 10))

        # Lời chào mặc định
        default_greeting = "Xin chào, mình là {my_name} bên công ty tài chính HDSAISON, vui lòng đồng ý kết bạn để được hỗ trợ hợp đồng {contract_id}"
        self.friend_greeting_textbox.insert("1.0", default_greeting)

        # Checkbox bỏ qua khách hàng đã xử lý
        self.skip_processed_var = customtkinter.BooleanVar(value=True)  # Mặc định bật
        self.skip_processed_checkbox = customtkinter.CTkCheckBox(
            friend_input_frame,
            text="Bỏ qua khách hàng đã kết bạn thành công",
            variable=self.skip_processed_var,
            font=customtkinter.CTkFont(size=12),
            text_color="#28A745"
        )
        self.skip_processed_checkbox.pack(anchor="w")

        # === 4. NHẮN TIN HÀNG LOẠT ===
        message_frame = customtkinter.CTkFrame(scrollable_frame)
        message_frame.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        message_frame.grid_columnconfigure(0, weight=0)  # Cột nút (trái)
        message_frame.grid_columnconfigure(1, weight=1)  # Cột nhập liệu (phải)

        message_title = customtkinter.CTkLabel(
            message_frame,
            text="💬 Nhắn Tin Hàng Loạt",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        message_title.grid(row=0, column=0, columnspan=2, pady=(10, 8), padx=10, sticky="w")

        # === CỘT TRÁI: CÁC NÚT ===
        message_buttons_left = customtkinter.CTkFrame(message_frame, fg_color="transparent")
        message_buttons_left.grid(row=1, column=0, padx=(15, 10), pady=(0, 15), sticky="n")

        # Nút gửi tin nhắn
        self.send_message_button = customtkinter.CTkButton(
            message_buttons_left,
            text="Gửi Tin Nhắn Hàng Loạt",
            command=self.send_bulk_messages,
            fg_color="#28A745",
            hover_color="#218838",
            height=36, width=150,
            font=customtkinter.CTkFont(size=13, weight="bold")
        )
        self.send_message_button.pack(pady=3)

        # Nút tạm dừng/tiếp tục cho gửi tin nhắn
        self.message_pause_button = customtkinter.CTkButton(
            message_buttons_left,
            text="Tạm dừng",
            command=self.toggle_pause,
            fg_color="#6C757D",
            hover_color="#5A6268",
            height=36, width=150,
            font=customtkinter.CTkFont(size=13, weight="bold"),
            state="disabled"  # Mặc định disabled
        )
        self.message_pause_button.pack(pady=3)

        # Nút lưu kịch bản
        save_template_button = customtkinter.CTkButton(
            message_buttons_left,
            text="💾 Lưu kịch bản",
            command=self.save_message_template,
            fg_color="#17A2B8",
            hover_color="#138496",
            height=32, width=150,
            font=customtkinter.CTkFont(size=12)
        )
        save_template_button.pack(pady=3)

        # === CỘT PHẢI: Ô NHẬP LIỆU ===
        message_input_frame = customtkinter.CTkFrame(message_frame, fg_color="transparent")
        message_input_frame.grid(row=1, column=1, padx=(10, 15), pady=(0, 15), sticky="nsew")
        
        # Dropdown chọn kịch bản đã lưu
        template_selection_frame = customtkinter.CTkFrame(message_input_frame, fg_color="transparent")
        template_selection_frame.pack(fill="x", pady=(0, 8))
        
        template_select_label = customtkinter.CTkLabel(
            template_selection_frame,
            text="Kịch bản đã lưu:",
            font=customtkinter.CTkFont(size=12)
        )
        template_select_label.pack(side="left", padx=(0, 8))
        
        # Load danh sách kịch bản
        self.template_names = self.get_saved_template_names()
        template_options = self.template_names if self.template_names else ["(Chưa có kịch bản)"]
        
        self.template_dropdown = customtkinter.CTkComboBox(
            template_selection_frame,
            values=template_options,
            command=self.on_template_selected,
            width=200,
            font=customtkinter.CTkFont(size=12)
        )
        self.template_dropdown.pack(side="left", padx=(0, 8))
        
        if self.template_names:
            self.template_dropdown.set(self.template_names[0])
        
        # Nút refresh danh sách
        refresh_template_button = customtkinter.CTkButton(
            template_selection_frame,
            text="🔄",
            command=self.refresh_template_list,
            width=40,
            height=28,
            font=customtkinter.CTkFont(size=14)
        )
        refresh_template_button.pack(side="left", padx=(0, 5))
        
        # Nút xoá kịch bản
        delete_template_button = customtkinter.CTkButton(
            template_selection_frame,
            text="🗑️",
            command=self.delete_message_template,
            width=40,
            height=28,
            fg_color="#DC3545",
            hover_color="#C82333",
            font=customtkinter.CTkFont(size=14)
        )
        delete_template_button.pack(side="left")

        # Hướng dẫn sử dụng biến
        help_label = customtkinter.CTkLabel(
            message_input_frame,
            text="Biến có sẵn: {name}, {phone}, {address}, {cccd}, {dob}, {contract_id}, {gender} (Nam→anh, Nữ→chị)",
            font=customtkinter.CTkFont(size=11),
            text_color="gray"
        )
        help_label.pack(anchor="w", pady=(0, 5))

        # TextBox nhập kịch bản
        message_label = customtkinter.CTkLabel(
            message_input_frame,
            text="Kịch bản tin nhắn:",
            font=customtkinter.CTkFont(size=12)
        )
        message_label.pack(anchor="w", pady=(0, 3))

        self.zalo_message_template = customtkinter.CTkTextbox(
            message_input_frame,
            height=120,
            font=customtkinter.CTkFont(size=12)
        )
        self.zalo_message_template.pack(fill="both", expand=True, pady=(0, 10))

        # Load kịch bản: Nếu có kịch bản đã chọn thì load, không thì dùng default
        if self.template_names:
            saved_template = self.load_message_template(self.template_names[0])
            if saved_template:
                self.zalo_message_template.insert("1.0", saved_template)
            else:
                self._insert_default_template()
        else:
            self._insert_default_template()

        # Checkbox bỏ qua khách hàng đã gửi tin nhắn
        self.skip_sent_messages_var = customtkinter.BooleanVar(value=True)  # Mặc định bật
        self.skip_sent_messages_checkbox = customtkinter.CTkCheckBox(
            message_input_frame,
            text="Bỏ qua khách hàng đã gửi tin nhắn thành công",
            variable=self.skip_sent_messages_var,
            font=customtkinter.CTkFont(size=12),
            text_color="#28A745"
        )
        self.skip_sent_messages_checkbox.pack(anchor="w")

    def create_gemini_sheet_tab(self):
        """Tạo nội dung cho tab Gemini & Sheet"""
        gemini_sheet_tab = self.tabview.tab("Gemini & Sheet")

        # Placeholder - Sẽ thêm nội dung sau
        placeholder = customtkinter.CTkLabel(
            gemini_sheet_tab,
            text="Tab Gemini & Sheet\n(Sẽ thêm nội dung sau)",
            font=customtkinter.CTkFont(size=14)
        )
        placeholder.pack(expand=True)

    def create_contract_check_tab(self):
        """Tạo nội dung cho tab Kiểm Tra Hợp Đồng"""
        contract_tab = self.tabview.tab("Kiểm Tra Hợp Đồng")

        # Tiêu đề
        title = customtkinter.CTkLabel(
            contract_tab,
            text="🔍 Kiểm Tra Hợp Đồng",
            font=customtkinter.CTkFont(weight="bold", size=18)
        )
        title.pack(pady=20)

        # Frame chính
        main_frame = customtkinter.CTkFrame(contract_tab)
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # === NHẬP LIỆU THỦ CÔNG ===
        manual_frame = customtkinter.CTkFrame(main_frame)
        manual_frame.pack(fill="x", padx=15, pady=15)

        manual_title = customtkinter.CTkLabel(
            manual_frame,
            text="✍️ Nhập Thủ Công",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        manual_title.grid(row=0, column=0, columnspan=2, pady=(10, 15), padx=10, sticky="w")

        # Nhập số hợp đồng
        contract_label = customtkinter.CTkLabel(
            manual_frame,
            text="Số hợp đồng:",
            font=customtkinter.CTkFont(size=12)
        )
        contract_label.grid(row=1, column=0, padx=15, pady=(0, 10), sticky="w")

        self.contract_number_entry = customtkinter.CTkEntry(
            manual_frame,
            placeholder_text="Nhập số hợp đồng",
            font=customtkinter.CTkFont(size=12),
            height=35
        )
        self.contract_number_entry.grid(row=1, column=1, padx=15, pady=(0, 10), sticky="ew")

        # Nhập số CCCD
        cccd_label = customtkinter.CTkLabel(
            manual_frame,
            text="Số CCCD:",
            font=customtkinter.CTkFont(size=12)
        )
        cccd_label.grid(row=2, column=0, padx=15, pady=(0, 15), sticky="w")

        self.cccd_entry = customtkinter.CTkEntry(
            manual_frame,
            placeholder_text="Nhập số CCCD",
            font=customtkinter.CTkFont(size=12),
            height=35
        )
        self.cccd_entry.grid(row=2, column=1, padx=15, pady=(0, 15), sticky="ew")

        manual_frame.grid_columnconfigure(1, weight=1)

        # === NHẬP LIỆU TỪ FILE ===
        file_frame = customtkinter.CTkFrame(main_frame)
        file_frame.pack(fill="x", padx=15, pady=(0, 15))
        file_frame.grid_columnconfigure(0, weight=0)  # Cột nút (trái)
        file_frame.grid_columnconfigure(1, weight=1)  # Cột thông tin (phải)

        file_title = customtkinter.CTkLabel(
            file_frame,
            text="📄 Nhập Dữ Liệu",
            font=customtkinter.CTkFont(weight="bold", size=14)
        )
        file_title.grid(row=0, column=0, columnspan=2, pady=(10, 15), padx=10, sticky="w")

        # === CỘT TRÁI: CÁC NÚT ===
        contract_buttons_frame = customtkinter.CTkFrame(file_frame, fg_color="transparent")
        contract_buttons_frame.grid(row=1, column=0, padx=(15, 10), pady=(0, 15), sticky="n")

        # Nút chọn file Excel
        self.select_contract_file_button = customtkinter.CTkButton(
            contract_buttons_frame,
            text="Chọn File Excel",
            command=self.select_contract_file,
            fg_color="#007BFF",
            hover_color="#0056B3",
            height=36, width=150,
            font=customtkinter.CTkFont(size=13, weight="bold")
        )
        self.select_contract_file_button.pack(pady=3)

        # Nút nhập từ Sheet
        self.import_contract_from_sheet_button = customtkinter.CTkButton(
            contract_buttons_frame,
            text="Nhập từ Sheet",
            command=self.import_contract_from_sheet,
            fg_color="#0F9D58",
            hover_color="#0B8043",
            height=36, width=150,
            font=customtkinter.CTkFont(size=13, weight="bold")
        )
        self.import_contract_from_sheet_button.pack(pady=3)

        # === CỘT PHẢI: THÔNG TIN ===
        contract_info_frame = customtkinter.CTkFrame(file_frame, fg_color="transparent")
        contract_info_frame.grid(row=1, column=1, padx=(10, 15), pady=(0, 15), sticky="ew")

        # Hiển thị file đã chọn
        self.contract_file_label = customtkinter.CTkLabel(
            contract_info_frame,
            text="Chưa chọn nguồn dữ liệu",
            font=customtkinter.CTkFont(size=12),
            text_color="gray"
        )
        self.contract_file_label.pack(anchor="w", pady=(0, 5))

        # Số lượng hợp đồng
        self.contract_count_label = customtkinter.CTkLabel(
            contract_info_frame,
            text="Số lượng: 0",
            font=customtkinter.CTkFont(size=12, weight="bold"),
            text_color="#0068FF"
        )
        self.contract_count_label.pack(anchor="w")

        # === NÚT KIỂM TRA ===
        self.check_contract_button = customtkinter.CTkButton(
            main_frame,
            text="Kiểm Tra",
            command=self.check_contracts,
            fg_color="#28A745",
            hover_color="#218838",
            height=40,
            font=customtkinter.CTkFont(size=14, weight="bold")
        )
        self.check_contract_button.pack(fill="x", padx=15, pady=(0, 15))

        # Biến lưu dữ liệu
        self.contract_data = []  # List of dict: [{contract_id, cccd}, ...]
        self.contract_excel_path = None

    def on_save_format_change(self, value):
        if value == "JSON":
            self.start_button.configure(text="Tải File (JSON)")
        else:
            self.start_button.configure(text="Tải File (PDF)")
    
    def open_excel_folder(self):
        """Mở thư mục chứa file Excel đã xuất"""
        import subprocess
        folder_path = self.folder_entry.get()
        if folder_path and os.path.exists(folder_path):
            try:
                # Mở thư mục trong File Explorer
                subprocess.Popen(f'explorer "{folder_path}"')
                self.log_to_gui(f"Đã mở thư mục: {folder_path}")
            except Exception as e:
                self.log_to_gui(f"Lỗi khi mở thư mục: {e}")
        else:
            self.log_to_gui("Thư mục không tồn tại hoặc chưa được chọn!")

    def log_to_gui(self, message):
        self.after(0, self._update_log_textbox, message)

    def _update_log_textbox(self, message):
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", message + "\n")
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")

        # Kiểm tra các thông điệp kết thúc
        end_messages = ["HOÀN TẤT", "LỖI", "DỪNG THEO YÊU CẦU", "KHÔNG TÌM THẤY", "THẤT BẠI"]
        if any(msg in message.upper() for msg in end_messages):
            self._enable_all_controls()

    def _get_common_inputs(self):
        """Hàm nội bộ để lấy các thông tin chung cho RPA"""
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        start_day = self.start_day_combo.get()
        start_month = self.start_month_combo.get()
        start_year = self.start_year_combo.get()
        end_day = self.end_day_combo.get()
        end_month = self.end_month_combo.get()
        end_year = self.end_year_combo.get()
        
        start_date_str = f"{start_day}{start_month}{start_year}"
        end_date_str = f"{end_day}{end_month}{end_year}"

        if not username or not password:
            self.log_to_gui("LỖI: Vui lòng nhập Tên đăng nhập và Mật khẩu.")
            return None
        
        if self.save_creds_check.get() == 1:
            self.save_config(username, password)
        else:
            self.clear_config()
            
        return username, password, start_date_str, end_date_str
    
    def _disable_all_controls(self, is_rpa_task=True):
        """Vô hiệu hóa tất cả các nút điều khiển chính"""
        self.start_button.configure(state="disabled")
        self.check_button.configure(state="disabled")
        self.scrape_details_button.configure(state="disabled")
        self.extract_button.configure(state="disabled") 
        
        if is_rpa_task:
            self.pause_button.configure(state="normal", text="Tạm Dừng", fg_color="#F9A825", hover_color="#F57F17")
            self.stop_button.configure(state="normal")
        else:
            self.pause_button.configure(state="disabled", text="Tạm Dừng", fg_color="gray")
            self.stop_button.configure(state="disabled")


    def _enable_all_controls(self):
        """Kích hoạt lại các nút điều khiển chính (trạng thái chờ)"""
        self.start_button.configure(state="normal")
        self.check_button.configure(state="normal")
        self.scrape_details_button.configure(state="normal")
        self.extract_button.configure(state="normal")
        self.pause_button.configure(state="disabled", text="Tạm Dừng", fg_color="gray")
        self.stop_button.configure(state="disabled")
    
    def _show_completion_notification(self, title, message, success=True):
        """Hiển thị thông báo hoàn thành tác vụ"""
        icon = "info" if success else "warning"
        messagebox.showinfo(title, message, parent=self) if success else messagebox.showwarning(title, message, parent=self)

    # === TÁC VỤ 1: CHẠY LUỒNG TRÍCH XUẤT LOCAL EXCEL ===
    
    def start_extraction_thread(self):
        """Bắt đầu luồng trích xuất (Excel) (chạy độc lập)"""
        self._disable_all_controls(is_rpa_task=False)
        
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")
        self.log_to_gui("--- BẮT ĐẦU TRÍCH XUẤT FILE LOCAL SANG EXCEL ---")

        extraction_thread = threading.Thread(
            target=self._run_extraction_logic,
            daemon=True 
        )
        extraction_thread.start()

    def _run_extraction_logic(self):
        """Hàm logic chạy trong luồng (thread) trích xuất"""
        try:
            self.log_to_gui("Vui lòng chọn thư mục chứa file (PDF, JSON)...")
            source_directory = filedialog.askdirectory(title="Chọn thư mục chứa file PDF/JSON cần trích xuất")
            
            if not source_directory:
                self.log_to_gui("Đã hủy. Tác vụ trích xuất dừng lại.")
                self.after(0, self._enable_all_controls)
                return

            self.log_to_gui(f"Đang quét thư mục: {source_directory}")

            results = logic_convert_pdf.process_directory(source_directory)

            if not results:
                self.log_to_gui("Không tìm thấy file hợp đồng (PDF/JSON) nào trong thư mục đã chọn.")
                self.after(0, self._enable_all_controls)
                return

            self.log_to_gui(f"Đã trích xuất xong {len(results)} hợp đồng.")

            self.log_to_gui("Vui lòng chọn nơi lưu file Excel...")
            save_path = filedialog.asksaveasfilename(
                title="Lưu file Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="Danh_sach_hop_dong_local.xlsx"
            )

            if not save_path:
                self.log_to_gui("Đã hủy lưu file. Tác vụ trích xuất dừng lại.")
                self.after(0, self._enable_all_controls)
                return

            self.log_to_gui(f"Đang xuất ra file Excel: {save_path}")
            logic_convert_pdf.export_data_to_excel(results, save_path)
            
            self.log_to_gui(f"HOÀN TẤT! Đã lưu file Excel thành công (từ file local).")
            
            # Hiển thị thông báo hoàn thành
            self.after(0, lambda: self._show_completion_notification(
                "Trích xuất hoàn tất",
                f"Đã trích xuất {len(results)} hợp đồng và lưu vào:\n{save_path}",
                success=True
            ))

        except Exception as e:
            if "openpyxl" in str(e):
                self.log_to_gui("LỖI: Vui lòng cài đặt 'openpyxl' để xuất Excel.")
                self.log_to_gui("Chạy lệnh: pip install openpyxl")
            else:
                self.log_to_gui(f"LỖI TRÍCH XUẤT: {e}")
        finally:
            self.after(0, self._enable_all_controls)
            
    # === TÁC VỤ 2: CHẠY LUỒNG KIỂM TRA (RPA) ===
    
    def start_check_thread(self):
        """Lấy thông tin và bắt đầu luồng CHỈ KIỂM TRA."""
        self.pause_event.set()
        self.stop_event.clear()
        inputs = self._get_common_inputs()
        if not inputs: return
        username, password, start_date_str, end_date_str = inputs
        
        # Lấy headless mode từ checkbox
        headless = self.headless_mode_var.get()
        
        self._disable_all_controls(is_rpa_task=True)
        self.pause_button.configure(state="disabled", text="Tạm Dừng", fg_color="gray") # Không thể tạm dừng khi Kiểm tra
        
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")
        self.log_to_gui("--- BẮT ĐẦU KIỂM TRA SỐ LƯỢNG HỢP ĐỒNG ---")
        if headless:
            self.log_to_gui("⚙️ Chế độ: Chạy ngầm (headless)")

        self.rpa_thread = threading.Thread(
            target=self._run_check_with_notification,
            args=(username, password, start_date_str, end_date_str, headless),
            daemon=True 
        )
        self.rpa_thread.start()
    
    def _run_check_with_notification(self, username, password, start_date_str, end_date_str, headless):
        """Wrapper để thêm notification cho check_contract_count"""
        try:
            # Biến để lưu kết quả
            self.check_result = {'count': 0, 'success': False}
            
            # Tạo callback wrapper để capture kết quả
            def callback_wrapper(message):
                self.log_to_gui(message)
                # Capture số lượng từ message
                if "TÌM THẤY TỔNG CỘNG" in message:
                    import re
                    match = re.search(r'(\d+)\s+HỢP ĐỒNG', message)
                    if match:
                        self.check_result['count'] = int(match.group(1))
                        self.check_result['success'] = True
            
            # Gọi hàm RPA với callback wrapper
            rpa_logic.check_contract_count(
                username, password, start_date_str, end_date_str,
                self.pause_event, self.stop_event, callback_wrapper, headless
            )
            
            # Hiển thị notification nếu thành công
            if self.check_result['success']:
                self.after(0, lambda: self._show_completion_notification(
                    "Kiểm tra hoàn tất",
                    f"Tìm thấy tổng cộng {self.check_result['count']} hợp đồng trong khoảng thời gian đã chọn.",
                    success=True
                ))
        except Exception as e:
            self.log_to_gui(f"Lỗi: {e}")

    # === TÁC VỤ 3: CHẠY LUỒNG TẢI FILE (RPA) ===
    
    def start_rpa_thread(self):
        """Lấy thông tin và bắt đầu luồng TẢI FILE/LƯU JSON."""
        self.pause_event.set()
        self.stop_event.clear()
        inputs = self._get_common_inputs()
        if not inputs: return
        username, password, start_date_str, end_date_str = inputs

        save_directory = self.folder_entry.get()
        if not save_directory:
            self.log_to_gui("LỖI: Vui lòng chọn thư mục lưu file.")
            return

        save_format_value = self.save_format_button.get()
        save_format = "JSON" if save_format_value == "JSON" else "PDF"
        
        # Lấy headless mode từ checkbox
        headless = self.headless_mode_var.get()

        self._disable_all_controls(is_rpa_task=True)
        
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")
        self.log_to_gui(f"--- BẮT ĐẦU KỊCH BẢN TẢI FILE (LƯU DẠNG: {save_format}) ---")
        if headless:
            self.log_to_gui("⚙️ Chế độ: Chạy ngầm (headless)")

        self.rpa_thread = threading.Thread(
            target=self._run_download_with_notification,
            args=(username, password, start_date_str, end_date_str,
                  save_directory, save_format, headless),
            daemon=True 
        )
        self.rpa_thread.start()
    
    def _run_download_with_notification(self, username, password, start_date_str, end_date_str, 
                                        save_directory, save_format, headless):
        """Wrapper để thêm notification cho run_scrape_and_download_files"""
        try:
            # Biến để lưu kết quả
            self.download_result = {'count': 0, 'success': False, 'format': save_format}
            
            # Tạo callback wrapper để capture kết quả
            def callback_wrapper(message):
                self.log_to_gui(message)
                # Capture kết quả
                if "HOÀN TẤT" in message and "ĐÃ TẢI XONG" in message:
                    import re
                    match = re.search(r'(\d+)\s+HỢP ĐỒNG', message)
                    if match:
                        self.download_result['count'] = int(match.group(1))
                        self.download_result['success'] = True
            
            # Gọi hàm RPA
            rpa_logic.run_scrape_and_download_files(
                username, password, start_date_str, end_date_str,
                save_directory, save_format,
                self.pause_event, self.stop_event, callback_wrapper, headless
            )
            
            # Hiển thị notification nếu thành công
            if self.download_result['success']:
                self.after(0, lambda: self._show_completion_notification(
                    "Tải file hoàn tất",
                    f"Đã tải xong {self.download_result['count']} hợp đồng ({save_format}).\nLưu tại: {save_directory}",
                    success=True
                ))
        except Exception as e:
            self.log_to_gui(f"Lỗi: {e}")

    # === TÁC VỤ 4: CHẠY LUỒNG CÀO CHI TIẾT (RPA) ===
    
    def start_detail_scrape_thread(self):
        """Lấy thông tin và bắt đầu luồng CÀO CHI TIẾT."""
        self.pause_event.set()
        self.stop_event.clear()
        inputs = self._get_common_inputs()
        if not inputs: return
        username, password, start_date_str, end_date_str = inputs

        save_directory = self.folder_entry.get()
        if not save_directory:
            self.log_to_gui("LỖI: Vui lòng chọn thư mục lưu file (Excel).")
            return
        
        # Lấy headless mode từ checkbox
        headless = self.headless_mode_var.get()

        self._disable_all_controls(is_rpa_task=True)
        
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")
        self.log_to_gui(f"--- BẮT ĐẦU KỊCH BẢN CÀO (SCRAPE) CHI TIẾT RA EXCEL ---")
        if headless:
            self.log_to_gui("⚙️ Chế độ: Chạy ngầm (headless)")

        self.rpa_thread = threading.Thread(
            target=self._run_scrape_details_with_notification,
            args=(username, password, start_date_str, end_date_str,
                  save_directory, headless),
            daemon=True 
        )
        self.rpa_thread.start()
    
    def _run_scrape_details_with_notification(self, username, password, start_date_str, end_date_str, 
                                              save_directory, headless):
        """Wrapper để thêm notification cho run_scrape_and_export_details"""
        try:
            # Biến để lưu kết quả
            self.scrape_result = {'count': 0, 'success': False}
            
            # Tạo callback wrapper để capture kết quả
            def callback_wrapper(message):
                self.log_to_gui(message)
                # Capture kết quả
                if "HOÀN TẤT" in message and "ĐÃ CÀO XONG" in message:
                    import re
                    match = re.search(r'(\d+)\s+HỢP ĐỒNG', message)
                    if match:
                        self.scrape_result['count'] = int(match.group(1))
                        self.scrape_result['success'] = True
            
            # Gọi hàm RPA
            rpa_logic.run_scrape_and_export_details(
                username, password, start_date_str, end_date_str,
                save_directory,
                self.pause_event, self.stop_event, callback_wrapper, headless
            )
            
            # Hiển thị notification nếu thành công
            if self.scrape_result['success']:
                self.after(0, lambda: self._show_completion_notification(
                    "Cào dữ liệu hoàn tất",
                    f"Đã cào xong {self.scrape_result['count']} hợp đồng và xuất ra Excel.\nLưu tại: {save_directory}",
                    success=True
                ))
        except Exception as e:
            self.log_to_gui(f"Lỗi: {e}")

    # --- Các hàm helper còn lại ---
    
    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r") as f:
                    config_data = json.load(f)
                    # Giải mã cả username và password
                    encoded_user = config_data.get("username", "")
                    encoded_pass = config_data.get("password", "")
                    username = base64.b64decode(encoded_user.encode()).decode() if encoded_user else ""
                    password = base64.b64decode(encoded_pass.encode()).decode() if encoded_pass else ""
                    self.username_entry.insert(0, username)
                    self.password_entry.insert(0, password)
                    self.save_creds_check.select()
            except Exception as e:
                self.log_to_gui(f"Lỗi: Không thể đọc tệp config.json. Lỗi: {e}")

    def save_config(self, username, password):
        try:
            # Mã hóa cả username và password
            encoded_user = base64.b64encode(username.encode()).decode()
            encoded_pass = base64.b64encode(password.encode()).decode()
            config_data = {"username": encoded_user, "password": encoded_pass}
            with open(self.config_file, "w") as f:
                json.dump(config_data, f)
            self.log_to_gui("Đã lưu thông tin đăng nhập (đã mã hóa).")
        except Exception as e:
            self.log_to_gui(f"Lỗi: Không thể lưu config. Lỗi: {e}")

    def load_zalo_session_info(self):
        """Load thông tin session Zalo khi khởi động ứng dụng"""
        try:
            import zalo_logic

            # Khởi tạo account manager
            self.account_manager = zalo_logic.ZaloAccountManager()

            # Load danh sách tài khoản
            self.load_account_list()

        except Exception as e:
            print(f"Lỗi khi load session info: {str(e)}")

    def clear_config(self):
        if os.path.exists(self.config_file):
            try:
                os.remove(self.config_file)
                self.log_to_gui("Đã xóa thông tin đăng nhập đã lưu.")
            except Exception as e:
                self.log_to_gui(f"Lỗi: Không thể xóa config. Lỗi: {e}")

    def toggle_password_visibility(self):
        if self.show_password_check.get() == 1:
            self.password_entry.configure(show="")
        else:
            self.password_entry.configure(show="*")

    def select_folder(self):
        path = filedialog.askdirectory(initialdir=self.folder_entry.get())
        if path:
            self.folder_entry.configure(state="normal")
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, path)
            self.folder_entry.configure(state="disabled")

    def toggle_pause(self):
        if self.pause_event.is_set():
            self.pause_event.clear()
            self.pause_button.configure(text="Tiếp Tục", fg_color="#4CAF50", hover_color="#388E3C")
            self.log_to_gui("...Đã tạm dừng. Nhấn 'Tiếp Tục' để chạy...")
        else:
            self.pause_event.set()
            self.pause_button.configure(text="Tạm Dừng", fg_color="#F9A825", hover_color="#F57F17")
            self.log_to_gui("...Đang tiếp tục...")

    def stop_rpa(self):
        if self.rpa_thread and self.rpa_thread.is_alive():
            self.log_to_gui("--- !!! ĐANG GỬI LỆNH DỪNG, VUI LÒNG CHỜ... !!! ---")
            self.stop_event.set()
            self.pause_event.set() 
            self._disable_all_controls(is_rpa_task=False) # Tắt hết nút
            self.stop_button.configure(text="Đang dừng...")

    # --- HÀM PLACEHOLDER CHO ZALO & GOOGLE SHEETS ---
    def open_zalo_window(self):
        """Mở cửa sổ mới để xử lý Zalo với Session Management"""
        self.log_to_gui("🔵 Đang mở Zalo với session management...")
        
        # Chạy trong thread riêng để không block UI
        zalo_thread = threading.Thread(
            target=self._run_zalo_login,
            daemon=True
        )
        zalo_thread.start()
    
    def check_zalo_status(self):
        """Kiểm tra trạng thái đăng nhập Zalo - Mở Zalo, kiểm tra cookie và cập nhật họ tên"""
        # Chạy trong thread riêng
        thread = threading.Thread(
            target=self._run_check_zalo_status,
            daemon=True
        )
        thread.start()

    def _run_check_zalo_status(self):
        """Thread worker để kiểm tra và cập nhật thông tin Zalo"""
        try:
            import zalo_logic
            import zalo_automation
            from datetime import datetime

            # Kiểm tra đã chọn tài khoản chưa
            if not self.current_account_id:
                self.log_to_gui("Vui lòng chọn tài khoản trước!")
                messagebox.showwarning(
                    "Cảnh báo",
                    "Vui lòng chọn tài khoản Zalo trước khi kiểm tra!",
                    parent=self
                )
                return

            self.log_to_gui("[CHO] Đang kiểm tra trạng thái Zalo...")

            # Lấy session manager cho tài khoản đã chọn
            session_manager = self.account_manager.get_session_manager(self.current_account_id)

            # Đăng nhập Zalo
            success, p, context, page = session_manager.login_with_session(max_wait_time=30)

            if not success:
                self.log_to_gui("Không thể kết nối Zalo")
                if context:
                    context.close()
                if p:
                    p.stop()
                return

            self.log_to_gui("Đã kết nối Zalo")

            # Tạo automation instance
            automation = zalo_automation.ZaloAutomation(page)

            # Lấy tên Zalo (sẽ tự động lưu vào session)
            my_zalo_name = automation.get_my_zalo_name(session_manager)

            # Cập nhật thời gian kiểm tra
            current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            session_info = session_manager.get_session_info() or {}
            session_info['last_check'] = current_time
            session_manager.save_session_info(session_info)

            # Cập nhật thông tin tài khoản trong account manager
            last_login = session_info.get('last_login', current_time)
            self.account_manager.update_account(
                self.current_account_id,
                zalo_name=my_zalo_name,
                last_login=last_login,
                status='active'
            )

            self.log_to_gui(f"Cập nhật thành công!")
            self.log_to_gui(f"   Họ tên: {my_zalo_name}")
            self.log_to_gui(f"   Phiên đăng nhập: {last_login}")
            self.log_to_gui(f"   [DANG XU LY] Kiểm tra lúc: {current_time}")

            # Reload danh sách tài khoản để cập nhật tên Zalo
            self.load_account_list()

            # Cleanup
            try:
                if context:
                    context.close()
                if p:
                    p.stop()
            except:
                pass

        except Exception as e:
            self.log_to_gui(f"Lỗi khi kiểm tra: {str(e)}")
    
    def clear_zalo_session(self):
        """Xóa session Zalo đã lưu"""
        try:
            # Kiểm tra đã chọn tài khoản chưa
            if not self.current_account_id:
                messagebox.showwarning(
                    "Cảnh báo",
                    "Vui lòng chọn tài khoản trước!",
                    parent=self
                )
                return

            # Xác nhận trước khi xóa
            result = messagebox.askyesno(
                "Xác nhận xóa session",
                "Bạn có chắc muốn xóa phiên đăng nhập Zalo?\nBạn sẽ phải quét QR lại lần sau.",
                parent=self
            )

            if result:
                session_manager = self.account_manager.get_session_manager(self.current_account_id)
                if session_manager.delete_session():
                    self.log_to_gui("Đã xóa phiên đăng nhập Zalo")

                    # Cập nhật trạng thái tài khoản
                    self.account_manager.update_account(
                        self.current_account_id,
                        status='inactive',
                        last_login=None
                    )

                    # Reload danh sách tài khoản
                    self.load_account_list()

                    messagebox.showinfo(
                        "Thành công",
                        "Đã xóa phiên đăng nhập Zalo!",
                        parent=self
                    )
                else:
                    self.log_to_gui("Không thể xóa phiên đăng nhập")
                    messagebox.showerror(
                        "Lỗi",
                        "Không thể xóa phiên đăng nhập!",
                        parent=self
                    )
            
        except Exception as e:
            self.log_to_gui(f"Lỗi khi xóa session: {str(e)}")
    
    def _run_zalo_login(self):
        """Logic đăng nhập Zalo với Session Management và Headless Mode"""
        playwright_instance = None
        context = None

        try:
            import zalo_logic

            # Kiểm tra đã chọn tài khoản chưa
            if not self.current_account_id:
                self.log_to_gui("Vui lòng chọn tài khoản trước!")
                messagebox.showwarning(
                    "Cảnh báo",
                    "Vui lòng chọn tài khoản Zalo trước khi mở Zalo!",
                    parent=self
                )
                return

            # Lấy headless mode từ checkbox trang chủ
            headless = self.headless_mode_var.get()
            
            self.log_to_gui("[CHO] Đang khởi tạo Zalo Session Manager...")
            if headless:
                self.log_to_gui("⚙️ Chế độ: Chạy ngầm (headless)")
                self.log_to_gui("ℹ️ Lưu ý: Nếu cần quét QR, trình duyệt sẽ tự động hiện lên")

            # Lấy session manager cho tài khoản đã chọn
            session_manager = self.account_manager.get_session_manager(self.current_account_id)

            # Kiểm tra session hiện có
            if session_manager.has_session():
                self.log_to_gui("Tìm thấy phiên đăng nhập đã lưu")
                session_info = session_manager.get_session_info()
                if session_info:
                    self.log_to_gui(f"  - Lần đăng nhập cuối: {session_info.get('last_login', 'N/A')}")
            else:
                self.log_to_gui("[ℹ️] Chưa có phiên đăng nhập, sẽ đăng nhập mới")
                if headless:
                    self.log_to_gui("[ℹ️] Trình duyệt sẽ hiện lên để bạn quét QR code")

            self.log_to_gui("📱 Đang mở Zalo...")

            # Đăng nhập với session persistence và headless mode
            success, playwright_instance, context, page = session_manager.login_with_session(
                max_wait_time=300,
                headless=headless
            )

            if success:
                self.log_to_gui("✅ Đăng nhập Zalo thành công!")
                self.log_to_gui("Phiên đăng nhập đã được lưu tự động")
                
                if headless:
                    self.log_to_gui("ℹ️ Trình duyệt đang chạy ở chế độ ngầm")
                else:
                    self.log_to_gui("Trình duyệt sẽ vẫn mở để bạn sử dụng Zalo")
                
                self.log_to_gui("Nhấn nút 'Kiểm tra & Cập nhật' để cập nhật thông tin tài khoản")
                self.log_to_gui("Đóng cửa sổ trình duyệt khi bạn hoàn tất công việc")

                # Giữ browser mở - chờ người dùng đóng
                # Context sẽ tự động lưu khi đóng
                try:
                    # Chờ cho đến khi context bị đóng
                    while not page.is_closed():
                        import time
                        import random
                        time.sleep(random.uniform(0.8, 1.2))
                except:
                    pass

                self.log_to_gui("[ℹ️] Đã đóng trình duyệt Zalo")

            else:
                self.log_to_gui("❌ Đăng nhập Zalo thất bại hoặc hết thời gian chờ")

        except ImportError as e:
            if "playwright" in str(e):
                self.log_to_gui("Lỗi: Chưa cài đặt Playwright!")
                self.log_to_gui("Vui lòng chạy: pip install playwright")
                self.log_to_gui("Sau đó chạy: playwright install chromium")
            else:
                self.log_to_gui(f"Lỗi import: {str(e)}")
        except Exception as e:
            self.log_to_gui(f"Lỗi khi mở Zalo: {str(e)}")
        finally:
            # Cleanup khi kết thúc
            try:
                if context:
                    context.close()
                if playwright_instance:
                    playwright_instance.stop()
            except:
                pass

    # === CÁC HÀM XỬ LÝ CHO TAB KIỂM TRA HỢP ĐỒNG ===

    def select_contract_file(self):
        """Chọn file Excel chứa dữ liệu hợp đồng"""
        try:
            file_path = filedialog.askopenfilename(
                title="Chọn file Excel chứa dữ liệu hợp đồng",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )

            if not file_path:
                return

            self.log_to_gui(f"[📄] Đang đọc file: {file_path}")

            # Đọc file Excel
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Đọc header
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else "")

            # Tìm các cột cần thiết (theo đúng tên cột trong file mẫu)
            contract_col = None
            cccd_col = None
            name_col = None
            loan_amount_col = None
            months_col = None

            for idx, header in enumerate(headers):
                header_stripped = header.strip()

                # Cột "ID Hợp đồng"
                if contract_col is None:
                    if header_stripped == "ID Hợp đồng" or "ID Hợp đồng" in header_stripped:
                        contract_col = idx

                # Cột "Số CCCD"
                if cccd_col is None:
                    if header_stripped == "Số CCCD" or "Số CCCD" in header_stripped:
                        cccd_col = idx

                # Cột "Tên KH (Profile)"
                if name_col is None:
                    if header_stripped == "Tên KH (Profile)" or "Tên KH" in header_stripped:
                        name_col = idx

                # Cột "Số tiền vay" (index 32)
                if loan_amount_col is None:
                    if header_stripped == "Số tiền vay":
                        loan_amount_col = idx

                # Cột "Số tháng"
                if months_col is None:
                    if header_stripped == "Số tháng" or "Số tháng" in header_stripped:
                        months_col = idx

            if contract_col is None and cccd_col is None:
                messagebox.showerror(
                    "Lỗi",
                    "Không tìm thấy cột 'Số hợp đồng' hoặc 'CCCD' trong file Excel!",
                    parent=self
                )
                return

            # Log các cột đã tìm thấy
            print(f"[EXCEL] Cột tìm thấy:")
            print(f"[EXCEL]   - ID Hợp đồng: {contract_col}")
            print(f"[EXCEL]   - Số CCCD: {cccd_col}")
            print(f"[EXCEL]   - Tên KH: {name_col}")
            print(f"[EXCEL]   - Số tiền vay: {loan_amount_col}")
            print(f"[EXCEL]   - Số tháng: {months_col}")

            # Đọc dữ liệu
            self.contract_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue

                contract_id = str(row[contract_col] or "").strip() if contract_col is not None else ""
                cccd = str(row[cccd_col] or "").strip() if cccd_col is not None else ""
                name = str(row[name_col] or "").strip() if name_col is not None else ""

                # Xử lý số tiền vay (loại bỏ dấu phẩy, dấu chấm, ký tự "đ")
                loan_amount = 0
                if loan_amount_col is not None and row[loan_amount_col]:
                    try:
                        # Loại bỏ tất cả ký tự không phải số
                        loan_str = str(row[loan_amount_col]).replace(",", "").replace(".", "").replace("đ", "").replace(" ", "").replace("\n", "").replace("\r", "").strip()
                        # Chỉ lấy các ký tự số
                        loan_str = ''.join(c for c in loan_str if c.isdigit())
                        loan_amount = float(loan_str) if loan_str else 0
                        if loan_amount > 0:
                            print(f"[EXCEL] Số tiền vay: {loan_amount:,.0f} VNĐ (từ '{row[loan_amount_col]}')")
                    except Exception as e:
                        print(f"[EXCEL] Lỗi đọc số tiền vay: {e}, giá trị: {row[loan_amount_col]}")
                        loan_amount = 0

                # Xử lý số tháng (loại bỏ chữ "tháng")
                months = 0
                if months_col is not None and row[months_col]:
                    try:
                        months_str = str(row[months_col]).replace("tháng", "").replace(" ", "").replace("\n", "").replace("\r", "").strip()
                        # Chỉ lấy các ký tự số
                        months_str = ''.join(c for c in months_str if c.isdigit())
                        months = int(months_str) if months_str else 0
                        if months > 0:
                            print(f"[EXCEL] Số tháng: {months} (từ '{row[months_col]}')")
                    except Exception as e:
                        print(f"[EXCEL] Lỗi đọc số tháng: {e}, giá trị: {row[months_col]}")
                        months = 0

                if contract_id or cccd:
                    self.contract_data.append({
                        'contract_id': contract_id,
                        'cccd': cccd,
                        'name': name,
                        'loan_amount': loan_amount,
                        'months': months
                    })

            # Cập nhật UI
            self.contract_excel_path = file_path
            file_name = file_path.split("/")[-1].split("\\")[-1]
            self.contract_file_label.configure(text=f"[TAI LIEU] {file_name}", text_color="#28A745")
            self.contract_count_label.configure(text=f"Số lượng: {len(self.contract_data)}", text_color="#28A745")

            self.log_to_gui(f"Đã tải {len(self.contract_data)} hợp đồng từ file Excel")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file Excel:\n{str(e)}", parent=self)
            self.log_to_gui(f"Lỗi khi đọc file: {str(e)}")

    def check_contracts(self):
        """Kiểm tra hợp đồng - ưu tiên nhập thủ công, nếu không có thì dùng file"""
        contract_number = self.contract_number_entry.get().strip()
        cccd = self.cccd_entry.get().strip()
        
        # Lấy giá trị headless trước khi tạo thread
        headless = self.headless_mode_var.get()

        # Trường hợp 1: Có nhập thủ công
        if contract_number or cccd:
            self.log_to_gui("\n" + "="*60)
            self.log_to_gui("🔍 Kiểm tra hợp đồng (Nhập thủ công)")
            self.log_to_gui(f"Số hợp đồng: {contract_number if contract_number else 'Không có'}")
            self.log_to_gui(f"Số CCCD: {cccd if cccd else 'Không có'}")

            # Chạy trong thread riêng
            import threading
            thread = threading.Thread(
                target=self._run_check_single_contract,
                args=(contract_number, cccd, headless),
                daemon=True
            )
            thread.start()

        # Trường hợp 2: Kiểm tra từ file
        elif self.contract_data:
            self.log_to_gui("\n" + "="*60)
            self.log_to_gui(f"🔍 Kiểm tra {len(self.contract_data)} hợp đồng từ file")

            # Chạy trong thread riêng
            import threading
            thread = threading.Thread(
                target=self._run_check_contracts_from_file,
                args=(headless,),
                daemon=True
            )
            thread.start()

        else:
            messagebox.showwarning(
                "Cảnh báo",
                "Vui lòng nhập thông tin hợp đồng hoặc chọn file Excel!",
                parent=self
            )

    def _run_check_single_contract(self, contract_number, cccd, headless):
        """Thread worker để kiểm tra 1 hợp đồng"""
        try:
            from playwright.sync_api import sync_playwright
            import time
            import random

            print("\n" + "="*80)
            print("KIỂM TRA HỢP ĐỒNG ĐƠN LẺ")
            print("="*80)
            
            # Hiển thị chế độ chạy
            mode_text = "chạy ngầm" if headless else "hiển thị trình duyệt"
            self.log_to_gui(f"[CHO] Đang khởi tạo trình duyệt ({mode_text})...")

            with sync_playwright() as p:
                # Launch với các tham số tối ưu
                launch_args = {
                    'headless': headless,
                    'args': [
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox',
                        '--disable-setuid-sandbox',
                        '--disable-web-security',
                        '--disable-features=IsolateOrigins,site-per-process'
                    ]
                }
                browser = p.chromium.launch(**launch_args)
                context = browser.new_context()
                page = context.new_page()

                print("[INIT] Đang mở trang tra cứu...")
                self.log_to_gui("[WEB] Đang truy cập trang tra cứu...")
                page.goto("https://www.hdsaison.com.vn/vn/khach-hang/tra-cuu-khoan-vay.html")
                page.wait_for_load_state("networkidle")

                # Đóng popup nếu có
                self._close_popup_if_exists(page)

                print("[INIT] Trang đã sẵn sàng")
                self.log_to_gui("Đã mở trang tra cứu")
                time.sleep(random.uniform(1.0, 1.5))

                # Điền số hợp đồng (nếu có)
                if contract_number:
                    print(f"[INPUT] Nhập số hợp đồng: {contract_number}")
                    self.log_to_gui(f"[✍️] Nhập số hợp đồng: {contract_number}")

                    self._close_popup_if_exists(page)
                    contract_input = page.wait_for_selector('input#contract_code', timeout=10000)
                    contract_input.click()
                    time.sleep(0.2)
                    self._close_popup_if_exists(page)
                    contract_input.fill(contract_number)
                    time.sleep(random.uniform(0.3, 0.5))

                # Điền số CCCD (nếu có)
                if cccd:
                    print(f"[INPUT] Nhập số CCCD: {cccd}")
                    self.log_to_gui(f"[✍️] Nhập số CCCD: {cccd}")

                    self._close_popup_if_exists(page)
                    cccd_input = page.wait_for_selector('input#cmnd', timeout=10000)
                    cccd_input.click()
                    time.sleep(0.2)
                    self._close_popup_if_exists(page)
                    cccd_input.fill(cccd)
                    time.sleep(random.uniform(0.3, 0.5))

                # Đóng popup trước khi tìm kiếm
                self._close_popup_if_exists(page)

                # Nhấn nút tìm kiếm
                print("[SEARCH] Nhấn nút Tìm kiếm...")
                self.log_to_gui("🔍 Nhấn nút Tìm kiếm...")
                search_button = page.wait_for_selector('button.btnStyAll.btnFull:has-text("Tìm kiếm")', timeout=10000)
                search_button.click()

                print("[SEARCH] Đang chờ kết quả...")
                self.log_to_gui("[CHO] Đang chờ kết quả...")
                time.sleep(3)

                # Đóng popup sau khi tìm kiếm
                self._close_popup_if_exists(page)

                # Kiểm tra có bảng thông tin không
                try:
                    print("[RESULT] Kiểm tra kết quả...")
                    info_table = page.wait_for_selector('div.boxTableBd', timeout=5000)
                    print("[RESULT] Tìm thấy bảng thông tin")
                    self.log_to_gui("Tìm thấy thông tin hợp đồng")

                    # Lấy thông tin từ bảng
                    name_element = page.query_selector('div.trBody:has-text("Họ Tên khách hàng") div.td')
                    contract_element = page.query_selector('div.trBody:has-text("Số hợp đồng") div.td')

                    if name_element and contract_element:
                        result_name = name_element.inner_text().strip()
                        result_contract = contract_element.inner_text().strip()

                        print(f"[RESULT] Họ tên: {result_name}")
                        print(f"[RESULT] Số hợp đồng: {result_contract}")

                        self.log_to_gui(f"[DANH SACH] Họ tên: {result_name}")
                        self.log_to_gui(f"[DANH SACH] Số hợp đồng: {result_contract}")

                        # Đóng popup trước khi nhấn nút
                        self._close_popup_if_exists(page)

                        # Nhấn nút "Lịch sử thanh toán thực tế"
                        print("[HISTORY] Nhấn nút 'Lịch sử thanh toán thực tế'...")
                        self.log_to_gui("🔍 Xem lịch sử thanh toán...")

                        # Kiểm tra xem link có mở tab mới không
                        history_button = page.wait_for_selector('a.btnStyAll.btnFull:has-text("Lịch sử thanh toán thực tế")', timeout=5000)

                        # Lấy href để kiểm tra
                        href = history_button.get_attribute('href')
                        print(f"[HISTORY] Link: {href}")

                        # Click và chờ navigation
                        try:
                            with page.expect_navigation(timeout=10000):
                                history_button.click()
                            print(f"[HISTORY] Đã chuyển trang")
                        except:
                            # Nếu không có navigation, có thể mở tab mới
                            print(f"[HISTORY] Không có navigation, thử cách khác...")
                            page.goto(f"https://www.hdsaison.com.vn{href}" if href.startswith('/') else href)

                        time.sleep(2)
                        page.wait_for_load_state("networkidle")

                        # Đóng popup sau khi chuyển trang
                        self._close_popup_if_exists(page)

                        # Đếm số lần đóng tiền và tổng tiền
                        print("[HISTORY] Đang đếm lịch sử thanh toán...")
                        payment_rows = page.query_selector_all('div.tableTra table tbody tr')
                        total_payments = len(payment_rows)
                        total_amount = 0

                        print(f"[HISTORY] Số lần đóng tiền: {total_payments}")
                        self.log_to_gui(f"[TIEN] Số lần đóng tiền: {total_payments}")

                        for row_idx, row in enumerate(payment_rows, 1):
                            amount_element = row.query_selector('span.number_vnd')
                            if amount_element:
                                amount_text = amount_element.inner_text().strip().replace(".", "").replace(",", "")
                                try:
                                    amount = float(amount_text)
                                    total_amount += amount
                                    print(f"[HISTORY]   Lần {row_idx}: {amount:,.0f} VNĐ")
                                except Exception as e:
                                    print(f"[HISTORY] [X] Lỗi đọc số tiền lần {row_idx}: {e}")

                        print(f"[HISTORY] Tổng tiền đã đóng: {total_amount:,.0f} VNĐ")
                        self.log_to_gui(f"[TIEN] Tổng số tiền đã đóng: {total_amount:,.0f} VNĐ")
                        self.log_to_gui("Đã hoàn thành tra cứu")

                    else:
                        print("[RESULT] [X] Không lấy được thông tin từ bảng")
                        self.log_to_gui("Không lấy được thông tin từ bảng")

                except Exception as e:
                    print(f"[ERROR] [X] LỖI: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    self.log_to_gui(f"Không tìm thấy thông tin hợp đồng hoặc lỗi: {str(e)}")

                # Giữ trình duyệt mở để xem
                print("[INFO] Giữ trình duyệt mở 5 phút...")
                time.sleep(300)  # Giữ mở 5 phút

                print("[INFO] Đóng trình duyệt...")
                context.close()
                browser.close()

        except Exception as e:
            print(f"[ERROR] [X] LỖI NGHIÊM TRỌNG: {str(e)}")
            import traceback
            traceback.print_exc()
            self.log_to_gui(f"Lỗi: {str(e)}")

    def _close_popup_if_exists(self, page):
        """Đóng popup tư vấn nếu xuất hiện"""
        try:
            popup = page.query_selector('div.dct-container')
            if popup and popup.is_visible():
                print("[POPUP] Phát hiện popup tư vấn, đang đóng...")
                close_button = page.query_selector('button#close-pop-support')
                if close_button:
                    close_button.click()
                    print("[POPUP] Đã đóng popup")
                    import time
                    time.sleep(0.5)
                    return True
        except:
            pass
        return False

    def _run_check_contracts_from_file(self, headless):
        """Thread worker để kiểm tra nhiều hợp đồng từ file"""
        try:
            from playwright.sync_api import sync_playwright
            import time
            import random

            print("\n" + "="*80)
            print("BẮT ĐẦU KIỂM TRA HỢP ĐỒNG HÀNG LOẠT")
            print("="*80)
            
            # Hiển thị chế độ chạy
            mode_text = "chạy ngầm" if headless else "hiển thị trình duyệt"
            self.log_to_gui(f"[CHO] Đang khởi tạo trình duyệt ({mode_text})...")

            # Danh sách lưu kết quả để ghi vào Excel
            results = []

            with sync_playwright() as p:
                # Launch với các tham số tối ưu
                launch_args = {
                    'headless': headless,
                    'args': [
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox',
                        '--disable-setuid-sandbox',
                        '--disable-web-security',
                        '--disable-features=IsolateOrigins,site-per-process'
                    ]
                }
                browser = p.chromium.launch(**launch_args)
                context = browser.new_context()
                page = context.new_page()

                print("\n[INIT] Đang mở trang tra cứu...")
                self.log_to_gui("[WEB] Đang mở trang tra cứu...")

                page.goto("https://www.hdsaison.com.vn/vn/khach-hang/tra-cuu-khoan-vay.html")
                page.wait_for_load_state("networkidle")
                print("[INIT] Trang đã sẵn sàng")

                # Đóng popup nếu có
                self._close_popup_if_exists(page)

                self.log_to_gui("Đã mở trang tra cứu")
                time.sleep(1)

                # Kiểm tra từng hợp đồng
                total_contracts = len(self.contract_data)

                for idx, item in enumerate(self.contract_data, 1):
                    print(f"\n{'='*80}")
                    print(f"[{idx}/{total_contracts}] KIỂM TRA HỢP ĐỒNG")
                    print(f"{'='*80}")

                    self.log_to_gui(f"\n{'='*60}")
                    self.log_to_gui(f"[{idx}/{total_contracts}] Kiểm tra:")
                    self.log_to_gui(f"  - Số hợp đồng: {item['contract_id']}")
                    self.log_to_gui(f"  - Số CCCD: {item['cccd']}")
                    self.log_to_gui(f"  - Tên KH (Excel): {item.get('name', 'N/A')}")
                    self.log_to_gui(f"  - Số tiền vay (Excel): {item.get('loan_amount', 0):,.0f} VNĐ")
                    self.log_to_gui(f"  - Số tháng (Excel): {item.get('months', 0)}")

                    print(f"[DATA] Số hợp đồng: {item['contract_id']}")
                    print(f"[DATA] Số CCCD: {item['cccd']}")
                    print(f"[DATA] Tên KH: {item.get('name', 'N/A')}")
                    print(f"[DATA] Số tiền vay: {item.get('loan_amount', 0):,.0f} VNĐ")
                    print(f"[DATA] Số tháng: {item.get('months', 0)}")

                    # Đóng popup trước khi bắt đầu
                    self._close_popup_if_exists(page)

                    # Điền số hợp đồng
                    if item['contract_id']:
                        print(f"[INPUT] Nhập số hợp đồng: {item['contract_id']}")
                        self.log_to_gui(f"  [✍️] Nhập số hợp đồng")

                        # Đóng popup trước khi nhập
                        self._close_popup_if_exists(page)

                        contract_input = page.wait_for_selector('input#contract_code', timeout=10000)
                        contract_input.click()
                        time.sleep(0.2)

                        # Đóng popup sau khi click
                        self._close_popup_if_exists(page)

                        contract_input.fill("")  # Clear
                        contract_input.fill(item['contract_id'])
                        print(f"[INPUT] Đã nhập số hợp đồng")
                        time.sleep(random.uniform(0.3, 0.5))

                    # Điền số CCCD
                    if item['cccd']:
                        print(f"[INPUT] Nhập số CCCD: {item['cccd']}")
                        self.log_to_gui(f"  [✍️] Nhập số CCCD")

                        # Đóng popup trước khi nhập
                        self._close_popup_if_exists(page)

                        cccd_input = page.wait_for_selector('input#cmnd', timeout=10000)
                        cccd_input.click()
                        time.sleep(0.2)

                        # Đóng popup sau khi click
                        self._close_popup_if_exists(page)

                        cccd_input.fill("")  # Clear
                        cccd_input.fill(item['cccd'])
                        print(f"[INPUT] Đã nhập số CCCD")
                        time.sleep(random.uniform(0.3, 0.5))

                    # Đóng popup trước khi nhấn tìm kiếm
                    self._close_popup_if_exists(page)

                    # Nhấn nút tìm kiếm
                    print(f"[SEARCH] Nhấn nút Tìm kiếm...")
                    self.log_to_gui(f"  🔍 Tìm kiếm...")
                    search_button = page.wait_for_selector('button.btnStyAll.btnFull:has-text("Tìm kiếm")', timeout=10000)
                    search_button.click()

                    print(f"[SEARCH] Đang chờ kết quả...")
                    self.log_to_gui(f"  [CHO] Chờ kết quả...")
                    time.sleep(3)

                    # Đóng popup sau khi tìm kiếm
                    self._close_popup_if_exists(page)

                    # Kiểm tra kết quả
                    try:
                        # Đóng popup trước khi đọc kết quả
                        self._close_popup_if_exists(page)

                        print(f"[RESULT] Kiểm tra kết quả...")
                        try:
                            info_table = page.wait_for_selector('div.boxTableBd', timeout=5000)
                            print(f"[RESULT] Tìm thấy bảng thông tin")
                            self.log_to_gui(f"  Tìm thấy thông tin hợp đồng")
                        except:
                            print(f"[RESULT] [X] Không tìm thấy thông tin hợp đồng")
                            self.log_to_gui(f"  Không tìm thấy thông tin hợp đồng")

                            # Lưu kết quả lỗi
                            results.append({
                                'contract_id': item['contract_id'],
                                'status': 'Không tìm thấy thông tin'
                            })
                            continue

                        # Lấy thông tin từ bảng
                        name_element = page.query_selector('div.trBody:has-text("Họ Tên khách hàng") div.td')
                        contract_element = page.query_selector('div.trBody:has-text("Số hợp đồng") div.td')

                        if name_element and contract_element:
                            result_name = name_element.inner_text().strip()
                            result_contract = contract_element.inner_text().strip()

                            print(f"[RESULT] Họ tên (Web): {result_name}")
                            print(f"[RESULT] Số HĐ (Web): {result_contract}")

                            self.log_to_gui(f"  [DANH SACH] Họ tên (Web): {result_name}")
                            self.log_to_gui(f"  [DANH SACH] Số HĐ (Web): {result_contract}")

                            # So sánh tên và số hợp đồng
                            name_match = result_name.upper() == item.get('name', '').upper()
                            contract_match = result_contract == item['contract_id']

                            print(f"[COMPARE] So sánh dữ liệu:")
                            print(f"[COMPARE]   - Tên khớp: {name_match}")
                            print(f"[COMPARE]   - Số HĐ khớp: {contract_match}")

                            if name_match and contract_match:
                                print(f"[COMPARE] Thông tin khớp với Excel")
                                self.log_to_gui(f"  Thông tin khớp với Excel")

                                # Đóng popup trước khi nhấn nút
                                self._close_popup_if_exists(page)

                                # Nhấn nút "Lịch sử thanh toán thực tế"
                                print(f"[HISTORY] Nhấn nút 'Lịch sử thanh toán thực tế'...")
                                self.log_to_gui(f"  🔍 Xem lịch sử thanh toán...")

                                # Kiểm tra xem link có mở tab mới không
                                history_button = page.wait_for_selector('a.btnStyAll.btnFull:has-text("Lịch sử thanh toán thực tế")', timeout=5000)

                                # Lấy href để kiểm tra
                                href = history_button.get_attribute('href')
                                print(f"[HISTORY] Link: {href}")

                                # Click và chờ navigation
                                try:
                                    with page.expect_navigation(timeout=10000):
                                        history_button.click()
                                    print(f"[HISTORY] Đã chuyển trang")
                                except:
                                    # Nếu không có navigation, có thể mở tab mới
                                    print(f"[HISTORY] Không có navigation, thử cách khác...")
                                    page.goto(f"https://www.hdsaison.com.vn{href}" if href.startswith('/') else href)

                                time.sleep(2)
                                page.wait_for_load_state("networkidle")

                                # Đóng popup sau khi chuyển trang
                                self._close_popup_if_exists(page)

                                # Kiểm tra xem page còn hoạt động không
                                try:
                                    if page.is_closed():
                                        print(f"[ERROR] [X] Page đã bị đóng!")
                                        self.log_to_gui(f"  Trang đã bị đóng")
                                        continue
                                except:
                                    print(f"[ERROR] [X] Không thể kiểm tra trạng thái page")
                                    continue

                                # Đếm số lần đóng tiền và tổng tiền
                                print(f"[HISTORY] Đang đếm lịch sử thanh toán...")

                                try:
                                    payment_rows = page.query_selector_all('div.tableTra table tbody tr')
                                    total_payments = len(payment_rows)
                                    total_amount = 0

                                    print(f"[HISTORY] Số lần đóng tiền: {total_payments}")
                                    self.log_to_gui(f"  [TIEN] Số lần đóng: {total_payments}")

                                    for row_idx, row in enumerate(payment_rows, 1):
                                        amount_element = row.query_selector('span.number_vnd')
                                        if amount_element:
                                            amount_text = amount_element.inner_text().strip().replace(".", "").replace(",", "")
                                            try:
                                                amount = float(amount_text)
                                                total_amount += amount
                                                print(f"[HISTORY]   Lần {row_idx}: {amount:,.0f} VNĐ")
                                            except Exception as e:
                                                print(f"[HISTORY] [X] Lỗi đọc số tiền lần {row_idx}: {e}")

                                    print(f"[HISTORY] Tổng tiền đã đóng: {total_amount:,.0f} VNĐ")
                                    self.log_to_gui(f"  [TIEN] Tổng đã đóng: {total_amount:,.0f} VNĐ")
                                except Exception as e:
                                    print(f"[HISTORY] [X] Lỗi khi đọc bảng thanh toán: {e}")
                                    self.log_to_gui(f"  Lỗi đọc bảng thanh toán")
                                    import traceback
                                    traceback.print_exc()
                                    continue

                                # So sánh với số tiền vay trong Excel
                                loan_amount = item.get('loan_amount', 0)
                                print(f"[COMPARE] So sánh với Excel:")
                                print(f"[COMPARE]   - Số tiền vay (Excel): {loan_amount:,.0f} VNĐ")
                                print(f"[COMPARE]   - Tổng đã đóng (Web): {total_amount:,.0f} VNĐ")

                                # Xác định tình trạng thanh toán
                                payment_status = ""
                                if loan_amount > 0:
                                    if total_amount >= loan_amount:
                                        payment_status = "Đã đóng đầy đủ"
                                        print(f"[COMPARE] Khách hàng đã đóng tiền đầy đủ")
                                        self.log_to_gui(f"  Đã đóng tiền đầy đủ")
                                    else:
                                        remaining = loan_amount - total_amount
                                        payment_status = f"Chưa thanh toán {remaining:,.0f} VNĐ"
                                        print(f"[COMPARE] ⚠ Chưa thanh toán: {remaining:,.0f} VNĐ")
                                        self.log_to_gui(f"  Chưa thanh toán: {remaining:,.0f} VNĐ")
                                else:
                                    payment_status = "Không có thông tin số tiền vay"
                                    print(f"[COMPARE] ⚠ Không có thông tin số tiền vay trong Excel")
                                    self.log_to_gui(f"  Không có thông tin số tiền vay")

                                # Lưu kết quả
                                results.append({
                                    'contract_id': item['contract_id'],
                                    'status': payment_status
                                })

                                # Quay lại trang tra cứu
                                print(f"[NAVIGATE] Quay lại trang tra cứu...")
                                self.log_to_gui(f"  🔙 Quay lại trang tra cứu...")
                                page.goto("https://www.hdsaison.com.vn/vn/khach-hang/tra-cuu-khoan-vay.html")
                                page.wait_for_load_state("networkidle")
                                time.sleep(random.uniform(1.0, 1.5))

                                # Đóng popup sau khi quay lại
                                self._close_popup_if_exists(page)
                                print(f"[NAVIGATE] Đã sẵn sàng cho tra cứu tiếp theo")

                            else:
                                print(f"[COMPARE] [X] Thông tin KHÔNG khớp với Excel")
                                self.log_to_gui(f"  Thông tin KHÔNG khớp")
                                if not name_match:
                                    print(f"[COMPARE]   - Tên không khớp: Excel='{item.get('name', '')}' vs Web='{result_name}'")
                                    self.log_to_gui(f"    - Tên không khớp")
                                if not contract_match:
                                    print(f"[COMPARE]   - Số HĐ không khớp: Excel='{item['contract_id']}' vs Web='{result_contract}'")
                                    self.log_to_gui(f"    - Số HĐ không khớp")

                                # Lưu kết quả lỗi
                                results.append({
                                    'contract_id': item['contract_id'],
                                    'status': 'Thông tin không khớp'
                                })
                        else:
                            print(f"[RESULT] [X] Không lấy được thông tin từ bảng")
                            self.log_to_gui(f"  Không lấy được thông tin từ bảng")

                            # Lưu kết quả lỗi
                            results.append({
                                'contract_id': item['contract_id'],
                                'status': 'Không lấy được thông tin'
                            })

                    except Exception as e:
                        print(f"[ERROR] [X] LỖI: {str(e)}")
                        import traceback
                        traceback.print_exc()
                        self.log_to_gui(f"  Lỗi - {str(e)}")

                        # Lưu kết quả lỗi
                        results.append({
                            'contract_id': item['contract_id'],
                            'status': f'Lỗi: {str(e)}'
                        })

                    # Delay giữa các lần tra cứu
                    if idx < total_contracts:
                        print(f"[WAIT] Chờ {2} giây trước khi tra cứu tiếp...")
                        time.sleep(2)

                print(f"\n{'='*80}")
                print(f"HOÀN THÀNH TRA CỨU {total_contracts} HỢP ĐỒNG")
                print(f"{'='*80}\n")

                self.log_to_gui(f"\n{'='*60}")
                self.log_to_gui(f"Đã hoàn thành tra cứu {total_contracts} hợp đồng")

                # Đóng trình duyệt
                print("[INFO] Đóng trình duyệt...")
                context.close()
                browser.close()
                print("[INFO] Đã đóng trình duyệt")

            # Ghi kết quả vào Excel
            if results and self.contract_excel_path:
                print(f"\n{'='*80}")
                print("GHI KẾT QUẢ VÀO FILE EXCEL")
                print(f"{'='*80}")
                self.log_to_gui("\n[✍️] Đang ghi kết quả vào Excel...")

                try:
                    import openpyxl

                    # Mở file Excel
                    wb = openpyxl.load_workbook(self.contract_excel_path)
                    ws = wb.active

                    # Tìm cột cuối cùng
                    last_col = ws.max_column + 1

                    # Thêm tiêu đề cột
                    ws.cell(1, last_col, "Tình trạng thanh toán")
                    print(f"[EXCEL] Thêm cột 'Tình trạng thanh toán' tại cột {last_col}")

                    # Tạo dict để tra cứu nhanh
                    result_dict = {r['contract_id']: r['status'] for r in results}

                    # Tìm cột "ID Hợp đồng" để map kết quả
                    contract_col = None
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value and "ID Hợp đồng" in str(cell.value).strip():
                            contract_col = col_idx
                            break

                    if contract_col:
                        # Ghi kết quả vào từng dòng
                        for row_idx in range(2, ws.max_row + 1):
                            contract_id = ws.cell(row_idx, contract_col).value
                            if contract_id and str(contract_id).strip() in result_dict:
                                status = result_dict[str(contract_id).strip()]
                                ws.cell(row_idx, last_col, status)
                                print(f"[EXCEL] Dòng {row_idx}: {contract_id} -> {status}")

                        # Lưu file
                        wb.save(self.contract_excel_path)
                        print(f"[EXCEL] Đã lưu file: {self.contract_excel_path}")
                        self.log_to_gui(f"Đã ghi kết quả vào Excel")
                        self.log_to_gui(f"File: {self.contract_excel_path}")
                    else:
                        print(f"[EXCEL] [X] Không tìm thấy cột 'ID Hợp đồng'")
                        self.log_to_gui(f"Không tìm thấy cột 'ID Hợp đồng'")

                except Exception as e:
                    print(f"[EXCEL] [X] Lỗi khi ghi Excel: {e}")
                    import traceback
                    traceback.print_exc()
                    self.log_to_gui(f"Lỗi khi ghi Excel: {str(e)}")

        except Exception as e:
            error_msg = f"LỖI NGHIÊM TRỌNG: {str(e)}"
            print(f"\n{'='*80}")
            print(error_msg)
            print(f"{'='*80}\n")
            import traceback
            traceback.print_exc()
            self.log_to_gui(f"Lỗi: {str(e)}")

    # === CÁC HÀM XỬ LÝ CHO TAB ZALO ===

    def select_zalo_excel(self):
        """Chọn file Excel chứa dữ liệu khách hàng"""
        try:
            file_path = filedialog.askopenfilename(
                title="Chọn file Excel chứa dữ liệu khách hàng",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )

            if not file_path:
                return

            self.log_to_gui(f"[📄] Đang đọc file: {file_path}")

            # Đọc file Excel
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Đọc header (dòng đầu tiên)
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else "")

            # Mapping các cột phổ biến
            column_mapping = {}

            # Pass 1: Tìm các cột ưu tiên cao
            for idx, header in enumerate(headers):
                header_lower = header.lower()

                # Tên khách hàng
                if 'name' not in column_mapping:
                    if any(x in header_lower for x in ['tên kh', 'họ tên', 'name', 'customer', 'profile']):
                        column_mapping['name'] = idx

                # Số điện thoại (ưu tiên SĐT chính)
                if 'phone' not in column_mapping:
                    if 'sđt (chính)' in header_lower or 'sđt chính' in header_lower:
                        column_mapping['phone'] = idx

                # Địa chỉ thường trú (ưu tiên cao nhất)
                if 'address' not in column_mapping:
                    if 'địa chỉ thường trú' in header_lower:
                        column_mapping['address'] = idx

                # CCCD
                if 'cccd' not in column_mapping:
                    if any(x in header_lower for x in ['cccd', 'cmnd', 'số cccd', 'số cmnd']):
                        column_mapping['cccd'] = idx

                # Ngày sinh
                if 'dob' not in column_mapping:
                    if any(x in header_lower for x in ['ngày sinh', 'dob', 'sinh']):
                        column_mapping['dob'] = idx

                # ID Hợp đồng
                if 'contract_id' not in column_mapping:
                    if any(x in header_lower for x in ['id hợp đồng', 'hợp đồng', 'contract', 'mã hợp đồng']):
                        column_mapping['contract_id'] = idx

                # Giới tính
                if 'gender' not in column_mapping:
                    if any(x in header_lower for x in ['giới tính', 'gender', 'sex']):
                        column_mapping['gender'] = idx

                # Ghi chú (để kiểm tra trạng thái đã xử lý)
                if 'note' not in column_mapping:
                    if any(x in header_lower for x in ['ghi chú', 'ghi chu', 'note', 'notes', 'status', 'trạng thái']):
                        column_mapping['note'] = idx

            # Pass 2: Tìm các cột fallback (nếu chưa tìm thấy)
            for idx, header in enumerate(headers):
                header_lower = header.lower()

                # SĐT fallback
                if 'phone' not in column_mapping:
                    if any(x in header_lower for x in ['số điện thoại', 'phone', 'sđt']):
                        column_mapping['phone'] = idx

                # Địa chỉ tạm trú (fallback 1)
                if 'address' not in column_mapping:
                    if 'địa chỉ tạm trú' in header_lower:
                        column_mapping['address'] = idx

                # Địa chỉ công ty hoặc địa chỉ chung (fallback 2)
                if 'address' not in column_mapping:
                    if any(x in header_lower for x in ['địa chỉ công ty', 'địa chỉ', 'address']):
                        column_mapping['address'] = idx

            # Đọc dữ liệu
            self.zalo_customer_data = []
            processed_count = 0  # Đếm số khách hàng đã xử lý

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue  # Bỏ qua dòng trống

                customer = {}

                # Lấy dữ liệu theo mapping
                customer['name'] = str(row[column_mapping.get('name', 0)] or "").strip()
                customer['phone'] = str(row[column_mapping.get('phone', 1)] or "").strip()
                customer['address'] = str(row[column_mapping.get('address', 2)] or "").strip()
                customer['cccd'] = str(row[column_mapping.get('cccd', 3)] or "").strip()
                customer['dob'] = str(row[column_mapping.get('dob', 4)] or "").strip()
                customer['contract_id'] = str(row[column_mapping.get('contract_id', 5)] or "").strip()
                customer['gender'] = str(row[column_mapping.get('gender', 6)] or "").strip()

                # Đọc cột "Ghi chú" nếu có
                note = ""
                if 'note' in column_mapping:
                    note = str(row[column_mapping.get('note')] or "").strip()
                customer['note'] = note

                # Kiểm tra trạng thái đã xử lý
                customer['is_processed'] = False
                if note:
                    # Kiểm tra nếu đã kết bạn thành công, đã là bạn bè, hoặc đã gửi lời mời
                    if 'Kết bạn thành công' in note or 'Đã là bạn bè' in note or 'Đã gửi lời mời trước đó' in note or 'Gửi tin nhắn thành công' in note:
                        customer['is_processed'] = True
                        processed_count += 1

                # Chỉ thêm nếu có ít nhất tên hoặc số điện thoại
                if customer['name'] or customer['phone']:
                    self.zalo_customer_data.append(customer)

            wb.close()

            # Cập nhật UI
            self.zalo_excel_path = file_path
            filename = file_path.split('/')[-1].split('\\')[-1]
            self.zalo_file_label.configure(
                text=f"[OK] {filename}",
                text_color="green"
            )

            # Hiển thị số lượng khách hàng và trạng thái
            unprocessed_count = len(self.zalo_customer_data) - processed_count
            status_text = f"Tổng: {len(self.zalo_customer_data)}"
            if processed_count > 0:
                status_text += f" ([OK] {processed_count} đã xử lý, [DANG XU LY] {unprocessed_count} chưa xử lý)"

            self.zalo_customer_count_label.configure(
                text=status_text,
                text_color="#28A745"
            )

            self.log_to_gui(f"[OK] Đã tải {len(self.zalo_customer_data)} khách hàng từ Excel")
            if processed_count > 0:
                self.log_to_gui(f"   [DU LIEU] Trạng thái: {processed_count} đã xử lý, {unprocessed_count} chưa xử lý")
            self.log_to_gui(f"   [DANH SACH] Các cột được nhận diện: {', '.join(column_mapping.keys())}")

        except ImportError:
            self.log_to_gui("[LỖI] Lỗi: Chưa cài đặt openpyxl. Chạy: pip install openpyxl")
            messagebox.showerror("Lỗi", "Chưa cài đặt thư viện openpyxl!", parent=self)
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi đọc file Excel: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể đọc file Excel:\n{str(e)}", parent=self)

    def save_results_to_excel(self, results):
        """
        Lưu kết quả kết bạn vào file Excel

        Args:
            results: List of dict với keys: phone, name, zalo_name, status
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            from datetime import datetime

            # Load workbook
            wb = load_workbook(self.zalo_excel_path)
            ws = wb.active

            # Tìm hoặc tạo cột "Ghi chú" và "Tên Zalo"
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else "")

            # Tìm cột "Ghi chú"
            note_col_idx = None
            for idx, header in enumerate(headers):
                if header.lower() in ['ghi chú', 'ghi chu', 'note', 'notes', 'status', 'trạng thái']:
                    note_col_idx = idx
                    break

            # Nếu không có cột "Ghi chú", tạo mới
            if note_col_idx is None:
                note_col_idx = len(headers)
                ws.cell(row=1, column=note_col_idx + 1, value="Ghi chú")
                headers.append("Ghi chú")

            # Tìm hoặc tạo cột "Tên Zalo"
            zalo_name_col_idx = None
            for idx, header in enumerate(headers):
                if header.lower() in ['tên zalo', 'ten zalo', 'zalo name', 'zalo']:
                    zalo_name_col_idx = idx
                    break

            # Nếu không có cột "Tên Zalo", tạo mới
            if zalo_name_col_idx is None:
                zalo_name_col_idx = len(headers)
                ws.cell(row=1, column=zalo_name_col_idx + 1, value="Tên Zalo")

            # Tìm cột số điện thoại để mapping
            phone_col_idx = None
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if any(x in header_lower for x in ['sđt', 'phone', 'số điện thoại', 'điện thoại']):
                    phone_col_idx = idx
                    break

            if phone_col_idx is None:
                self.log_to_gui("[⚠️] Không tìm thấy cột số điện thoại trong Excel")
                return

            # Tạo mapping từ phone -> result
            phone_to_result = {}
            for result in results:
                phone = str(result.get('phone', '')).strip()
                if phone:
                    phone_to_result[phone] = result

            # Định nghĩa màu sắc
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Cập nhật từng dòng
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                phone_cell = ws.cell(row=row_idx, column=phone_col_idx + 1)
                phone = str(phone_cell.value or "").strip()

                if phone in phone_to_result:
                    result = phone_to_result[phone]
                    status = result.get('status', 'unknown')
                    zalo_name = result.get('zalo_name', '')

                    # Tạo message ghi chú và chọn màu
                    if status == 'success':
                        note = f"[OK] Kết bạn thành công ({timestamp})"
                        fill_color = green_fill
                    elif status == 'already_friend':
                        note = f"[OK] Đã là bạn bè ({timestamp})"
                        fill_color = green_fill
                    elif status == 'already_sent':
                        note = f"[⚠️] Đã gửi lời mời trước đó ({timestamp})"
                        fill_color = yellow_fill
                    elif status == 'failed':
                        note = f"[LỖI] Kết bạn thất bại ({timestamp})"
                        fill_color = red_fill
                    else:
                        note = f"❓ Không xác định ({timestamp})"
                        fill_color = None

                    # Ghi vào cột "Ghi chú"
                    note_cell = ws.cell(row=row_idx, column=note_col_idx + 1, value=note)
                    if fill_color:
                        note_cell.fill = fill_color

                    # Ghi vào cột "Tên Zalo" nếu có
                    if zalo_name:
                        zalo_cell = ws.cell(row=row_idx, column=zalo_name_col_idx + 1, value=zalo_name)
                        if fill_color:
                            zalo_cell.fill = fill_color

                    updated_count += 1

            # Lưu file
            wb.save(self.zalo_excel_path)
            wb.close()

            self.log_to_gui(f"   [✍️] Đã cập nhật {updated_count}/{len(results)} dòng trong Excel")

        except Exception as e:
            raise Exception(f"Lỗi khi lưu Excel: {str(e)}")

    def save_message_template(self):
        """Lưu kịch bản tin nhắn vào file với tên riêng"""
        try:
            # Hỏi tên kịch bản với custom dialog
            dialog = InputDialog(
                self,
                "Lưu kịch bản",
                "Nhập tên cho kịch bản:"
            )
            template_name = dialog.get_input()
            
            if not template_name:
                return
            
            # Làm sạch tên file
            template_name = template_name.strip()
            if not template_name:
                messagebox.showwarning("Cảnh báo", "Tên kịch bản không được để trống!", parent=self)
                return
            
            template = self.zalo_message_template.get("1.0", "end-1c")

            # Lưu vào file JSON với tên riêng
            templates_dir = os.path.join(self.app_data_dir, "message_templates")
            if not os.path.exists(templates_dir):
                os.makedirs(templates_dir)
            
            template_file = os.path.join(templates_dir, f"{template_name}.json")
            with open(template_file, "w", encoding="utf-8") as f:
                json.dump({"name": template_name, "template": template}, f, ensure_ascii=False, indent=2)

            self.log_to_gui(f"[OK] Đã lưu kịch bản '{template_name}'")
            messagebox.showinfo(
                "Thành công",
                f"Đã lưu kịch bản '{template_name}'!",
                parent=self
            )
            
            # Refresh danh sách
            self.refresh_template_list()
            
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi lưu kịch bản: {str(e)}")
            messagebox.showerror(
                "Lỗi",
                f"Không thể lưu kịch bản: {str(e)}",
                parent=self
            )

    def load_message_template(self, template_name=None):
        """Load kịch bản tin nhắn từ file theo tên"""
        try:
            if not template_name:
                return ""
            
            templates_dir = os.path.join(self.app_data_dir, "message_templates")
            template_file = os.path.join(templates_dir, f"{template_name}.json")
            
            if os.path.exists(template_file):
                with open(template_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get("template", "")
            return ""
        except Exception as e:
            print(f"Lỗi khi load kịch bản: {str(e)}")
            return ""
    
    def get_saved_template_names(self):
        """Lấy danh sách tên các kịch bản đã lưu"""
        try:
            templates_dir = os.path.join(self.app_data_dir, "message_templates")
            if not os.path.exists(templates_dir):
                return []
            
            template_files = [f for f in os.listdir(templates_dir) if f.endswith('.json')]
            template_names = [f[:-5] for f in template_files]  # Bỏ .json
            return sorted(template_names)
        except Exception as e:
            print(f"Lỗi khi lấy danh sách kịch bản: {str(e)}")
            return []
    
    def on_template_selected(self, selected_name):
        """Callback khi chọn kịch bản từ dropdown"""
        try:
            if selected_name == "(Chưa có kịch bản)":
                return
            
            template = self.load_message_template(selected_name)
            if template:
                # Xóa nội dung cũ và insert mới
                self.zalo_message_template.delete("1.0", "end")
                self.zalo_message_template.insert("1.0", template)
                self.log_to_gui(f"[OK] Đã load kịch bản '{selected_name}'")
            else:
                messagebox.showwarning(
                    "Cảnh báo",
                    f"Không thể load kịch bản '{selected_name}'",
                    parent=self
                )
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi load kịch bản: {str(e)}")
    
    def refresh_template_list(self):
        """Refresh danh sách kịch bản trong dropdown"""
        try:
            self.template_names = self.get_saved_template_names()
            template_options = self.template_names if self.template_names else ["(Chưa có kịch bản)"]
            
            self.template_dropdown.configure(values=template_options)
            if self.template_names:
                self.template_dropdown.set(self.template_names[0])
            else:
                self.template_dropdown.set("(Chưa có kịch bản)")
            
            self.log_to_gui("[OK] Đã cập nhật danh sách kịch bản")
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi refresh danh sách: {str(e)}")
    
    def delete_message_template(self):
        """Xoá kịch bản tin nhắn đã lưu"""
        try:
            # Lấy kịch bản đang chọn
            selected_template = self.template_dropdown.get()
            
            if not selected_template or selected_template == "(Chưa có kịch bản)":
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn kịch bản cần xoá!", parent=self)
                return
            
            # Xác nhận xoá
            confirm = messagebox.askyesno(
                "Xác nhận xoá",
                f"Bạn có chắc muốn xoá kịch bản '{selected_template}'?\n\nHành động này không thể hoàn tác!",
                parent=self,
                icon='warning'
            )
            
            if not confirm:
                return
            
            # Xoá file
            templates_dir = os.path.join(self.app_data_dir, "message_templates")
            template_file = os.path.join(templates_dir, f"{selected_template}.json")
            
            if os.path.exists(template_file):
                os.remove(template_file)
                self.log_to_gui(f"[OK] Đã xoá kịch bản '{selected_template}'")
                
                # Refresh danh sách
                self.refresh_template_list()
                
                # Load kịch bản mới hoặc default
                if self.template_names:
                    new_template = self.load_message_template(self.template_names[0])
                    if new_template:
                        self.zalo_message_template.delete("1.0", "end")
                        self.zalo_message_template.insert("1.0", new_template)
                else:
                    # Nếu không còn kịch bản nào, load default
                    self.zalo_message_template.delete("1.0", "end")
                    self._insert_default_template()
                
                messagebox.showinfo("Thành công", f"Đã xoá kịch bản '{selected_template}'", parent=self)
            else:
                messagebox.showerror("Lỗi", f"Không tìm thấy file kịch bản '{selected_template}'", parent=self)
                
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi xoá kịch bản: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể xoá kịch bản: {str(e)}", parent=self)
    
    def _insert_default_template(self):
        """Insert template mặc định vào textbox"""
        default_template = """{name} ơi, HD SAISON có ưu đãi 35 triệu cho khách hàng thân thiết!
Chào {gender} {name},
Chúc mừng {gender} đã thanh toán hoàn tất khoản vay số {contract_id}! HD SAISON rất trân trọng uy tín và sự đồng hành của Quý khách.
Để tri ân, HD SAISON xin gửi tặng {gender} {name} ưu đãi vay tiền mặt ĐỘC QUYỀN dành cho khách hàng cũ:
Khoản vay lên đến 35 TRIỆU ĐỒNG.
Thời hạn vay lên đến 36 tháng.
Lãi suất chỉ từ 1.67%/tháng.
Giải ngân NHANH CHÓNG trong 1 giờ, không cần thế chấp.
Miễn phí hoàn toàn với khách hàng thân thiết.

{gender} {name} vui lòng nhắn tin lại cho tôi để được hỗ trợ hoặc gọi Hotline: 1900 6249 (máy lẻ 0) để được tư vấn chi tiết.

Xin cảm ơn và chúc {gender} {name} luôn thành công!"""
        self.zalo_message_template.insert("1.0", default_template)

    def save_message_results_to_excel(self, details):
        """
        Lưu kết quả gửi tin nhắn vào file Excel

        Args:
            details: List of dict với keys: phone, name, status, friend_status
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            from datetime import datetime

            # Load workbook
            wb = load_workbook(self.zalo_excel_path)
            ws = wb.active

            # Tìm hoặc tạo cột "Ghi chú"
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else "")

            # Tìm cột "Ghi chú"
            note_col_idx = None
            for idx, header in enumerate(headers):
                if header.lower() in ['ghi chú', 'ghi chu', 'note', 'notes', 'status', 'trạng thái']:
                    note_col_idx = idx
                    break

            # Nếu không có cột "Ghi chú", tạo mới
            if note_col_idx is None:
                note_col_idx = len(headers)
                ws.cell(row=1, column=note_col_idx + 1, value="Ghi chú")
                headers.append("Ghi chú")

            # Tìm hoặc tạo cột "Trạng thái kết bạn"
            friend_status_col_idx = None
            for idx, header in enumerate(headers):
                if header.lower() in ['trạng thái kết bạn', 'trang thai ket ban', 'friend status']:
                    friend_status_col_idx = idx
                    break

            # Nếu không có cột "Trạng thái kết bạn", tạo mới
            if friend_status_col_idx is None:
                friend_status_col_idx = len(headers)
                ws.cell(row=1, column=friend_status_col_idx + 1, value="Trạng thái kết bạn")

            # Tìm cột số điện thoại để mapping
            phone_col_idx = None
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if any(x in header_lower for x in ['sđt', 'phone', 'số điện thoại', 'điện thoại']):
                    phone_col_idx = idx
                    break

            if phone_col_idx is None:
                self.log_to_gui("[⚠️] Không tìm thấy cột số điện thoại trong Excel")
                return

            # Tạo mapping từ phone -> detail
            phone_to_detail = {}
            for detail in details:
                phone = str(detail.get('phone', '')).strip()
                if phone:
                    phone_to_detail[phone] = detail

            # Định nghĩa màu sắc
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # Cập nhật từng dòng
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                phone_cell = ws.cell(row=row_idx, column=phone_col_idx + 1)
                phone = str(phone_cell.value or "").strip()

                if phone in phone_to_detail:
                    detail = phone_to_detail[phone]
                    status = detail.get('status', 'unknown')
                    friend_status = detail.get('friend_status', None)

                    # Tạo message ghi chú và chọn màu
                    if status == 'success':
                        note = f"[OK] Gửi tin nhắn thành công ({timestamp})"
                        fill_color = green_fill
                    elif status == 'not_registered':
                        note = f"[⚠️] Số chưa đăng ký hoặc không cho phép tìm kiếm ({timestamp})"
                        fill_color = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Màu vàng nhạt
                    elif status == 'not_found':
                        note = f"[⚠️] Không tìm thấy kết quả ({timestamp})"
                        fill_color = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Màu vàng nhạt
                    elif status == 'failed':
                        note = f"[LỖI] Gửi tin nhắn thất bại ({timestamp})"
                        fill_color = red_fill
                    elif status == 'no_phone':
                        note = f"[⚠️] Không có số điện thoại ({timestamp})"
                        fill_color = gray_fill
                    elif status == 'error':
                        note = f"[LỖI] Lỗi khi gửi tin nhắn ({timestamp})"
                        fill_color = red_fill
                    else:
                        note = f"❓ Không xác định ({timestamp})"
                        fill_color = None

                    # Ghi vào cột "Ghi chú"
                    note_cell = ws.cell(row=row_idx, column=note_col_idx + 1, value=note)
                    if fill_color:
                        note_cell.fill = fill_color

                    # Ghi vào cột "Trạng thái kết bạn"
                    if friend_status:
                        if friend_status == 'friend':
                            friend_status_text = "👥 Bạn bè"
                        elif friend_status == 'stranger':
                            friend_status_text = "[NGUOI DUNG] Người lạ"
                        else:
                            friend_status_text = "❓ Không xác định"

                        friend_status_cell = ws.cell(row=row_idx, column=friend_status_col_idx + 1, value=friend_status_text)
                        if fill_color:
                            friend_status_cell.fill = fill_color

                    updated_count += 1

            # Lưu file
            wb.save(self.zalo_excel_path)
            wb.close()

            self.log_to_gui(f"   [✍️] Đã cập nhật {updated_count}/{len(details)} dòng trong Excel")

        except Exception as e:
            raise Exception(f"Lỗi khi lưu Excel: {str(e)}")

    def send_bulk_messages(self):
        """Gửi tin nhắn hàng loạt cho khách hàng"""
        # Kiểm tra dữ liệu
        if not self.zalo_customer_data:
            messagebox.showwarning(
                "Cảnh báo",
                "Vui lòng chọn file Excel chứa dữ liệu khách hàng trước!",
                parent=self
            )
            return

        # Lọc khách hàng theo checkbox
        skip_sent = self.skip_sent_messages_var.get()
        customers_to_process = []

        if skip_sent:
            # Chỉ lấy khách hàng chưa gửi tin nhắn thành công
            customers_to_process = [c for c in self.zalo_customer_data if not c.get('is_processed', False)]
            skipped_count = len(self.zalo_customer_data) - len(customers_to_process)

            if len(customers_to_process) == 0:
                messagebox.showinfo(
                    "Thông báo",
                    "Tất cả khách hàng đã được gửi tin nhắn thành công!\n\n"
                    "Bỏ tick '☑️ Bỏ qua khách hàng đã gửi tin nhắn thành công' nếu muốn gửi lại.",
                    parent=self
                )
                return
        else:
            # Xử lý tất cả
            customers_to_process = self.zalo_customer_data
            skipped_count = 0

        # Lấy template
        template = self.zalo_message_template.get("1.0", "end-1c").strip()
        if not template:
            messagebox.showwarning(
                "Cảnh báo",
                "Vui lòng nhập kịch bản tin nhắn!",
                parent=self
            )
            return

        # Xác nhận
        confirm_msg = f"Bạn có chắc muốn gửi tin nhắn đến {len(customers_to_process)} khách hàng?"
        if skipped_count > 0:
            confirm_msg += f"\n\n(Bỏ qua {skipped_count} khách hàng đã gửi thành công)"
        confirm_msg += "\n\nLưu ý: Bạn cần đăng nhập Zalo trước!"

        result = messagebox.askyesno(
            "Xác nhận",
            confirm_msg,
            parent=self
        )

        if not result:
            return

        # Tự động lưu kịch bản trước khi gửi
        try:
            template_file = os.path.join(self.app_data_dir, "message_template.json")
            with open(template_file, "w", encoding="utf-8") as f:
                json.dump({"template": template}, f, ensure_ascii=False, indent=2)
            self.log_to_gui("[💾] Đã tự động lưu kịch bản tin nhắn")
        except Exception as e:
            self.log_to_gui(f"[⚠️] Không thể tự động lưu kịch bản: {str(e)}")

        # Lưu danh sách cần xử lý vào biến tạm
        self.current_customers_to_process = customers_to_process

        # Reset trạng thái tạm dừng
        self.is_paused = False

        # Enable nút tạm dừng cho gửi tin nhắn
        self.message_pause_button.configure(text="Tạm dừng", fg_color="#6C757D", hover_color="#5A6268")
        self.message_pause_button.configure(state="normal")

        # Cũng enable nút tạm dừng chung (để tương thích)
        self.zalo_pause_button.configure(text="Tạm dừng", fg_color="#6C757D", hover_color="#5A6268")
        self.zalo_pause_button.configure(state="normal")

        # Chạy trong thread riêng
        thread = threading.Thread(
            target=self._run_send_bulk_messages,
            args=(template,),
            daemon=True
        )
        thread.start()

    def add_friends_bulk(self):
        """Kết bạn hàng loạt với khách hàng"""
        # Kiểm tra dữ liệu
        if not self.zalo_customer_data:
            messagebox.showwarning(
                "Cảnh báo",
                "Vui lòng chọn file Excel chứa dữ liệu khách hàng trước!",
                parent=self
            )
            return

        # Lọc khách hàng theo checkbox
        skip_processed = self.skip_processed_var.get()
        customers_to_process = []

        if skip_processed:
            # Chỉ lấy khách hàng chưa kết bạn thành công
            customers_to_process = [c for c in self.zalo_customer_data if not c.get('is_processed', False)]
            skipped_count = len(self.zalo_customer_data) - len(customers_to_process)

            if len(customers_to_process) == 0:
                messagebox.showinfo(
                    "Thông báo",
                    "Tất cả khách hàng đã được kết bạn thành công!\n\n"
                    "Bỏ tick '☑️ Bỏ qua khách hàng đã kết bạn thành công' nếu muốn gửi lại.",
                    parent=self
                )
                return
        else:
            # Xử lý tất cả
            customers_to_process = self.zalo_customer_data
            skipped_count = 0

        # Đếm số khách hàng có số điện thoại
        phone_count = sum(1 for c in customers_to_process if c.get('phone'))

        if phone_count == 0:
            messagebox.showwarning(
                "Cảnh báo",
                "Không tìm thấy số điện thoại nào trong dữ liệu chưa xử lý!",
                parent=self
            )
            return

        # Xác nhận
        confirm_msg = f"Bạn có chắc muốn gửi lời mời kết bạn đến {phone_count} số điện thoại?"
        if skipped_count > 0:
            confirm_msg += f"\n\n(Bỏ qua {skipped_count} khách hàng đã kết bạn thành công)"
        confirm_msg += "\n\nLưu ý: Bạn cần đăng nhập Zalo trước!"

        result = messagebox.askyesno(
            "Xác nhận",
            confirm_msg,
            parent=self
        )

        if not result:
            return

        # Lưu danh sách cần xử lý vào biến tạm
        self.current_customers_to_process = customers_to_process

        # Reset trạng thái tạm dừng
        self.is_paused = False
        self.zalo_pause_button.configure(text="Tạm dừng", fg_color="#6C757D", hover_color="#5A6268")
        self.zalo_pause_button.configure(state="normal")  # Enable nút tạm dừng

        # Chạy trong thread riêng
        thread = threading.Thread(
            target=self._run_add_friends_bulk,
            daemon=True
        )
        thread.start()

    def toggle_pause(self):
        """Chuyển đổi trạng thái tạm dừng/tiếp tục"""
        self.is_paused = not self.is_paused

        if self.is_paused:
            # Cập nhật cả 2 nút
            self.zalo_pause_button.configure(text="Tiếp tục", fg_color="#28A745", hover_color="#218838")
            self.message_pause_button.configure(text="Tiếp tục", fg_color="#28A745", hover_color="#218838")
            self.log_to_gui("[TAM DUNG] Đã tạm dừng - Click 'Tiếp tục' để chạy tiếp")
        else:
            # Cập nhật cả 2 nút
            self.zalo_pause_button.configure(text="Tạm dừng", fg_color="#6C757D", hover_color="#5A6268")
            self.message_pause_button.configure(text="Tạm dừng", fg_color="#6C757D", hover_color="#5A6268")
            self.log_to_gui("[TIEP TUC] Tiếp tục chạy...")



    def _run_send_bulk_messages(self, template):
        """Thread worker để gửi tin nhắn hàng loạt"""
        try:
            import time
            import zalo_logic
            import zalo_automation

            # Kiểm tra đã chọn tài khoản chưa
            if not self.current_account_id:
                self.log_to_gui("[LỖI] Vui lòng chọn tài khoản trước!")
                messagebox.showerror(
                    "Lỗi",
                    "Vui lòng chọn tài khoản Zalo trước khi gửi tin nhắn!",
                    parent=self
                )
                return

            self.log_to_gui("[CHO] Đang khởi tạo Zalo...")

            # Lấy session manager cho tài khoản đã chọn
            session_manager = self.account_manager.get_session_manager(self.current_account_id)
            success, p, context, page = session_manager.login_with_session(max_wait_time=60)

            if not success:
                self.log_to_gui("[LỖI] Không thể đăng nhập Zalo. Vui lòng đăng nhập thủ công trước!")
                if context:
                    context.close()
                if p:
                    p.stop()
                return

            self.log_to_gui("[OK] Đã đăng nhập Zalo")

            # Sử dụng danh sách đã lọc
            customers = getattr(self, 'current_customers_to_process', self.zalo_customer_data)
            self.log_to_gui(f"[TAI LEN] Bắt đầu gửi tin nhắn đến {len(customers)} khách hàng...")

            # Tạo automation instance
            automation = zalo_automation.ZaloAutomation(page)

            # Gửi tin nhắn hàng loạt (với hỗ trợ pause và kiểm tra trạng thái bạn bè)
            result = automation.send_bulk_messages(
                customers,
                template,
                callback=self.log_to_gui,
                delay=3,
                is_paused_func=lambda: self.is_paused,
                check_friend_status=True  # Kiểm tra và ghi nhận trạng thái bạn bè/người lạ
            )

            # Hiển thị kết quả
            self.log_to_gui("\n" + "="*50)
            self.log_to_gui(f"[OK] HOÀN TẤT GỬI TIN NHẮN HÀNG LOẠT")
            self.log_to_gui(f"   - Thành công: {result['success']}")
            self.log_to_gui(f"   - Thất bại: {result['failed']}")
            if result['errors']:
                self.log_to_gui(f"\n[LỖI] Các lỗi:")
                for error in result['errors'][:10]:  # Hiển thị tối đa 10 lỗi
                    self.log_to_gui(f"   - {error}")
            self.log_to_gui("="*50)

            # Ghi kết quả vào file Excel
            if self.zalo_excel_path and result.get('details'):
                self.log_to_gui("\n[💾] Đang lưu kết quả vào file Excel...")
                try:
                    self.save_message_results_to_excel(result['details'])
                    self.log_to_gui("[OK] Đã lưu kết quả vào file Excel")
                except Exception as e:
                    self.log_to_gui(f"[⚠️] Không thể lưu kết quả vào Excel: {str(e)}")

            # Disable nút điều khiển
            self.zalo_pause_button.configure(state="disabled")
            self.message_pause_button.configure(state="disabled")

            # Giữ trình duyệt mở
            self.log_to_gui("\n[ℹ️] Trình duyệt vẫn mở, bạn có thể tiếp tục sử dụng Zalo")
            self.log_to_gui("[⚠️] Đóng cửa sổ trình duyệt khi hoàn tất")

            # Chờ người dùng đóng
            try:
                import random
                while not page.is_closed():
                    time.sleep(random.uniform(0.8, 1.2))
            except:
                pass

            # Cleanup
            try:
                if context:
                    context.close()
                if p:
                    p.stop()
            except:
                pass

        except ImportError as e:
            self.log_to_gui(f"[LỖI] Lỗi import: {str(e)}")
            self.log_to_gui("[📝] Vui lòng cài đặt: pip install playwright")
            self.zalo_pause_button.configure(state="disabled")
            self.message_pause_button.configure(state="disabled")
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi gửi tin nhắn: {str(e)}")
            self.zalo_pause_button.configure(state="disabled")
            self.message_pause_button.configure(state="disabled")

    def _run_add_friends_bulk(self):
        """Thread worker để kết bạn hàng loạt"""
        try:
            import time
            import zalo_logic
            import zalo_automation

            # Kiểm tra đã chọn tài khoản chưa
            if not self.current_account_id:
                self.log_to_gui("[LỖI] Vui lòng chọn tài khoản trước!")
                messagebox.showerror(
                    "Lỗi",
                    "Vui lòng chọn tài khoản Zalo trước khi kết bạn!",
                    parent=self
                )
                return

            # Lấy lời chào từ textbox
            greeting_template = self.friend_greeting_textbox.get("1.0", "end-1c").strip()
            if not greeting_template:
                greeting_template = "Xin chào, mình là {my_name} bên công ty tài chính HDSAISON, vui lòng đồng ý kết bạn để được hỗ trợ hợp đồng {contract_id}"
                self.log_to_gui("[⚠️] Lời chào trống, sử dụng lời chào mặc định")

            self.log_to_gui("[CHO] Đang khởi tạo Zalo...")

            # Lấy session manager cho tài khoản đã chọn
            session_manager = self.account_manager.get_session_manager(self.current_account_id)
            success, p, context, page = session_manager.login_with_session(max_wait_time=60)

            if not success:
                self.log_to_gui("[LỖI] Không thể đăng nhập Zalo. Vui lòng đăng nhập thủ công trước!")
                if context:
                    context.close()
                if p:
                    p.stop()
                return

            self.log_to_gui("[OK] Đã đăng nhập Zalo")

            # Sử dụng danh sách đã lọc và lọc thêm khách hàng có số điện thoại
            customers = getattr(self, 'current_customers_to_process', self.zalo_customer_data)
            customers_with_phone = [c for c in customers if c.get('phone')]
            self.log_to_gui(f"[THEM] Bắt đầu kết bạn với {len(customers_with_phone)} số điện thoại...")

            # Tạo automation instance
            automation = zalo_automation.ZaloAutomation(page)

            # Lấy tên Zalo của tài khoản đang đăng nhập (truyền session_manager để lưu tên)
            my_zalo_name = automation.get_my_zalo_name(session_manager)
            self.log_to_gui(f"[NGUOI DUNG] Tài khoản Zalo: {my_zalo_name}")

            # Kết bạn từng người
            success_count = 0
            failed_count = 0
            already_sent_count = 0
            already_friend_count = 0  # Đếm số người đã là bạn bè
            results = []  # Lưu kết quả chi tiết

            for idx, customer in enumerate(customers_with_phone, 1):
                # Kiểm tra tạm dừng
                import random
                while self.is_paused:
                    time.sleep(random.uniform(0.4, 0.6))

                phone = customer.get('phone', '').strip()
                name = customer.get('name', 'N/A')
                contract_id = customer.get('contract_id', '')

                # Chuyển đổi giới tính từ Nam/Nữ sang anh/chị
                gender_raw = customer.get('gender', '').strip()
                gender_pronoun = ''
                if gender_raw:
                    gender_lower = gender_raw.lower()
                    if 'nam' in gender_lower or 'male' in gender_lower:
                        gender_pronoun = 'anh'
                    elif 'nữ' in gender_lower or 'nv' in gender_lower or 'female' in gender_lower:
                        gender_pronoun = 'chị'
                    else:
                        gender_pronoun = 'anh/chị'
                else:
                    gender_pronoun = 'anh/chị'

                # Format greeting template với các biến
                formatted_greeting = greeting_template.format(
                    name=name,
                    phone=phone,
                    contract_id=contract_id,
                    my_name=my_zalo_name,
                    gender=gender_pronoun,
                    address=customer.get('address', ''),
                    cccd=customer.get('cccd', ''),
                    dob=customer.get('dob', '')
                )

                self.log_to_gui(f"\n{'='*60}")
                self.log_to_gui(f"[THEM] [{idx}/{len(customers_with_phone)}] Đang kết bạn: {name} ({phone})")

                # Gọi hàm kết bạn (trả về tuple: success/status, display_name)
                # Truyền formatted_greeting thay vì greeting_template
                result, display_name = automation.add_friend_by_phone(phone, contract_id, my_zalo_name, formatted_greeting)

                # Xử lý kết quả
                if result == "already_sent":
                    # Đã gửi lời mời trước đó
                    already_sent_count += 1
                    result_msg = f"[⚠️] [{idx}/{len(customers_with_phone)}] Đã gửi lời mời trước đó: {phone}"
                    if display_name:
                        result_msg += f" - Tên Zalo: {display_name}"
                    self.log_to_gui(result_msg)
                    results.append({
                        'phone': phone,
                        'name': name,
                        'zalo_name': display_name,
                        'status': 'already_sent'
                    })
                elif result == "already_friend":
                    # Đã là bạn bè
                    already_friend_count += 1
                    result_msg = f"[OK] [{idx}/{len(customers_with_phone)}] Đã là bạn bè: {phone}"
                    if display_name:
                        result_msg += f" - Tên Zalo: {display_name}"
                    self.log_to_gui(result_msg)
                    results.append({
                        'phone': phone,
                        'name': name,
                        'zalo_name': display_name,
                        'status': 'already_friend'
                    })
                elif result:
                    # Thành công
                    success_count += 1
                    result_msg = f"[OK] [{idx}/{len(customers_with_phone)}] Thành công: {phone}"
                    if display_name:
                        result_msg += f" - Tên Zalo: {display_name}"
                    self.log_to_gui(result_msg)
                    results.append({
                        'phone': phone,
                        'name': name,
                        'zalo_name': display_name,
                        'status': 'success'
                    })
                else:
                    # Thất bại
                    failed_count += 1
                    self.log_to_gui(f"[LỖI] [{idx}/{len(customers_with_phone)}] Thất bại: {phone}")
                    results.append({
                        'phone': phone,
                        'name': name,
                        'zalo_name': None,
                        'status': 'failed'
                    })

                # Đóng modal sau mỗi lần kết bạn (thành công hoặc thất bại)
                self.log_to_gui("[DANG XU LY] Đóng modal và chuẩn bị kết bạn tiếp...")
                automation.close_modal_after_add_friend()

                # Delay giữa các lần kết bạn (random 2.5-3.5s)
                if idx < len(customers_with_phone):
                    import random
                    delay = random.uniform(2.5, 3.5)
                    self.log_to_gui(f"[CHO] Chờ {delay:.1f} giây trước khi kết bạn tiếp...")
                    time.sleep(delay)

            # Hiển thị kết quả
            self.log_to_gui("\n" + "="*60)
            self.log_to_gui(f"[OK] HOÀN TẤT KẾT BẠN HÀNG LOẠT")
            self.log_to_gui(f"   - Tổng số: {len(customers_with_phone)}")
            self.log_to_gui(f"   - Thành công: {success_count}")
            self.log_to_gui(f"   - Đã là bạn bè: {already_friend_count}")
            self.log_to_gui(f"   - Đã gửi lời mời trước đó: {already_sent_count}")
            self.log_to_gui(f"   - Thất bại: {failed_count}")

            # Hiển thị danh sách thành công
            if success_count > 0:
                self.log_to_gui(f"\n[DANH SACH] Danh sách kết bạn thành công ({success_count}):")
                for idx, result in enumerate([r for r in results if r['status'] == 'success'], 1):
                    zalo_name_info = f" - Zalo: {result['zalo_name']}" if result['zalo_name'] else ""
                    self.log_to_gui(f"   {idx}. {result['name']} ({result['phone']}){zalo_name_info}")

            # Hiển thị danh sách đã gửi lời mời trước đó
            if already_sent_count > 0:
                self.log_to_gui(f"\n[⚠️] Danh sách đã gửi lời mời trước đó ({already_sent_count}):")
                for idx, result in enumerate([r for r in results if r['status'] == 'already_sent'], 1):
                    zalo_name_info = f" - Zalo: {result['zalo_name']}" if result['zalo_name'] else ""
                    self.log_to_gui(f"   {idx}. {result['name']} ({result['phone']}){zalo_name_info}")

            # Hiển thị danh sách thất bại
            if failed_count > 0:
                self.log_to_gui(f"\n[LỖI] Danh sách kết bạn thất bại ({failed_count}):")
                for idx, result in enumerate([r for r in results if r['status'] == 'failed'], 1):
                    self.log_to_gui(f"   {idx}. {result['name']} ({result['phone']})")

            self.log_to_gui("="*60)

            # Ghi kết quả vào file Excel
            if self.zalo_excel_path and results:
                self.log_to_gui("\n[💾] Đang lưu kết quả vào file Excel...")
                try:
                    self.save_results_to_excel(results)
                    self.log_to_gui("[OK] Đã lưu kết quả vào file Excel")
                except Exception as e:
                    self.log_to_gui(f"[⚠️] Không thể lưu kết quả vào Excel: {str(e)}")

            # Disable nút điều khiển
            self.zalo_pause_button.configure(state="disabled")

            # Giữ trình duyệt mở
            self.log_to_gui("\n[ℹ️] Trình duyệt vẫn mở, bạn có thể tiếp tục sử dụng Zalo")
            self.log_to_gui("[⚠️] Đóng cửa sổ trình duyệt khi hoàn tất")

            # Chờ người dùng đóng
            try:
                import random
                while not page.is_closed():
                    time.sleep(random.uniform(0.8, 1.2))
            except:
                pass

            # Cleanup
            try:
                if context:
                    context.close()
                if p:
                    p.stop()
            except:
                pass

        except ImportError as e:
            self.log_to_gui(f"[LỖI] Lỗi import: {str(e)}")
            self.log_to_gui("[📝] Vui lòng cài đặt: pip install playwright")
            self.zalo_pause_button.configure(state="disabled")
        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi kết bạn: {str(e)}")
            self.zalo_pause_button.configure(state="disabled")

    # === QUẢN LÝ NHIỀU TÀI KHOẢN ZALO ===

    def load_account_list(self):
        """Load danh sách tài khoản vào combobox"""
        try:
            accounts = self.account_manager.get_all_accounts()

            if not accounts:
                self.account_combobox.configure(values=["Chưa có tài khoản"])
                self.account_combobox.set("Chưa có tài khoản")
                return

            # Tạo danh sách hiển thị: "Tên tài khoản (Tên Zalo)"
            account_list = []
            for account in accounts:
                zalo_name = account.get('zalo_name', 'Chưa cập nhật')
                display_name = f"{account['account_name']} ({zalo_name})"
                account_list.append(display_name)

            self.account_combobox.configure(values=account_list)

            # Chọn tài khoản đầu tiên nếu chưa có
            if not self.current_account_id and accounts:
                self.current_account_id = accounts[0]['id']
                self.account_combobox.set(account_list[0])
                self.update_account_info_display(accounts[0])

        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi load danh sách tài khoản: {str(e)}")

    def on_account_selected(self, choice):
        """Xử lý khi chọn tài khoản từ combobox"""
        if choice == "Chưa có tài khoản":
            return

        try:
            # Lấy account_name từ choice (format: "Tên tài khoản (Tên Zalo)")
            account_name = choice.split(" (")[0]

            # Tìm tài khoản theo tên
            accounts = self.account_manager.get_all_accounts()
            for account in accounts:
                if account['account_name'] == account_name:
                    self.current_account_id = account['id']
                    self.update_account_info_display(account)
                    self.log_to_gui(f"[OK] Đã chọn tài khoản: {account_name}")
                    break

        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi chọn tài khoản: {str(e)}")

    def update_account_info_display(self, account):
        """Cập nhật hiển thị thông tin tài khoản"""
        # Cập nhật họ tên Zalo
        zalo_name = account.get('zalo_name', 'Chưa cập nhật')
        if zalo_name and zalo_name != 'Chưa cập nhật':
            self.zalo_name_label.configure(text=zalo_name, text_color="#0068FF")
        else:
            self.zalo_name_label.configure(text="Chưa cập nhật", text_color="gray")

        # Cập nhật phiên đăng nhập
        last_login = account.get('last_login', 'Chưa đăng nhập')
        if last_login and last_login != 'Chưa đăng nhập':
            self.zalo_session_label.configure(text=last_login, text_color="white")
        else:
            self.zalo_session_label.configure(text="Chưa đăng nhập", text_color="gray")

        # Cập nhật trạng thái
        status = account.get('status', 'inactive')
        if status == 'active':
            self.zalo_status_label.configure(text="[OK] Active", text_color="green")
        else:
            self.zalo_status_label.configure(text="[LỖI] Inactive", text_color="red")

    def add_zalo_account(self):
        """Thêm tài khoản Zalo mới"""
        # Tạo dialog nhập tên tài khoản
        dialog = customtkinter.CTkInputDialog(
            text="Nhập tên tài khoản (ví dụ: Tài khoản 1, Zalo công ty, ...):",
            title="Thêm Tài Khoản Zalo"
        )
        account_name = dialog.get_input()

        if not account_name:
            return

        try:
            import zalo_logic

            # Khởi tạo account manager nếu chưa có
            if not self.account_manager:
                self.account_manager = zalo_logic.ZaloAccountManager()

            # Thêm tài khoản mới
            new_account = self.account_manager.add_account(account_name)

            self.log_to_gui(f"[OK] Đã thêm tài khoản: {account_name}")
            self.log_to_gui(f"   ID: {new_account['id']}")
            self.log_to_gui(f"   Session directory: {new_account['session_dir']}")

            # Reload danh sách
            self.load_account_list()

            # Chọn tài khoản mới
            self.current_account_id = new_account['id']
            self.account_combobox.set(f"{account_name} (Chưa cập nhật)")
            self.update_account_info_display(new_account)

            messagebox.showinfo(
                "Thành công",
                f"Đã thêm tài khoản: {account_name}\n\nVui lòng click 'Kiểm tra & Cập nhật' để đăng nhập!",
                parent=self
            )

        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi thêm tài khoản: {str(e)}")
            messagebox.showerror(
                "Lỗi",
                f"Không thể thêm tài khoản!\n{str(e)}",
                parent=self
            )

    def delete_zalo_account(self):
        """Xóa tài khoản Zalo hiện tại"""
        if not self.current_account_id:
            messagebox.showwarning(
                "Cảnh báo",
                "Vui lòng chọn tài khoản cần xóa!",
                parent=self
            )
            return

        try:
            account = self.account_manager.get_account_by_id(self.current_account_id)
            if not account:
                return

            # Xác nhận xóa
            result = messagebox.askyesno(
                "Xác nhận xóa",
                f"Bạn có chắc muốn xóa tài khoản:\n{account['account_name']}?\n\nSession và dữ liệu sẽ bị xóa vĩnh viễn!",
                parent=self
            )

            if result:
                # Xóa tài khoản
                if self.account_manager.delete_account(self.current_account_id):
                    self.log_to_gui(f"[OK] Đã xóa tài khoản: {account['account_name']}")

                    # Reset current account
                    self.current_account_id = None

                    # Reload danh sách
                    self.load_account_list()

                    messagebox.showinfo(
                        "Thành công",
                        f"Đã xóa tài khoản: {account['account_name']}",
                        parent=self
                    )
                else:
                    messagebox.showerror(
                        "Lỗi",
                        "Không thể xóa tài khoản!",
                        parent=self
                    )

        except Exception as e:
            self.log_to_gui(f"[LỖI] Lỗi khi xóa tài khoản: {str(e)}")

    def import_zalo_from_sheet(self):
        """Nhập dữ liệu khách hàng từ Google Sheet cho Auto Zalo"""
        self.log_to_gui("[ℹ️] Chức năng 'Nhập từ Sheet' sẽ được thêm sau")
        messagebox.showinfo(
            "Thông báo",
            "Chức năng 'Nhập từ Google Sheet' sẽ được thêm sau",
            parent=self
        )

    def import_contract_from_sheet(self):
        """Nhập dữ liệu hợp đồng từ Google Sheet cho Kiểm Tra Hợp Đồng"""
        self.log_to_gui("[ℹ️] Chức năng 'Nhập từ Sheet' sẽ được thêm sau")
        messagebox.showinfo(
            "Thông báo",
            "Chức năng 'Nhập từ Google Sheet' sẽ được thêm sau",
            parent=self
        )


if __name__ == "__main__":
    app = App()
    app.mainloop()

