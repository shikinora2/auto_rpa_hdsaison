import os
import time
import random
from playwright.sync_api import sync_playwright 
import base64 
import json
import re # Thêm thư viện Regex

# Thêm thư viện để xuất Excel
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Lỗi: Thư viện 'openpyxl' chưa được cài đặt. Vui lòng chạy: pip install openpyxl")
    # Sẽ ném lỗi sau nếu người dùng cố gắng chạy hàm
    pass

# --- CẤU HÌNH TRANG WEB ---
LOGIN_URL = "https://hpo.hdsaison.com.vn/login"
DASHBOARD_URL = "https://hpo.hdsaison.com.vn/dashboard"
CONTRACT_DETAIL_URL_BASE = "https://hpo.hdsaison.com.vn/contracts" 
CONTRACTS_URL_TEMPLATE = "https://hpo.hdsaison.com.vn/contracts#keyword=&startDate={START_DATE_HERE}&endDate={END_DATE_HERE}&filter=APPROVED_HARD_COPY_STATUS"

# --- CÁC SELECTOR (BỘ CHỌN) ---
# Trang danh sách
CONTRACT_CARDS_CONTAINER_SELECTOR = "div.cards"
USERNAME_SELECTOR = "[formcontrolname='username']"
PASSWORD_SELECTOR = "[formcontrolname='password']"
LOGIN_BUTTON_SELECTOR = "text=Đăng nhập"
SEE_MORE_BUTTON_SELECTOR = "button:has-text('Xem thêm...')"
CONTRACT_CARD_SELECTOR = "div.fuse-card"
CONTRACT_ID_SELECTOR = "span.fl-right"

# Trang chi tiết - Các Tab chính
CUSTOMER_INFO_TAB_SELECTOR = "div.mat-tab-label:has-text('THÔNG TIN KHÁCH HÀNG')"
# === C. SELECTOR CHO TAB "THÔNG TIN HÀNG HÓA" ===
GOODS_INFO_TAB_SELECTOR = "div.mat-tab-label:has-text('THÔNG TIN HÀNG HÓA')"

# Thông tin POS
GOODS_POS_ID_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Mã POS:']/following-sibling::div"""
GOODS_POS_NAME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Tên POS:']/following-sibling::div"""
GOODS_POS_ADDRESS_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Địa chỉ:']/following-sibling::div"""
GOODS_USERNAME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Username:']/following-sibling::div"""
GOODS_FULLNAME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Họ tên:']/following-sibling::div"""
GOODS_SCHEME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Scheme:']/following-sibling::div"""

# Thông tin sản phẩm (hàng hóa)
GOODS_PRODUCT_CARD_SELECTOR = """//div[contains(@class, 'products')]/mat-card"""
# Các selector con (dùng CSS selector cho query_selector)
GOODS_PRODUCT_NAME_CSS = """div[class*='title'] h3"""
GOODS_PRODUCT_PRICE_CSS = """div[class*='currency']"""
GOODS_PRODUCT_TYPE_CSS = """span[class*='product-type']"""

# Thông tin tài chính - CẢI TIẾN SELECTOR
GOODS_TOTAL_AMOUNT_SELECTOR = """//div[@class='label-info' and contains(text(), 'Tổng')]/following-sibling::div"""
GOODS_DOWN_PAYMENT_SELECTOR = """//div[@class='label-info' and contains(text(), 'trả trước')]/following-sibling::div"""
GOODS_LOAN_AMOUNT_SELECTOR = """//div[contains(@class, 'label-info-strong') and contains(text(), 'tiền vay')]/following-sibling::div"""

# Thông tin gói vay - CẢI TIẾN SELECTOR
GOODS_INSTALLMENT_AMOUNT_SELECTOR = """//div[contains(@class, 'monthly-installment')]"""
GOODS_INSTALLMENT_MONTHS_SELECTOR = """//div[contains(@class, 'period-installment')]//span"""
GOODS_INTEREST_RATE_SELECTOR = """//span[contains(text(), 'Lãi suất')]/ancestor::div[1]/following-sibling::div//span"""
GOODS_INSURANCE_SELECTOR = """//span[contains(text(), 'Bảo hiểm')]/ancestor::div[1]/following-sibling::div//span"""
GOODS_BONUS_SCHEME_SELECTOR = """//span[contains(text(), 'Bonus')]/ancestor::div[1]/following-sibling::div//span"""

# Trang chi tiết - Các Panel (trong Tab Thông tin khách hàng)
PERSONAL_INFO_PANEL_SELECTOR = "mat-expansion-panel-header:has-text('1. Thông tin cá nhân')"
RESIDENCE_INFO_PANEL_SELECTOR = "mat-expansion-panel-header:has-text('2. Thông tin cư trú')"
REFERENCE_INFO_PANEL_SELECTOR = "mat-expansion-panel-header:has-text('3. Người tham chiếu')"
DOCUMENT_INFO_PANEL_SELECTOR = "mat-expansion-panel-header:has-text('4. Thông tin giấy tờ')"

# --- 1. Panel "Thông tin cá nhân" ---
CUSTOMER_NAME_SELECTOR = """//div[contains(@class, 'customer-profile-picture')]//span[@class='caption']"""
PROFILE_PIC_SELECTOR = """//div[contains(@class, 'customer-profile-picture')]//img[contains(@class, 'mat-menu-trigger')]"""
GENDER_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Giới tính:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
DOB_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Ngày sinh:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
CCCD_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Số CCCD:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
CCCD_ISSUE_DATE_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Ngày cấp:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
CCCD_EXPIRY_DATE_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Ngày hết hạn:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
CCCD_FRONT_IMG_SELECTOR = """//mat-panel-title[contains(., 'Ảnh chụp CCCD (mặt trước)')]/following-sibling::div//img"""
CCCD_BACK_IMG_SELECTOR = """//mat-panel-title[contains(., 'Ảnh chụp CCCD (mặt sau)')]/following-sibling::div//img"""
PHONE_SELECTOR_PRIMARY = """//div[@class='detail-column-1' and normalize-space(.) = 'Số điện thoại:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
PHONE_SELECTOR_OTHER = """//div[@class='detail-column-1' and normalize-space(.) = 'Số điện thoại khác:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
EMAIL_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Email:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
MARITAL_STATUS_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Tình trạng hôn nhân:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
EDUCATION_LEVEL_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Cấp bậc học vấn:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
JOB_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Nghề làm việc:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
COMPANY_NAME_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Tên công ty:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
COMPANY_ADDRESS_SELECTOR = """//div[contains(@class, 'detail-column-1') and contains(@class, 'align-top') and normalize-space(.) = 'Địa chỉ công ty:']/following-sibling::div[contains(@class, 'detail-column-2')]"""
INCOME_SELECTOR = """//div[@class='detail-column-1' and normalize-space(.) = 'Thu nhập:']/following-sibling::div[contains(@class, 'detail-column-2')]"""

# --- 2. Panel "Thông tin cư trú" ---
RES_PERMANENT_PROVINCE_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ thường trú')]/following-sibling::div[div[normalize-space(.)='Tỉnh/Thành phố:']]/div[contains(@class, 'detail-column-2')]"""
RES_PERMANENT_DISTRICT_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ thường trú')]/following-sibling::div[div[normalize-space(.)='Quận/Huyện:']]/div[contains(@class, 'detail-column-2')]"""
RES_PERMANENT_WARD_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ thường trú')]/following-sibling::div[div[normalize-space(.)='Phường/Xã:']]/div[contains(@class, 'detail-column-2')]"""
RES_PERMANENT_ADDRESS_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ thường trú')]/following-sibling::div[div[normalize-space(.)='Địa chỉ:']]/div[contains(@class, 'detail-column-2')]"""
RES_TEMP_PROVINCE_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ tạm trú')]/following-sibling::div[div[normalize-space(.)='Tỉnh/Thành phố:']]/div[contains(@class, 'detail-column-2')]"""
RES_TEMP_DISTRICT_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ tạm trú')]/following-sibling::div[div[normalize-space(.)='Quận/Huyện:']]/div[contains(@class, 'detail-column-2')]"""
RES_TEMP_WARD_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ tạm trú')]/following-sibling::div[div[normalize-space(.)='Phường/Xã:']]/div[contains(@class, 'detail-column-2')]"""
RES_TEMP_ADDRESS_SELECTOR = """//mat-expansion-panel[contains(., '2. Thông tin cư trú')]//mat-panel-title[contains(., 'Địa chỉ tạm trú')]/following-sibling::div[div[normalize-space(.)='Địa chỉ:']]/div[contains(@class, 'detail-column-2')]"""

# --- 3. Panel "Người tham chiếu" ---
REF1_NAME_SELECTOR = """//mat-expansion-panel[contains(., '3. Người tham chiếu')]//mat-panel-title[contains(., 'Người tham chiếu 1')]/following-sibling::div[div[normalize-space(.)='Họ tên:']]/div[contains(@class, 'detail-column-2')]"""
REF1_PHONE_SELECTOR = """//mat-expansion-panel[contains(., '3. Người tham chiếu')]//mat-panel-title[contains(., 'Người tham chiếu 1')]/following-sibling::div[div[normalize-space(.)='Số điện thoại:']]/div[contains(@class, 'detail-column-2')]"""
REF1_RELATION_SELECTOR = """//mat-expansion-panel[contains(., '3. Người tham chiếu')]//mat-panel-title[contains(., 'Người tham chiếu 1')]/following-sibling::div[div[contains(., 'Mối quan hệ')]]/div[contains(@class, 'detail-column-2')]"""
REF2_NAME_SELECTOR = """//mat-expansion-panel[contains(., '3. Người tham chiếu')]//mat-panel-title[contains(., 'Người tham chiếu 2')]/following-sibling::div[div[normalize-space(.)='Họ tên:']]/div[contains(@class, 'detail-column-2')]"""
REF2_PHONE_SELECTOR = """//mat-expansion-panel[contains(., '3. Người tham chiếu')]//mat-panel-title[contains(., 'Người tham chiếu 2')]/following-sibling::div[div[normalize-space(.)='Số điện thoại:']]/div[contains(@class, 'detail-column-2')]"""
REF2_RELATION_SELECTOR = """//mat-expansion-panel[contains(., '3. Người tham chiếu')]//mat-panel-title[contains(., 'Người tham chiếu 2')]/following-sibling::div[div[contains(., 'Mối quan hệ')]]/div[contains(@class, 'detail-column-2')]"""

# --- 4. Panel "Thông tin giấy tờ" ---
FIRST_PAYMENT_DATE_SELECTOR = """//mat-expansion-panel[contains(., '4. Thông tin giấy tờ')]//div[@class='detail-column-1' and contains(., 'Ngày đóng tiền đầu tiên:')]/following-sibling::div[contains(@class, 'detail-column-2')]"""

# --- 5. Tab "Thông tin hàng hóa" ---
GOODS_POS_ID_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Mã POS:']/following-sibling::div"""
GOODS_POS_NAME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Tên POS:']/following-sibling::div"""
GOODS_POS_ADDRESS_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Địa chỉ:']/following-sibling::div"""
GOODS_USERNAME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Username:']/following-sibling::div"""
GOODS_FULLNAME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Họ tên:']/following-sibling::div"""
GOODS_SCHEME_SELECTOR = """//div[contains(@class, 'pos-info')]//div[@class='label-info' and normalize-space(.) = 'Scheme:']/following-sibling::div"""
GOODS_PRODUCT_CARD_SELECTOR = """//div[contains(@class, 'products')]/mat-card"""
# Selector cho Tên và Giá (dùng bên trong card)
GOODS_PRODUCT_NAME_SELECTOR = """.//div[contains(@class, 'title')]/h3"""
GOODS_PRODUCT_PRICE_SELECTOR = """.//div[contains(@class, 'currency')]"""
GOODS_PRODUCT_TYPE_SELECTOR = """.//span[contains(@class, 'product-type')]"""
# Selector cho các mục tài chính (khớp với HTML của bạn)
GOODS_TOTAL_AMOUNT_SELECTOR = """//div[contains(@class, 'contain-text')]//div[@class='label-info' and normalize-space(.) = 'Tổng:']/following-sibling::div"""
GOODS_DOWN_PAYMENT_SELECTOR = """//div[contains(@class, 'contain-text')]//div[@class='label-info' and normalize-space(.) = 'Số tiền trả trước:']/following-sibling::div"""
GOODS_LOAN_AMOUNT_SELECTOR = """//div[contains(@class, 'contain-text')]//div[contains(@class, 'label-info-strong') and normalize-space(.) = 'Số tiền vay:']/following-sibling::div"""
GOODS_INSTALLMENT_AMOUNT_SELECTOR = """//div[contains(@class, 'summary')]//div[contains(@class, 'monthly-installment')]"""
GOODS_INSTALLMENT_MONTHS_SELECTOR = """//div[contains(@class, 'summary')]//div[contains(@class, 'period-installment')]/span"""
GOODS_INTEREST_RATE_SELECTOR = """//div[contains(@class, 'summary')]//span[contains(@class, 'm-0') and normalize-space(.) = 'Lãi suất']/parent::div/following-sibling::div/span"""
GOODS_INSURANCE_SELECTOR = """//div[contains(@class, 'summary')]//span[contains(@class, 'm-0') and normalize-space(.) = 'Bảo hiểm']/parent::div/following-sibling::div/span"""
GOODS_BONUS_SCHEME_SELECTOR = """//div[contains(@class, 'summary')]//span[contains(@class, 'm-0') and normalize-space(.) = 'Bonus scheme']/parent::div/following-sibling::div/span"""
# -----------------------------------------------------------------


# === HÀM HELPER ĐỂ GỬI LOG VỀ GUI ===
def _callback_handler(message, status_callback=None):
    print(message) # Vẫn in ra console để debug
    if status_callback:
        status_callback(message) # Gửi message về UI

# === HÀM HELPER MỚI: LẤY TEXT TỪ TRANG (AN TOÀN) ===
def _get_text(page, selector: str, timeout: int = 3000) -> str | None: # Sửa: Tăng timeout
    """
    Hàm helper để lấy inner_text một cách an toàn.
    Trả về None nếu không tìm thấy hoặc hết thời gian.
    """
    try:
        # Sử dụng page.inner_text với timeout (tăng lên 3s)
        # Nó sẽ tự động chờ element xuất hiện
        text = page.inner_text(selector, timeout=timeout)
        return text.strip()
    except Exception:
        # Trả về None nếu element không tìm thấy
        return None

# === HÀM HELPER MỚI: XUẤT EXCEL (CHO DỮ LIỆU CÀO) ===
def _export_details_to_excel(data_list: list, save_directory: str, start_date_ddmmyyyy: str, end_date_ddmmyyyy: str):
    """
    Hàm nội bộ để xuất dữ liệu chi tiết đã cào ra file Excel.
    """
    if not data_list:
        _callback_handler("Không có dữ liệu chi tiết để xuất Excel.")
        return

    try:
        import time
        
        # 1. Tạo Workbook và Worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chi tiết Hợp đồng"

        # 2. Định nghĩa Headers (Tiêu đề cột) và Keys (Khóa dictionary)
        headers = [
            'STT', 'ID Hợp đồng', 'Tên KH (Profile)', 'Giới tính', 'Ngày sinh', 'Số CCCD', 'Ngày cấp', 'Ngày hết hạn', 
            'SĐT (Chính)', 'Tình trạng hôn nhân', 'Học vấn', 'Nghề nghiệp', 'Tên công ty', 
            'Địa chỉ công ty', 'Thu nhập',
            'Địa chỉ thường trú', 
            'Địa chỉ tạm trú', 
            'Tham chiếu 1: Tên', 'Tham chiếu 1: SĐT', 'Tham chiếu 1: Quan hệ',
            'Tham chiếu 2: Tên', 'Tham chiếu 2: SĐT', 'Tham chiếu 2: Quan hệ',
            'Ngày đóng tiền đầu tiên',
            'Mã POS', 'Tên POS', 'Địa chỉ POS', 'Username', 'Scheme',
            'Sản phẩm (Gộp)', 'Tổng tiền', 'Trả trước', 'Số tiền vay',
            'Góp mỗi tháng', 'Số tháng', 'Lãi suất', 'Bảo hiểm', 'Bonus Scheme'
        ]
        
        keys = [
            'stt', 'contract_id', 'customer_name', 'gender', 'dob', 'cccd', 'cccd_issue_date', 'cccd_expiry_date',
            'phone_primary', 'marital_status', 'education_level', 'job', 'company_name',
            'company_address', 'income',
            'dia_chi_thuong_tru', 
            'dia_chi_tam_tru', 
            'ref1_name', 'ref1_phone', 'ref1_relation',
            'ref2_name', 'ref2_phone', 'ref2_relation',
            'first_payment_date',
            'goods_pos_id', 'goods_pos_name', 'goods_pos_address', 'goods_username', 'goods_scheme',
            'products_joined', 'goods_total_amount', 'goods_down_payment', 'goods_loan_amount',
            'goods_installment_amount', 'goods_installment_months', 'goods_interest_rate', 'goods_insurance', 'goods_bonus_scheme'
        ]
        
        ws.append(headers)

        # 3. Ghi Dữ liệu
        for i, data in enumerate(data_list):
            data['stt'] = i + 1 # Thêm STT
            # Xây dựng hàng (row) dựa trên thứ tự của 'keys'
            row = [data.get(key, None) for key in keys]
            ws.append(row)

        # 4. Tự động điều chỉnh độ rộng cột (ước lượng)
        for i, header in enumerate(headers):
            try:
                # Lấy độ rộng của header hoặc 15 (cái nào lớn hơn), tối đa 40
                width = max(len(str(header)), 15)
                if "Địa chỉ" in header or "Tên POS" in header: width = 30
                if "Sản phẩm" in header: width = 40
                ws.column_dimensions[get_column_letter(i + 1)].width = width
            except Exception:
                pass # Bỏ qua nếu lỗi

        # 5. Định dạng font Times New Roman cho toàn bộ bảng
        from openpyxl.styles import Font, Alignment
        
        times_font = Font(name='Times New Roman', size=11)
        header_font = Font(name='Times New Roman', size=11, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Định dạng header (hàng đầu tiên)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Định dạng các hàng dữ liệu
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = times_font
                # Căn giữa cho cột STT
                if cell.column == 1:
                    cell.alignment = center_alignment

        # 6. Lưu file
        # Tạo tên file theo định dạng: daystart_to_dayend_month_timestamp.xlsx
        # Ví dụ: 01_to_31_Jan2025_20250103_143022.xlsx
        day_start = start_date_ddmmyyyy[0:2]
        day_end = end_date_ddmmyyyy[0:2]
        month = start_date_ddmmyyyy[2:4]
        year = start_date_ddmmyyyy[4:8]
        
        # Tạo timestamp để mỗi file là duy nhất
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        
        # Tên tháng (số)
        month_name = f"{month}{year}"
        
        filename = f"{day_start}_to_{day_end}_{month_name}_{timestamp}.xlsx"
        output_path = os.path.join(save_directory, filename)
        
        wb.save(output_path)
        _callback_handler(f"✅ ĐÃ XUẤT EXCEL THÀNH CÔNG: {output_path}")

    except ImportError:
        _callback_handler("❌ LỖI: Vui lòng cài đặt 'openpyxl' (pip install openpyxl)")
    except Exception as e:
        _callback_handler(f"❌ Lỗi khi xuất file Excel: {e}")


def _perform_login_and_scrape_ids(
    username, password, 
    start_date_url, end_date_url, 
    pause_event, stop_event, 
    status_callback=None
):
    """
    Hàm nội bộ: Thực hiện Giai đoạn 1 (Đăng nhập, Lọc, Lấy hết ID).
    Trả về (playwright_instance, browser, page, contract_ids_list) nếu thành công.
    Trả về (None, None, None, None) nếu thất bại.
    """
    callback = lambda msg: _callback_handler(msg, status_callback)
    
    dynamic_filtered_url = CONTRACTS_URL_TEMPLATE.format(
        START_DATE_HERE=start_date_url,
        END_DATE_HERE=end_date_url
    )
    callback(f"URL sẽ sử dụng là: {dynamic_filtered_url}")
    callback("-----------------------------------")
    
    p = sync_playwright().start()
    browser = p.chromium.launch(headless=False, slow_mo=250)
    page = browser.new_page()

    try:
        # === KIỂM TRA DỪNG/TẠM DỪNG ===
        if stop_event.is_set(): raise InterruptedError("Dừng trước khi bắt đầu")
        pause_event.wait()
        # ===============================

        # === TIẾN TRÌNH 1: LOGIN ===
        callback(f"Tiến trình 1: Đang mở trang đăng nhập: {LOGIN_URL}")
        page.goto(LOGIN_URL)
        callback("Đang điền thông tin đăng nhập...")
        page.fill(USERNAME_SELECTOR, username)
        page.fill(PASSWORD_SELECTOR, password)
        page.click(LOGIN_BUTTON_SELECTOR)
        
        callback(f"\nTiến trình 2: Đang chờ trang Dashboard...")
        page.wait_for_url(DASHBOARD_URL, timeout=30000) 
        callback("✔ Đăng nhập thành công! Đã vào Dashboard.")
        
        # === GIAI ĐOẠN 1: THU THẬP DANH SÁCH ID HỢP ĐỒNG ===
        callback(f"\n--- GIAI ĐOẠN 1: THU THẬP DANH SÁCH ID HỢP ĐỒNG ---")
        
        # === KIỂM TRA DỪNG/TẠM DỪNG ===
        if stop_event.is_set(): raise InterruptedError("Dừng bởi người dùng")
        pause_event.wait()
        # ===============================
        
        callback(f"\nĐang điều hướng đến URL đã lọc theo ngày...")
        page.goto(dynamic_filtered_url)
        page.wait_for_url(dynamic_filtered_url, timeout=30000)
        
        callback(f"\nĐang ép F5 (Tải lại) trang để áp dụng bộ lọc...")
        page.reload()
        
        callback(f"   Đang chờ DỮ LIỆU ĐỢT ĐẦU tải...")
        page.wait_for_selector(CONTRACT_CARDS_CONTAINER_SELECTOR, timeout=30000)
        callback("✔ Khu vực dữ liệu ('div.cards') đã tải.")
        
        callback(f"\nBắt đầu tìm và nhấn nút 'Xem thêm...'")
        
        while page.is_visible(SEE_MORE_BUTTON_SELECTOR):
            if stop_event.is_set():
                callback("...Đã nhận lệnh dừng (trong khi tải 'Xem thêm')...")
                break 
            pause_event.wait() 

            try:
                callback("   Đang nhấn 'Xem thêm...' để tải thêm hợp đồng...")
                page.click(SEE_MORE_BUTTON_SELECTOR)
                page.wait_for_load_state('networkidle', timeout=15000) 
            except Exception as e:
                callback(f"   Lỗi khi click 'Xem thêm' (có thể nút đã biến mất): {e}")
                break 
        
        if stop_event.is_set():
            raise InterruptedError("Kịch bản bị dừng bởi người dùng")

        callback("   (Nút 'Xem thêm...' không còn hiển thị)")
        callback("✔ ĐÃ TẢI HẾT TẤT CẢ HỢP ĐỒNG!")

        callback(f"\nBắt đầu trích xuất tất cả ID hợp đồng...")
        contract_ids_to_process = []
        all_cards = page.query_selector_all(CONTRACT_CARD_SELECTOR)
        
        for card in all_cards:
            try:
                ed_id = card.query_selector(CONTRACT_ID_SELECTOR).inner_text()
                contract_ids_to_process.append(ed_id)
            except Exception as e:
                callback(f"   Lỗi khi trích xuất ID từ 1 thẻ: {e}")

        if not contract_ids_to_process:
            callback("KHÔNG TÌM THẤY HỢP ĐỒNG NÀO. Kịch bản dừng lại.")
            browser.close()
            p.stop()
            return None, None, None, None

        callback(f"   ==> Đã thu thập xong! Tìm thấy {len(contract_ids_to_process)} ID hợp đồng.")
        
        return p, browser, page, contract_ids_to_process

    except Exception as e:
        callback(f"\n❌ LỖI TRONG GIAI ĐOẠN 1: {e}")
        if browser: browser.close()
        p.stop()
        return None, None, None, None


# === HÀM 1: CHỈ KIỂM TRA SỐ LƯỢNG ===
def check_contract_count(
    username, password, 
    start_date_ddmmyyyy, end_date_ddmmyyyy, 
    pause_event, stop_event, 
    status_callback=None
):
    """Chức năng cho nút 'Kiểm tra số lượng'"""
    callback = lambda msg: _callback_handler(msg, status_callback)
    
    try:
        start_date_url = f"{start_date_ddmmyyyy[0:2]}-{start_date_ddmmyyyy[2:4]}-{start_date_ddmmyyyy[4:8]}"
        end_date_url = f"{end_date_ddmmyyyy[0:2]}-{end_date_ddmmyyyy[2:4]}-{end_date_ddmmyyyy[4:8]}"
    except Exception as e:
        callback(f"Lỗi định dạng ngày: {e}")
        return

    callback("--- BẮT ĐẦU CHẾ ĐỘ KIỂM TRA SỐ LƯỢNG ---")
    
    playwright_instance, browser, page, contract_ids = _perform_login_and_scrape_ids(
        username, password, start_date_url, end_date_url,
        pause_event, stop_event, status_callback
    )

    if browser: # Nếu hàm trên chạy thành công
        callback("\n===================================")
        callback(f"✅ KIỂM TRA HOÀN TẤT: TÌM THẤY TỔNG CỘNG {len(contract_ids)} HỢP ĐỒNG.")
        callback("===================================")
        browser.close()
        playwright_instance.stop()
    else:
        callback("\n===================================")
        callback(f"❌ KIỂM TRA THẤT BẠI.")
        callback("===================================")


# === HÀM 2: CHẠY VÀ TẢI FILE ===
def run_scrape_and_download_files(
    username, password, 
    start_date_ddmmyyyy, end_date_ddmmyyyy, 
    save_directory, 
    save_format, # (PDF hoặc JSON)
    pause_event, stop_event, 
    status_callback=None
):
    """Chức năng cho nút 'Bắt Đầu Tải'"""
    callback = lambda msg: _callback_handler(msg, status_callback)

    # === BƯỚC 1 & 2: Định dạng ngày và tạo thư mục ===
    try:
        start_date_url = f"{start_date_ddmmyyyy[0:2]}-{start_date_ddmmyyyy[2:4]}-{start_date_ddmmyyyy[4:8]}"
        end_date_url = f"{end_date_ddmmyyyy[0:2]}-{end_date_ddmmyyyy[2:4]}-{end_date_ddmmyyyy[4:8]}"
    except Exception as e:
        callback(f"Lỗi định dạng ngày. Lỗi: {e}")
        return
    callback(f"\nOK. Sẽ lọc từ ngày: {start_date_url} đến {end_date_url}")

    try:
        month_year_folder = f"{start_date_ddmmyyyy[2:4]}{start_date_ddmmyyyy[4:8]}"
        base_download_dir = save_directory
        final_save_dir = os.path.join(base_download_dir, month_year_folder)
        if not os.path.exists(final_save_dir):
            os.makedirs(final_save_dir)
            callback(f"   -> Đã tạo thư mục lưu trữ: '{final_save_dir}'")
        else:
            callback(f"   -> Thư mục lưu trữ '{final_save_dir}' đã tồn tại.")
    except Exception as e:
        callback(f"Lỗi nghiêm trọng: Không thể tạo thư mục lưu trữ '{final_save_dir}'. Lỗi: {e}")
        return

    # === GỌI GIAI ĐOẠN 1 ===
    playwright_instance, browser, page, contract_ids_to_process = _perform_login_and_scrape_ids(
        username, password, start_date_url, end_date_url,
        pause_event, stop_event, status_callback
    )

    if not browser:
        callback("❌ Kịch bản dừng do Giai đoạn 1 thất bại.")
        return

    # === GIAI ĐOẠN 2: XỬ LÝ TỪNG ID ===
    try:
        callback(f"\n--- GIAI ĐOẠN 2: BẮT ĐẦU XỬ LÝ {len(contract_ids_to_process)} HỢP ĐỒNG ---")
        
        for i, ed_id in enumerate(contract_ids_to_process):
            
            # === KIỂM TRA DỪNG/TẠM DỪNG (TRONG VÒNG LẶP) ===
            if stop_event.is_set():
                callback("...Đã nhận lệnh dừng (trước khi xử lý ID tiếp theo)...")
                break 
            pause_event.wait()
            # ============================================

            callback(f"\n   Đang xử lý HĐ #{i + 1}/{len(contract_ids_to_process)} (ID: {ed_id})...")

            try:
                detail_url = f"{CONTRACT_DETAIL_URL_BASE}/{ed_id}"
                callback(f"      Đang điều hướng thẳng đến: {detail_url}")
                page.goto(detail_url)
                
                callback(f"      Đang chờ trang chi tiết của HĐ {ed_id} tải...")
                page.wait_for_selector(f"text={ed_id}", timeout=30000)
                
                callback(f"      1. Click tab 'BỘ HỢP ĐỒNG'...")
                page.click("div.mat-tab-label-content:has-text('BỘ HỢP ĐỒNG')")

                callback(f"      2. Click nút 'Lấy file in hợp đồng'...")
                page.click("button:has-text('Lấy file in hợp đồng')")

                callback(f"      3. Đang chờ file tải về...")
                with page.expect_download() as download_info:
                    download_link_selector = f"a:has-text('Contract_{ed_id}')"
                    callback(f"         Đang chờ link tải về '{download_link_selector}' xuất hiện...")
                    page.wait_for_selector(download_link_selector, timeout=60000) 
                    
                    callback(f"         Đang click vào link để bắt đầu tải...")
                    page.click(download_link_selector)

                download = download_info.value
                
                # === LOGIC MỚI: LƯU PDF hoặc JSON/BASE64 ===
                if save_format == "PDF":
                    # Cách 1: Lưu file PDF
                    save_path = os.path.join(final_save_dir, download.suggested_filename)
                    download.save_as(save_path)
                    callback(f"      ✔ Đã TẢI XONG và lưu vào: {save_path}")
                
                elif save_format == "JSON":
                    # Cách 2: Lưu Base64 vào JSON
                    temp_file_path = download.path() # Lấy đường dẫn file tạm
                    
                    pdf_bytes = b""
                    with open(temp_file_path, 'rb') as f:
                        pdf_bytes = f.read() # Đọc toàn bộ file
                    
                    # Mã hóa sang Base64
                    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
                    
                    # Tạo đối tượng JSON
                    json_data = {
                        "contract_id": ed_id,
                        "filename": download.suggested_filename,
                        "pdf_base64": pdf_base64
                    }
                    
                    # Lưu file JSON
                    json_save_path = os.path.join(final_save_dir, f"{ed_id}.json")
                    with open(json_save_path, 'w') as f:
                        json.dump(json_data, f)
                        
                    callback(f"      ✔ Đã LƯU DỮ LIỆU BASE64 vào: {json_save_path}")
                    
                    # (Playwright sẽ tự động xóa file tạm sau khi hàm 'with' kết thúc)

                # =========================================
                
                # Random delay giữa 2-5 giây để tránh spam
                delay = random.uniform(2, 5)
                callback(f"      ⏳ Chờ {delay:.1f}s trước khi xử lý HĐ tiếp theo...")
                time.sleep(delay)
                
            except Exception as e:
                callback(f"   *** Lỗi khi xử lý hợp đồng ID {ed_id}: {e}")
                callback(f"   *** Bỏ qua và tiếp tục với HĐ tiếp theo...")
                continue 

        # === THÔNG BÁO KẾT THÚC ===
        if stop_event.is_set():
            callback("\n===================================")
            callback(f"🛑 KỊCH BẢN ĐÃ DỪNG THEO YÊU CẦU CỦA NGƯỜI DÙNG.")
        else:
            callback("\n===================================")
            callback(f"🎉 HOÀN TẤT! ĐÃ XỬ LÝ TẤT CẢ {len(contract_ids_to_process)} HỢP ĐỒNG!")
        
        callback("===================================")
        close_delay = random.uniform(8, 12)
        callback(f"\n(Trình duyệt sẽ tự đóng sau {close_delay:.1f} giây...)")
        time.sleep(close_delay)

    except InterruptedError: # Bắt lỗi chủ động dừng
        callback("\n===================================")
        callback(f"🛑 KỊCH BẢN ĐÃ DỪNG THEO YÊU CẦU CỦA NGƯỜI DÙNG.")
    
    except Exception as e:
        callback("\n===================================")
        callback(f"❌ ĐÃ XAY RA LỖI LỚN TRONG GIAI ĐOẠN 2!")
        callback(f"Lỗi: {e}")
    
    finally:
        if browser:
            browser.close()
            playwright_instance.stop()
        callback("Đã đóng trình duyệt.")

# === HÀM 3 (MỚI): CHẠY, CÀO DỮ LIỆU VÀ XUẤT EXCEL ===
def run_scrape_and_export_details(
    username, password, 
    start_date_ddmmyyyy, end_date_ddmmyyyy, 
    save_directory, 
    pause_event, stop_event, 
    status_callback=None
):
    """
    Chức năng cho nút 'Lấy Chi Tiết Hợp Đồng (Excel)'.
    Kịch bản: Đăng nhập -> Lấy IDs -> Lặp qua từng ID -> 
    Vào trang chi tiết -> Cào (scrape) dữ liệu -> Xuất Excel.
    """
    callback = lambda msg: _callback_handler(msg, status_callback)

    # === BƯỚC 1: Định dạng ngày và tạo thư mục ===
    try:
        start_date_url = f"{start_date_ddmmyyyy[0:2]}-{start_date_ddmmyyyy[2:4]}-{start_date_ddmmyyyy[4:8]}"
        end_date_url = f"{end_date_ddmmyyyy[0:2]}-{end_date_ddmmyyyy[2:4]}-{end_date_ddmmyyyy[4:8]}"
    except Exception as e:
        callback(f"Lỗi định dạng ngày. Lỗi: {e}")
        return

    try:
        # Tạo thư mục lưu trữ (ví dụ: downloads_contracts/012025)
        month_year_folder = f"{start_date_ddmmyyyy[2:4]}{start_date_ddmmyyyy[4:8]}"
        final_save_dir = os.path.join(save_directory, month_year_folder)
        if not os.path.exists(final_save_dir):
            os.makedirs(final_save_dir)
            callback(f"   -> Đã tạo thư mục lưu trữ: '{final_save_dir}'")
        else:
            callback(f"   -> Thư mục lưu trữ '{final_save_dir}' đã tồn tại.")
    except Exception as e:
        callback(f"Lỗi nghiêm trọng: Không thể tạo thư mục lưu trữ '{final_save_dir}'. Lỗi: {e}")
        return

    # === GỌI GIAI ĐOẠN 1: Lấy danh sách ID ===
    playwright_instance, browser, page, contract_ids_to_process = _perform_login_and_scrape_ids(
        username, password, start_date_url, end_date_url,
        pause_event, stop_event, status_callback
    )

    if not browser:
        callback("❌ Kịch bản dừng do Giai đoạn 1 thất bại.")
        return

    all_scraped_data = [] # List để lưu trữ dữ liệu của tất cả hợp đồng

    # === GIAI ĐOẠN 2 (MỚI): CÀO DỮ LIỆU CHI TIẾT TỪNG ID ===
    try:
        callback(f"\n--- GIAI ĐOẠN 2: BẮT ĐẦU CÀO (SCRAPE) CHI TIẾT {len(contract_ids_to_process)} HỢP ĐỒNG ---")
        
        for i, ed_id in enumerate(contract_ids_to_process):
            
            scraped_data = {"contract_id": ed_id} # Tạo dictionary cho hợp đồng này

            # === KIỂM TRA DỪNG/TẠM DỪNG (TRONG VÒNG LẶP) ===
            if stop_event.is_set():
                callback("...Đã nhận lệnh dừng (trước khi xử lý ID tiếp theo)...")
                break 
            pause_event.wait()
            # ============================================

            callback(f"\n   Đang cào HĐ #{i + 1}/{len(contract_ids_to_process)} (ID: {ed_id})...")

            try:
                detail_url = f"{CONTRACT_DETAIL_URL_BASE}/{ed_id}"
                callback(f"      Đang điều hướng thẳng đến: {detail_url}")
                page.goto(detail_url)
                
                callback(f"      Đang chờ trang chi tiết của HĐ {ed_id} tải...")
                page.wait_for_selector(f"text={ed_id}", timeout=30000)
                
                # --- A. CÀO TAB "THÔNG TIN KHÁCH HÀNG" ---
                try:
                    callback("      A. Đang cào Tab 'Thông tin khách hàng'...")
                    page.click(CUSTOMER_INFO_TAB_SELECTOR)
                    
                    # Click mở tất cả các panel
                    page.click(PERSONAL_INFO_PANEL_SELECTOR)
                    page.click(RESIDENCE_INFO_PANEL_SELECTOR)
                    page.click(REFERENCE_INFO_PANEL_SELECTOR)
                    page.click(DOCUMENT_INFO_PANEL_SELECTOR)
                    
                    # Chờ 1 panel bất kỳ (ví dụ: Giới tính) xuất hiện
                    page.wait_for_selector(GENDER_SELECTOR, timeout=5000)

                    # 1. Thông tin cá nhân
                    scraped_data['customer_name'] = _get_text(page, CUSTOMER_NAME_SELECTOR)
                    scraped_data['gender'] = _get_text(page, GENDER_SELECTOR)
                    scraped_data['dob'] = _get_text(page, DOB_SELECTOR)
                    scraped_data['cccd'] = _get_text(page, CCCD_SELECTOR)
                    # Sửa: Định dạng lại CCCD (bỏ gạch ngang)
                    if scraped_data['cccd']:
                        scraped_data['cccd'] = scraped_data['cccd'].replace('-', '')
                    
                    scraped_data['cccd_issue_date'] = _get_text(page, CCCD_ISSUE_DATE_SELECTOR)
                    scraped_data['cccd_expiry_date'] = _get_text(page, CCCD_EXPIRY_DATE_SELECTOR)
                    scraped_data['phone_primary'] = _get_text(page, PHONE_SELECTOR_PRIMARY)
                    scraped_data['phone_other'] = _get_text(page, PHONE_SELECTOR_OTHER)
                    scraped_data['email'] = _get_text(page, EMAIL_SELECTOR)
                    scraped_data['marital_status'] = _get_text(page, MARITAL_STATUS_SELECTOR)
                    scraped_data['education_level'] = _get_text(page, EDUCATION_LEVEL_SELECTOR)
                    scraped_data['job'] = _get_text(page, JOB_SELECTOR)
                    scraped_data['company_name'] = _get_text(page, COMPANY_NAME_SELECTOR)
                    scraped_data['company_address'] = _get_text(page, COMPANY_ADDRESS_SELECTOR)
                    scraped_data['income'] = _get_text(page, INCOME_SELECTOR)

                    # 2. Thông tin cư trú - Sửa: Gộp địa chỉ
                    p_tinh = _get_text(page, RES_PERMANENT_PROVINCE_SELECTOR)
                    p_huyen = _get_text(page, RES_PERMANENT_DISTRICT_SELECTOR)
                    p_xa = _get_text(page, RES_PERMANENT_WARD_SELECTOR)
                    p_diachi = _get_text(page, RES_PERMANENT_ADDRESS_SELECTOR)
                    
                    # Gộp địa chỉ thường trú, lọc bỏ các phần None hoặc rỗng
                    dia_chi_thuong_tru_parts = [p_diachi, p_xa, p_huyen, p_tinh]
                    # Lọc các giá trị rỗng, None, hoặc chỉ chứa dấu '.'
                    scraped_data['dia_chi_thuong_tru'] = ", ".join(part for part in dia_chi_thuong_tru_parts if part and part.strip() and part.strip() != '.')

                    # Gộp địa chỉ tạm trú
                    t_tinh = _get_text(page, RES_TEMP_PROVINCE_SELECTOR)
                    t_huyen = _get_text(page, RES_TEMP_DISTRICT_SELECTOR)
                    t_xa = _get_text(page, RES_TEMP_WARD_SELECTOR)
                    t_diachi = _get_text(page, RES_TEMP_ADDRESS_SELECTOR)
                    
                    dia_chi_tam_tru_parts = [t_diachi, t_xa, t_huyen, t_tinh]
                    scraped_data['dia_chi_tam_tru'] = ", ".join(part for part in dia_chi_tam_tru_parts if part and part.strip() and part.strip() != '.')
                    
                    # 3. Người tham chiếu
                    scraped_data['ref1_name'] = _get_text(page, REF1_NAME_SELECTOR)
                    scraped_data['ref1_phone'] = _get_text(page, REF1_PHONE_SELECTOR)
                    scraped_data['ref1_relation'] = _get_text(page, REF1_RELATION_SELECTOR)
                    scraped_data['ref2_name'] = _get_text(page, REF2_NAME_SELECTOR)
                    # Sửa: Thêm 2 trường cho Tham chiếu 2
                    scraped_data['ref2_phone'] = _get_text(page, REF2_PHONE_SELECTOR)
                    scraped_data['ref2_relation'] = _get_text(page, REF2_RELATION_SELECTOR)
                    
                    # 4. Thông tin giấy tờ
                    scraped_data['first_payment_date'] = _get_text(page, FIRST_PAYMENT_DATE_SELECTOR)
                
                except Exception as e:
                    callback(f"      Lỗi khi cào Tab 'Thông tin khách hàng': {e}")
                    
                # --- B. CÀO TAB "THÔNG TIN HÀNG HÓA" ---
                try:
                    callback("      B. Đang cào Tab 'Thông tin hàng hóa'...")
                    page.click(GOODS_INFO_TAB_SELECTOR)
                    
                    # Chờ 1 element (Mã POS) xuất hiện
                    page.wait_for_selector(GOODS_POS_ID_SELECTOR, timeout=5000)

                    # 1. Thông tin POS
                    scraped_data['goods_pos_id'] = _get_text(page, GOODS_POS_ID_SELECTOR)
                    scraped_data['goods_pos_name'] = _get_text(page, GOODS_POS_NAME_SELECTOR)
                    scraped_data['goods_pos_address'] = _get_text(page, GOODS_POS_ADDRESS_SELECTOR)
                    scraped_data['goods_username'] = _get_text(page, GOODS_USERNAME_SELECTOR)
                    scraped_data['goods_fullname'] = _get_text(page, GOODS_FULLNAME_SELECTOR)
                    scraped_data['goods_scheme'] = _get_text(page, GOODS_SCHEME_SELECTOR)

                    # 2. Thông tin Hàng hóa (lặp)
                    products_list = []
                    product_cards = page.query_selector_all(GOODS_PRODUCT_CARD_SELECTOR)
                    callback(f"         Tìm thấy {len(product_cards)} sản phẩm.")
                    for card in product_cards:
                        try:
                            # Dùng CSS selector cho query_selector
                            name_element = card.query_selector(GOODS_PRODUCT_NAME_CSS)
                            if name_element:
                                name = name_element.inner_text().strip()
                                products_list.append(name)
                        except Exception as e:
                            callback(f"         ⚠ Lỗi khi lấy tên sản phẩm: {e}")
                    scraped_data['products_joined'] = ", ".join(products_list) 

                    # 3. Thông tin Tổng tiền - CẢI TIẾN
                    callback("         Đang lấy thông tin tài chính...")
                    
                    # Thử chờ element xuất hiện trước khi lấy
                    try:
                        page.wait_for_selector(GOODS_TOTAL_AMOUNT_SELECTOR, timeout=8000)
                    except:
                        callback("         ⚠ Không tìm thấy thông tin 'Tổng tiền', thử selector khác...")
                    
                    scraped_data['goods_total_amount'] = _get_text(page, GOODS_TOTAL_AMOUNT_SELECTOR)
                    scraped_data['goods_down_payment'] = _get_text(page, GOODS_DOWN_PAYMENT_SELECTOR)
                    scraped_data['goods_loan_amount'] = _get_text(page, GOODS_LOAN_AMOUNT_SELECTOR)
                    
                    callback(f"         Tổng tiền: {scraped_data['goods_total_amount']}")
                    callback(f"         Trả trước: {scraped_data['goods_down_payment']}")
                    callback(f"         Số tiền vay: {scraped_data['goods_loan_amount']}")

                    # 4. Thông tin Gói vay - CẢI TIẾN
                    callback("         Đang lấy thông tin gói vay...")
                    
                    # Thử chờ element xuất hiện
                    try:
                        page.wait_for_selector(GOODS_INSTALLMENT_AMOUNT_SELECTOR, timeout=8000)
                    except:
                        callback("         ⚠ Không tìm thấy thông tin 'Góp mỗi tháng', thử selector khác...")
                    
                    scraped_data['goods_installment_amount'] = _get_text(page, GOODS_INSTALLMENT_AMOUNT_SELECTOR)
                    scraped_data['goods_installment_months'] = _get_text(page, GOODS_INSTALLMENT_MONTHS_SELECTOR)
                    scraped_data['goods_interest_rate'] = _get_text(page, GOODS_INTEREST_RATE_SELECTOR)
                    scraped_data['goods_insurance'] = _get_text(page, GOODS_INSURANCE_SELECTOR)
                    scraped_data['goods_bonus_scheme'] = _get_text(page, GOODS_BONUS_SCHEME_SELECTOR)
                    
                    callback(f"         Góp mỗi tháng: {scraped_data['goods_installment_amount']}")
                    callback(f"         Số tháng: {scraped_data['goods_installment_months']}")
                    callback(f"         Lãi suất: {scraped_data['goods_interest_rate']}")
                    callback(f"         Bảo hiểm: {scraped_data['goods_insurance']}")
                    callback(f"         Bonus Scheme: {scraped_data['goods_bonus_scheme']}")
                    
                except Exception as e:
                    callback(f"      Lỗi khi cào Tab 'Thông tin hàng hóa': {e}")
                    import traceback
                    callback(f"      Chi tiết lỗi: {traceback.format_exc()}")
                
                # Thêm dữ liệu đã cào vào danh sách tổng
                all_scraped_data.append(scraped_data)
                callback(f"      ✔ Đã cào xong HĐ: {ed_id}")
                
                # Random delay giữa 1.5-3.5 giây để tránh spam
                delay = random.uniform(1.5, 3.5)
                callback(f"      ⏳ Chờ {delay:.1f}s trước khi xử lý HĐ tiếp theo...")
                time.sleep(delay)
                
            except Exception as e:
                callback(f"   *** Lỗi khi xử lý hợp đồng ID {ed_id}: {e}")
                callback(f"   *** Bỏ qua và tiếp tục với HĐ tiếp theo...")
                continue 

        # === GIAI ĐOẠN 3: XUẤT EXCEL ===
        if stop_event.is_set():
            callback("\n===================================")
            callback(f"🛑 KỊCH BẢN ĐÃ DỪNG THEO YÊU CẦU CỦA NGƯỜI DÙNG.")
        else:
            callback("\n===================================")
            callback(f"🎉 HOÀN TẤT! ĐÃ CÀO XONG {len(all_scraped_data)} HỢP ĐỒNG!")
            callback("Đang bắt đầu xuất ra file Excel...")
            # Sửa: dùng final_save_dir thay vì save_directory
            _export_details_to_excel(all_scraped_data, final_save_dir, start_date_ddmmyyyy, end_date_ddmmyyyy)
            callback("===================================")
        
        close_delay = random.uniform(8, 12)
        callback(f"\n(Trình duyệt sẽ tự đóng sau {close_delay:.1f} giây...)")
        time.sleep(close_delay)

    except InterruptedError: # Bắt lỗi chủ động dừng
        callback("\n===================================")
        callback(f"🛑 KỊCH BẢN ĐÃ DỪNG THEO YÊU CẦU CỦA NGƯỜI DÙNG.")
    
    except Exception as e:
        callback("\n===================================")
        callback(f"❌ ĐÃ XAY RA LỖI LỚN TRONG GIAI ĐOẠN 2 (CÀO DỮ LIỆU)! Lỗi: {e}")
    
    finally:
        if browser:
            browser.close()
            playwright_instance.stop()
        callback("Đã đóng trình duyệt.")

