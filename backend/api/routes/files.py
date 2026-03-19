"""
Files API Routes
API endpoints cho upload/download files
"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from typing import List
import os
import json
import io
import aiofiles
import re
import unicodedata
from datetime import datetime

from config.settings import DOWNLOADS_DIR, APP_DATA_DIR, ALLOWED_EXTENSIONS, MAX_UPLOAD_SIZE

router = APIRouter()

# Thư mục temp cho upload
UPLOAD_DIR = APP_DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload file Excel hoặc PDF
    Trả về đường dẫn file và dữ liệu nếu là Excel
    """
    # Kiểm tra extension
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"File type not allowed. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Kiểm tra size
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Max size: {MAX_UPLOAD_SIZE / 1024 / 1024}MB"
        )
    
    # Lưu file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{file.filename}"
    filepath = UPLOAD_DIR / filename
    
    async with aiofiles.open(filepath, 'wb') as f:
        await f.write(contents)
    
    result = {
        "status": "success",
        "filename": filename,
        "original_name": file.filename,
        "filepath": str(filepath),
        "size": len(contents)
    }
    
    # Nếu là Excel, đọc và trả về data
    if ext in [".xlsx", ".xls"]:
        try:
            data = parse_excel_file(str(filepath))
            result["data"] = data
            result["row_count"] = len(data)
        except Exception as e:
            result["parse_error"] = str(e)
    
    return result


# Mapping tên cột Excel → key chuẩn cho Zalo customer
_COLUMN_ALIAS: dict = {
    # phone
    'sđt (chính)': 'phone', 'sdt (chinh)': 'phone',
    'sđt': 'phone', 'sdt': 'phone',
    'số điện thoại': 'phone', 'so dien thoai': 'phone',
    'điện thoại': 'phone', 'dien thoai': 'phone',
    'phone': 'phone', 'phone number': 'phone',
    # name
    'tên kh (profile)': 'name', 'ten kh (profile)': 'name',
    'tên kh': 'name', 'ten kh': 'name',
    'họ tên': 'name', 'ho ten': 'name',
    'họ và tên': 'name', 'ho va ten': 'name',
    'tên': 'name', 'ten': 'name',
    'khách hàng': 'name', 'khach hang': 'name',
    'name': 'name',
    # contract_id
    'id hợp đồng': 'contract_id', 'id hop dong': 'contract_id',
    'số hợp đồng': 'contract_id', 'so hop dong': 'contract_id',
    'mã hđ': 'contract_id', 'ma hd': 'contract_id',
    'mã hợp đồng': 'contract_id', 'ma hop dong': 'contract_id',
    'hợp đồng': 'contract_id', 'hop dong': 'contract_id',
    'contract_id': 'contract_id', 'contract id': 'contract_id',
    # gender
    'giới tính': 'gender', 'gioi tinh': 'gender',
    'giới tính khách hàng': 'gender', 'gioi tinh khach hang': 'gender',
    'gender': 'gender', 'sex': 'gender', 'gt': 'gender',
    # address
    'địa chỉ': 'address', 'dia chi': 'address',
    'địa chỉ thường trú': 'address', 'dia chi thuong tru': 'address',
    'địa chỉ tạm trú': 'address', 'dia chi tam tru': 'address',
    'địa chỉ hiện tại': 'address', 'dia chi hien tai': 'address',
    'địa chỉ nhà': 'address', 'dia chi nha': 'address',
    'address': 'address',
    # cccd
    'số cccd': 'cccd', 'so cccd': 'cccd',
    'cccd': 'cccd',
    'số cmnd': 'cccd', 'so cmnd': 'cccd',
    'cmnd': 'cccd',
    'căn cước': 'cccd', 'can cuoc': 'cccd',
    'số căn cước': 'cccd', 'so can cuoc': 'cccd',
    # dob
    'ngày sinh': 'dob', 'ngay sinh': 'dob',
    'dob': 'dob', 'date of birth': 'dob',
    'ngay sinh khach hang': 'dob', 'ngày sinh khách hàng': 'dob',
}


def _normalize_header(h: str) -> str:
    """Chuẩn hoá header để so sánh (thường, bỏ khoảng trắng thừa)"""
    text = unicodedata.normalize('NFKD', h or '')
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace('đ', 'd').replace('Đ', 'D')
    text = text.lower().strip()
    text = re.sub(r'[:\-_]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text


def _map_customer_row(row_data: dict) -> dict:
    """
    Thêm các key chuẩn (phone, name, contract_id, address, cccd, dob, gender)
    dựa trên tên cột gốc.
    Giữ lại toàn bộ cột gốc để không mất dữ liệu.
    """
    result = dict(row_data)
    for orig_key, value in row_data.items():
        norm = _normalize_header(orig_key)
        mapped = _COLUMN_ALIAS.get(norm)
        if mapped and value and not result.get(mapped):
            result[mapped] = value
    return result


def parse_excel_file(filepath: str) -> List[dict]:
    """Parse file Excel thành list of dicts, tự động map cột sang key chuẩn"""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    data = []
    headers = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell).strip() if cell else f"col_{j}" for j, cell in enumerate(row)]
        else:
            row_data = {}
            for j, cell in enumerate(row):
                if j < len(headers):
                    row_data[headers[j]] = str(cell).strip() if cell else ""
            if any(row_data.values()):
                data.append(_map_customer_row(row_data))

    wb.close()
    return data


@router.get("/list")
async def list_files(directory: str = "downloads"):
    """Liệt kê các file trong thư mục"""
    if directory == "downloads":
        target_dir = DOWNLOADS_DIR
    elif directory == "uploads":
        target_dir = UPLOAD_DIR
    else:
        raise HTTPException(status_code=400, detail="Invalid directory")
    
    files = []
    if target_dir.exists():
        for f in target_dir.iterdir():
            if f.is_file():
                stat = f.stat()
                files.append({
                    "name": f.name,
                    "size": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "extension": f.suffix
                })
    
    # Sắp xếp theo thời gian mới nhất
    files.sort(key=lambda x: x["modified"], reverse=True)
    
    return {"files": files, "directory": str(target_dir)}


@router.get("/download/{filename}")
async def download_file(filename: str, directory: str = "downloads"):
    """Download file"""
    if directory == "downloads":
        target_dir = DOWNLOADS_DIR
    elif directory == "uploads":
        target_dir = UPLOAD_DIR
    else:
        raise HTTPException(status_code=400, detail="Invalid directory")
    
    filepath = target_dir / filename
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/octet-stream"
    )


@router.delete("/{filename}")
async def delete_file(filename: str, directory: str = "uploads"):
    """Xóa file"""
    if directory == "downloads":
        target_dir = DOWNLOADS_DIR
    elif directory == "uploads":
        target_dir = UPLOAD_DIR
    else:
        raise HTTPException(status_code=400, detail="Invalid directory")
    
    filepath = target_dir / filename
    
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        os.remove(filepath)
        return {"status": "success", "message": f"Deleted {filename}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/parse-excel")
async def parse_excel_endpoint(file: UploadFile = File(...)):
    """
    Parse file Excel thành JSON (không lưu file)
    Dùng cho việc preview data trước khi gửi
    """
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".xlsx", ".xls"]:
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large")
    
    # Lưu tạm để parse
    temp_path = UPLOAD_DIR / f"temp_{file.filename}"
    async with aiofiles.open(temp_path, 'wb') as f:
        await f.write(contents)
    
    try:
        data = parse_excel_file(str(temp_path))
        return {
            "status": "success",
            "filename": file.filename,
            "data": data,
            "row_count": len(data)
        }
    finally:
        # Xóa file tạm
        if temp_path.exists():
            os.remove(temp_path)


@router.get("/template")
async def download_template():
    """
    Tạo và tải về file Excel mẫu cùng định dạng với tab RPA (scrape details).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    headers = [
        'STT', 'ID Hợp đồng', 'Tên KH (Profile)', 'Giới tính', 'Ngày sinh', 'Số CCCD',
        'Ngày cấp', 'Ngày hết hạn', 'SĐT (Chính)', 'Tình trạng hôn nhân', 'Học vấn',
        'Nghề nghiệp', 'Tên công ty', 'Địa chỉ công ty', 'Thu nhập',
        'Địa chỉ thường trú', 'Địa chỉ tạm trú',
        'Tham chiếu 1: Tên', 'Tham chiếu 1: SĐT', 'Tham chiếu 1: Quan hệ',
        'Tham chiếu 2: Tên', 'Tham chiếu 2: SĐT', 'Tham chiếu 2: Quan hệ',
        'Ngày đóng tiền đầu tiên',
        'Mã POS', 'Tên POS', 'Địa chỉ POS', 'Username', 'Scheme',
        'Sản phẩm (Gộp)', 'Tổng tiền', 'Trả trước', 'Số tiền vay',
        'Góp mỗi tháng', 'Số tháng', 'Lãi suất', 'Bảo hiểm', 'Bonus Scheme',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiết Hợp đồng"

    ws.append(headers)

    header_font = Font(name='Times New Roman', size=11, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        col_letter = get_column_letter(i + 1)
        header_text = headers[i]
        width = max(len(header_text) + 2, 14)
        if 'Địa chỉ' in header_text or 'Tên POS' in header_text:
            width = 30
        if 'Sản phẩm' in header_text:
            width = 40
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''mau_danh_sach_khach_hang.xlsx"
        },
    )
